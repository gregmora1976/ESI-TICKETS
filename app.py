from flask import Flask, render_template, jsonify, request, send_file, abort, redirect, url_for
from pathlib import Path
import json, webbrowser, os, urllib.request, urllib.parse
import csv, re, time
import hashlib, threading
from io import StringIO, BytesIO
import smtplib
from email.mime.text import MIMEText
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from datetime import datetime
import html

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / 'data'
CONFIG_FILE = DATA_DIR / 'config.json'
TICKETS_SUB = 'tickets'
FILES_SUB = 'fichiers'
SUPABASE_URL = (os.getenv("SUPABASE_URL") or "").rstrip("/")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or ""
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "uploads")

app = Flask(__name__, template_folder='templates', static_folder='static')

# Cache court + verrou pour éviter de relancer plusieurs OCR lourds sur le même PDF.
_RECEPTION_OCR_LOCK = threading.Lock()
_RECEPTION_PDF_CACHE = {}
_RECEPTION_PDF_CACHE_TTL = 300  # 5 minutes
_RECEPTION_PDF_CACHE_MAX = 20


def _reception_cache_get(pdf_bytes):
    key = hashlib.sha256(pdf_bytes).hexdigest()
    item = _RECEPTION_PDF_CACHE.get(key)
    if not item:
        return key, None
    if time.time() - item.get('ts', 0) > _RECEPTION_PDF_CACHE_TTL:
        _RECEPTION_PDF_CACHE.pop(key, None)
        return key, None
    return key, dict(item.get('parsed') or {})


def _reception_cache_set(key, parsed):
    if len(_RECEPTION_PDF_CACHE) >= _RECEPTION_PDF_CACHE_MAX:
        oldest = min(_RECEPTION_PDF_CACHE.items(), key=lambda kv: kv[1].get('ts', 0))[0]
        _RECEPTION_PDF_CACHE.pop(oldest, None)
    _RECEPTION_PDF_CACHE[key] = {'ts': time.time(), 'parsed': dict(parsed)}


# -----------------------------------------------------------------------------
# Google Sheet public - suivi fournisseur caisserie
# -----------------------------------------------------------------------------
GOOGLE_SHEET_PUBLIC_ID = os.getenv(
    "GOOGLE_SHEET_PUBLIC_ID",
    "2PACX-1vQSiTSLN-AtXoa4GrscgSM_2VwFzO12Bh-UFyUKNLihyRZSocciqe8OHHIZCKvs5r77ynFqd5NZI29Q"
)
GOOGLE_SHEET_PUBHTML_URL = (
    f"https://docs.google.com/spreadsheets/d/e/{GOOGLE_SHEET_PUBLIC_ID}/pubhtml"
)
GOOGLE_SHEET_CSV_BASE_URL = (
    f"https://docs.google.com/spreadsheets/d/e/{GOOGLE_SHEET_PUBLIC_ID}/pub"
)

# Petit cache mémoire pour éviter de relire la page des onglets à chaque clic.
_GOOGLE_GIDS_CACHE = {"gids": [], "expires_at": 0}


def _http_get_text(url, timeout=10):
    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "User-Agent": "Mozilla/5.0 ESI-Tickets/1.0",
            "Accept": "text/html,text/csv,*/*"
        }
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
        charset = resp.headers.get_content_charset() or "utf-8"
        return raw.decode(charset, errors="replace")


def _google_sheet_gids(force_refresh=False):
    now = time.time()
    if (
        not force_refresh
        and _GOOGLE_GIDS_CACHE["gids"]
        and now < _GOOGLE_GIDS_CACHE["expires_at"]
    ):
        return list(_GOOGLE_GIDS_CACHE["gids"])

    html_page = _http_get_text(GOOGLE_SHEET_PUBHTML_URL, timeout=10)
    gids = re.findall(r"gid=(\d+)", html_page)

    # Supprime les doublons tout en conservant l'ordre du classeur publié.
    gids = list(dict.fromkeys(gids))

    _GOOGLE_GIDS_CACHE["gids"] = gids
    _GOOGLE_GIDS_CACHE["expires_at"] = now + 600  # 10 minutes
    return list(gids)


def get_caisse_fournisseur_status(caisse_ref):
    """
    Recherche une caisse dans tous les onglets publiés du Google Sheet.

    Exemple : 100872-01
      - colonne B = 100872
      - colonne C = 1
      - colonne D = état fournisseur
    """
    caisse_ref = _as_text(caisse_ref).strip()
    if not caisse_ref or "-" not in caisse_ref:
        return {
            "success": False,
            "caisse": caisse_ref,
            "error": "Format de caisse invalide"
        }

    try:
        reference_recherchee, numero_recherche = caisse_ref.rsplit("-", 1)
        reference_recherchee = reference_recherchee.strip()
        numero_recherche = str(int(numero_recherche.strip()))
    except Exception:
        return {
            "success": False,
            "caisse": caisse_ref,
            "error": "Format de caisse invalide"
        }

    try:
        gids = _google_sheet_gids()
    except Exception as e:
        print(f"[GOOGLE SHEET] Impossible de récupérer les onglets : {e}")
        return {
            "success": False,
            "caisse": caisse_ref,
            "error": "Suivi fournisseur indisponible"
        }

    if not gids:
        return {
            "success": False,
            "caisse": caisse_ref,
            "error": "Aucun onglet fournisseur disponible"
        }

    for gid in gids:
        csv_url = (
            f"{GOOGLE_SHEET_CSV_BASE_URL}?gid={urllib.parse.quote(str(gid), safe='')}"
            "&single=true&output=csv"
        )

        try:
            csv_text = _http_get_text(csv_url, timeout=10)
        except Exception as e:
            print(f"[GOOGLE SHEET] Erreur lecture gid={gid} : {e}")
            continue

        reader = csv.reader(StringIO(csv_text))
        for row in reader:
            if len(row) < 4:
                continue

            reference = _as_text(row[1]).strip()
            numero = _as_text(row[2]).strip()
            etat = _as_text(row[3]).strip()

            if not reference or not numero:
                continue

            try:
                numero_normalise = str(int(float(numero.replace(",", "."))))
            except Exception:
                continue

            if (
                reference == reference_recherchee
                and numero_normalise == numero_recherche
            ):
                return {
                    "success": True,
                    "caisse": caisse_ref,
                    "etat": etat or "Non renseigné",
                    "gid": str(gid)
                }

    return {
        "success": False,
        "caisse": caisse_ref,
        "error": "Caisse introuvable dans le suivi fournisseur"
    }

def safe_filename(name):
    """Nettoie le nom du fichier pour Supabase tout en gardant le vrai nom affiché côté appli."""
    name = str(name or "fichier")
    return "".join(
        c if c.isalnum() or c in "._-" else "_"
        for c in name
    )


def supabase_upload_bytes(storage_path, content, content_type="application/octet-stream"):
    """Envoie un fichier dans Supabase Storage sans dépendre du SDK Python."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{safe_path}"

    req = urllib.request.Request(
        url,
        data=content,
        method="POST",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY,
            "Content-Type": content_type or "application/octet-stream",
            "x-upsert": "true"
        }
    )

    print("[SUPABASE UPLOAD URL]", url)
    print("[SUPABASE UPLOAD BUCKET]", SUPABASE_BUCKET)
    print("[SUPABASE UPLOAD PATH]", safe_path)

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        print(f"[SUPABASE UPLOAD ERROR] HTTP {e.code} - {e.reason} - {body}")
        raise

def supabase_download_bytes(storage_path):
    """Télécharge un fichier depuis Supabase Storage."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{safe_path}"

    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY
        }
    )

    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def supabase_signed_download_url(storage_path, expires_in=300):
    """Crée une URL signée Supabase Storage pour éviter de faire transiter le fichier par Render."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/sign/{SUPABASE_BUCKET}/{safe_path}"

    payload = json.dumps({"expiresIn": int(expires_in)}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY,
            "Content-Type": "application/json"
        }
    )

    with urllib.request.urlopen(req, timeout=30) as resp:
        body = json.loads(resp.read().decode("utf-8", errors="replace"))

    signed = body.get("signedURL") or body.get("signedUrl") or body.get("url")
    if not signed:
        raise RuntimeError(f"Réponse URL signée invalide : {body}")
    if signed.startswith("http"):
        return signed
    return SUPABASE_URL + "/storage/v1" + signed

def choose_shared_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askdirectory(title="Choisis le dossier partagé ESI Tickets")
        root.destroy()
        if path:
            return path
    except Exception:
        pass
    return ''

def load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {}

def save_config(cfg):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding='utf-8')

def ensure_shared_root():
    root = APP_DIR
    (root / TICKETS_SUB).mkdir(parents=True, exist_ok=True)
    (root / FILES_SUB).mkdir(parents=True, exist_ok=True)
    return root

def tickets_dir():
    return ensure_shared_root() / TICKETS_SUB

def files_dir():
    return ensure_shared_root() / FILES_SUB

def ticket_file(ticket_id):
    return tickets_dir() / f'{ticket_id}.json'

def ticket_folder(ticket_id):
    path = files_dir() / ticket_id
    path.mkdir(parents=True, exist_ok=True)
    return path



def _as_text(value, default=''):
    if value is None:
        return default
    return str(value)


# -----------------------------------------------------------------------------
# Référentiel global des articles ESI
# -----------------------------------------------------------------------------
_ARTICLE_LOCK = threading.Lock()


def _article_quantity(value, default=1):
    try:
        n = int(float(str(value or default).replace(",", ".")))
        return max(1, n)
    except Exception:
        return default


def _article_search_text(article):
    """Concatène toutes les données connues afin qu'une seule recherche retrouve l'article."""
    values = []
    for key, value in (article or {}).items():
        if key == "raw_json":
            try:
                values.append(json.dumps(value, ensure_ascii=False, sort_keys=True))
            except Exception:
                values.append(_as_text(value))
        elif isinstance(value, (dict, list)):
            try:
                values.append(json.dumps(value, ensure_ascii=False, sort_keys=True))
            except Exception:
                values.append(_as_text(value))
        else:
            values.append(_as_text(value))
    return " | ".join(x.strip() for x in values if _as_text(x).strip())


def _article_payload_from_item(ticket, item, source_module=None, source_index=None, unit_index=1):
    """Transforme une ligne de marchandise en article physique (quantité = 1)."""
    item = dict(item or {})
    module = source_module or ticket.get("module") or ""
    avis = ticket.get("avisArrivee") or ticket.get("avis_arrivee") or {}
    enl = ticket.get("enlevement") or {}

    if module == "Avis d'arrivée":
        dossier = avis.get("dossier_ref") or ticket.get("dossier") or ""
        client = avis.get("client") or ticket.get("dossier") or ""
        projet = avis.get("projet") or ticket.get("expo") or ticket.get("objet") or ""
        reference = item.get("reference") or ""
        description = item.get("description") or ""
        poids = item.get("poids_kg") or ""
        longueur = item.get("longueur_cm") or ""
        largeur = item.get("largeur_cm") or ""
        hauteur = item.get("hauteur_cm") or ""
        volume = item.get("volume_m3") or ""
        surface = item.get("surface_m2") or ""
        ref_caisse = item.get("ref_caisse") or ""
        transporteur_ref = ((avis.get("transporteur") or {}).get("reference") or "")
    else:
        dossier = ticket.get("numeroDossier") or ticket.get("numero_dossier") or ""
        client = enl.get("client") or ticket.get("dossier") or ""
        projet = enl.get("exhibition") or ticket.get("expo") or ticket.get("objet") or ""
        reference = item.get("reference") or item.get("ref") or ""
        description = item.get("designation") or item.get("description") or ""
        poids = item.get("poids_kg") or item.get("poids") or ""
        dims = _as_text(item.get("dimensions")).strip()
        longueur = item.get("longueur_cm") or ""
        largeur = item.get("largeur_cm") or ""
        hauteur = item.get("hauteur_cm") or ""
        if dims and not any((longueur, largeur, hauteur)):
            parts = [x.strip() for x in re.split(r"[xX×]", dims)]
            if len(parts) >= 1: longueur = parts[0]
            if len(parts) >= 2: largeur = parts[1]
            if len(parts) >= 3: hauteur = parts[2]
        volume = item.get("volume_m3") or ""
        surface = item.get("surface_m2") or ""
        ref_caisse = item.get("ref_caisse") or ""
        transporteur_ref = enl.get("numero_bon") or ticket.get("ref") or ""

    created_at = ticket.get("createdAt") or datetime.now().isoformat()
    payload = {
        "ticket_id": _as_text(ticket.get("id")).strip(),
        "source_module": _as_text(module).strip(),
        "source_index": int(source_index) if source_index is not None else None,
        "unit_index": int(unit_index),
        "reference": _as_text(reference).strip(),
        "description": _as_text(description).strip(),
        "dossier": _as_text(dossier).strip(),
        "client": _as_text(client).strip(),
        "projet": _as_text(projet).strip(),
        "ref_caisse": _as_text(ref_caisse).strip(),
        "transporteur_ref": _as_text(transporteur_ref).strip(),
        "longueur_cm": _as_text(longueur).strip(),
        "largeur_cm": _as_text(largeur).strip(),
        "hauteur_cm": _as_text(hauteur).strip(),
        "volume_m3": _as_text(volume).strip(),
        "surface_m2": _as_text(surface).strip(),
        "poids_kg": _as_text(poids).strip(),
        "lieu_stockage": _as_text(item.get("lieu_stockage")).strip(),
        "statut_logistique": "Créé",
        "created_at": _as_text(created_at).strip(),
        "updated_at": datetime.now().isoformat(),
        "raw_json": {
            "ticket_id": ticket.get("id"),
            "module": module,
            "source_index": source_index,
            "unit_index": unit_index,
            "item": item,
        },
    }
    payload["search_text"] = _article_search_text(payload)
    return payload


def _article_row_to_public(row):
    row = dict(row or {})
    article_no = row.get("article_no")
    if not row.get("esi_id") and article_no is not None:
        row["esi_id"] = f"ESI-{article_no}"
    return row


def _create_article_record(payload):
    """Insère un article et retourne sa ligne avec son identifiant ESI."""
    rows = supabase_rest_request(
        "POST",
        "articles",
        "",
        [payload],
        prefer="return=representation"
    ) or []
    if not rows:
        raise RuntimeError("Supabase n'a pas retourné l'article créé.")
    return _article_row_to_public(rows[0])


def _normalise_article_reference(value):
    """Normalise une référence uniquement pour le rapprochement bon d'enlèvement / base articles."""
    return re.sub(r"[^A-Za-z0-9]", "", _as_text(value)).upper().strip()


def _find_existing_articles_for_reference(reference, limit=25):
    """Retourne les articles existants dont la référence correspond exactement après normalisation."""
    reference = _as_text(reference).strip()
    if not reference:
        return []

    # Le filtre ilike limite la quantité de données lues, puis la comparaison normalisée
    # évite de rater ABC-01 / ABC 01 ou une différence de casse issue de l'OCR.
    wanted = _normalise_article_reference(reference)
    token = wanted[:4] or reference.replace('*', '').strip()
    pattern = '*' + token + '*'
    rows = supabase_rest_request(
        'GET', 'articles',
        'select=*&reference=ilike.' + urllib.parse.quote(pattern, safe='*') +
        '&order=article_no.desc&limit=' + str(int(limit))
    ) or []

    matches = []
    for row in rows:
        if _normalise_article_reference(row.get('reference')) != wanted:
            continue
        article = _article_row_to_public(row)
        matches.append({
            'esi_id': _as_text(article.get('esi_id')).strip(),
            'reference': _as_text(article.get('reference')).strip(),
            'description': _as_text(article.get('description')).strip(),
            'dossier': _as_text(article.get('dossier')).strip(),
            'client': _as_text(article.get('client')).strip(),
            'projet': _as_text(article.get('projet')).strip(),
            'lieu_stockage': _as_text(article.get('lieu_stockage')).strip(),
            'statut_logistique': _as_text(article.get('statut_logistique')).strip(),
        })
    return matches


def _enlevement_with_article_candidates(parsed):
    """Ajoute les propositions de la base articles aux lignes reconnues par l'OCR."""
    result = dict(parsed or {})
    enriched = []
    for index, source in enumerate(result.get('items') or []):
        item = dict(source or {})
        item['source_index'] = index
        item['article_candidates'] = _find_existing_articles_for_reference(item.get('reference'))
        enriched.append(item)
    result['items'] = enriched
    return result


def _apply_enlevement_article_selections(ticket, selections):
    """
    Applique les ESI existants choisis par le demandeur.
    Les unités marquées 'create' restent sans ESI et seront créées ensuite par
    _ensure_articles_for_ticket().
    """
    enl = dict(ticket.get('enlevement') or {})
    items = list(enl.get('items') or [])
    by_index = {}
    for entry in selections or []:
        if not isinstance(entry, dict):
            continue
        try:
            idx = int(entry.get('index'))
        except Exception:
            continue
        by_index[idx] = entry

    used_esi = set()
    for idx, original in enumerate(items):
        item = dict(original or {})
        qty = _article_quantity(item.get('quantite'), 1)
        selection = by_index.get(idx) or {}
        units = selection.get('units') or []
        chosen = []

        for unit in units[:qty]:
            if not isinstance(unit, dict) or unit.get('mode') != 'existing':
                continue
            esi_id = _as_text(unit.get('esi_id')).strip()
            if not esi_id or esi_id in used_esi:
                continue

            safe_esi = urllib.parse.quote(esi_id, safe='-')
            rows = supabase_rest_request(
                'GET', 'articles', f'select=esi_id,reference&esi_id=eq.{safe_esi}&limit=1'
            ) or []
            if not rows:
                continue
            if _normalise_article_reference(rows[0].get('reference')) != _normalise_article_reference(item.get('reference')):
                continue

            chosen.append(esi_id)
            used_esi.add(esi_id)

        item['esi_ids'] = chosen
        item['esi_id'] = chosen[0] if chosen else ''
        # Les candidats n'ont pas vocation à être stockés dans le ticket final.
        item.pop('article_candidates', None)
        item.pop('source_index', None)
        items[idx] = item

    enl['items'] = items
    enl['references'] = [
        _as_text(x.get('reference')).strip() for x in items
        if _as_text(x.get('reference')).strip()
    ]
    ticket['enlevement'] = enl


def _extract_enlevement_pdf_preview_low_memory(pdf_bytes):
    """
    Analyse légère utilisée uniquement avant la création du ticket.

    Le flux historique faisait deux OCR complets (300 dpi puis OCR spatial 200 dpi),
    ce qui peut dépasser la mémoire/timeout d'un worker Render. Ici on :
      - tente d'abord le texte natif ;
      - OCRise page par page à 180 dpi si nécessaire ;
      - ne lance PAS le second OCR spatial ;
      - extrait tout de même les références et les principaux champs texte.
    """
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise RuntimeError("Le module pypdf n'est pas installé.") from e

    try:
        reader = PdfReader(BytesIO(pdf_bytes))
    except Exception as e:
        raise ValueError(f"PDF illisible : {e}")

    pages_native = []
    for page in reader.pages:
        try:
            pages_native.append(page.extract_text() or "")
        except Exception:
            pages_native.append("")

    native_text = "\n".join(pages_native).strip()
    text = native_text
    ocr_used = False

    if len(native_text) < 120:
        print("[ENLEVEMENT PREVIEW] Texte natif insuffisant, OCR basse mémoire")
        try:
            from pdf2image import convert_from_bytes
            import pytesseract
        except Exception as e:
            raise RuntimeError(
                "OCR indisponible. Vérifie pdf2image, pytesseract, Pillow, tesseract-ocr et poppler-utils."
            ) from e

        ocr_pages = []
        page_count = len(reader.pages)
        with _RECEPTION_OCR_LOCK:
            for page_no in range(1, page_count + 1):
                images = []
                try:
                    images = convert_from_bytes(
                        pdf_bytes,
                        dpi=180,
                        grayscale=True,
                        first_page=page_no,
                        last_page=page_no,
                        thread_count=1,
                    )
                    if not images:
                        ocr_pages.append("")
                        continue
                    image = images[0]
                    page_text = pytesseract.image_to_string(
                        image, lang="fra", config="--psm 4"
                    )
                    if len((page_text or "").strip()) < 80:
                        page_text = pytesseract.image_to_string(
                            image, lang="fra", config="--psm 11"
                        )
                    ocr_pages.append(page_text or "")
                    print(f"[ENLEVEMENT PREVIEW] Page {page_no}/{page_count} OCRisée")
                finally:
                    for image in images:
                        try:
                            image.close()
                        except Exception:
                            pass
                    images.clear()
        text = "\n".join(ocr_pages).strip()
        ocr_used = True

    if not text:
        raise ValueError("Aucun texte exploitable trouvé dans le bon d'enlèvement.")

    clean_text = text.replace("\r", "")
    cutoff = re.search(r"\bAssur[eé]\s+par\b", clean_text, re.I)
    if cutoff:
        clean_text = clean_text[:cutoff.start()].rstrip()

    lines = _enlevement_lines(clean_text)
    label_numero = [
        r"Num[eé]ro\s+de\s+r[eé]f[eé]r(?:ence)?",
        r"N[°ºo]\s*de\s*r[eé]f[eé]rence",
        r"R[eé]f[eé]rence\s+du\s+bon",
    ]
    common_stops = [
        r"Client", r"Coordinateur", r"Exhibition", r"Programme\s+du\s+chantier",
        r"Instructions?", r"Adresse", r"Service"
    ]

    numero_bon = _value_after_label(lines, label_numero, common_stops)
    if numero_bon:
        m = re.search(r"\b([A-Za-z0-9][A-Za-z0-9_-]{3,})\b", numero_bon)
        numero_bon = m.group(1) if m else ""

    client = _value_after_label(
        lines, [r"\bClient\b"],
        [r"Coordinateur", r"Exhibition", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )
    coordinateur = _value_after_label(
        lines, [r"Coordinateur"],
        [r"Client", r"Exhibition", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )
    exhibition = _value_after_label(
        lines, [r"Exhibition"],
        [r"Client", r"Coordinateur", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )

    if client:
        client = re.split(r"\bExhibition\b", client, maxsplit=1, flags=re.I)[0].strip(" :-|")
    if coordinateur:
        coordinateur = re.split(r"\b(?:Client|Exhibition)\b", coordinateur, maxsplit=1, flags=re.I)[0].strip(" :-|")
    if exhibition:
        exhibition = re.split(r"\b(?:Client|Coordinateur)\b", exhibition, maxsplit=1, flags=re.I)[0].strip(" :-|")

    programme = _extract_programme_chantier(clean_text)
    instructions = _extract_instructions_block(clean_text)
    items = _extract_enlevement_items(instructions) if instructions else []
    contact_data = _extract_contact_blocks(clean_text)

    display_name = " - ".join(
        x for x in [_clean_ocr_line(client), _clean_ocr_line(numero_bon)] if x
    )

    return {
        "numero_bon": numero_bon,
        "client": client,
        "display_name": display_name,
        "coordinateur": coordinateur,
        "exhibition": exhibition,
        "date_enlevement": programme.get("date_enlevement", ""),
        "service": programme.get("service", ""),
        "assigne_a": programme.get("assigne_a", ""),
        "vehicules": programme.get("vehicules", ""),
        "notes": programme.get("notes", ""),
        "instructions": instructions,
        "items": items,
        "references": [x.get("reference") for x in items if x.get("reference")],
        "page_count": len(reader.pages),
        "ocr_used": ocr_used,
        "raw_text": clean_text,
        **contact_data,
    }


@app.route('/api/enlevement/analyser-articles', methods=['POST'])
def api_enlevement_analyser_articles():
    """Analyse légère d'un bon d'enlèvement et propose les articles existants."""
    fs = request.files.get('file')
    if not fs or not fs.filename:
        return jsonify({'ok': False, 'error': "Bon d'enlèvement PDF manquant"}), 400
    if not fs.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'error': "Le bon d'enlèvement doit être un PDF"}), 400
    content = fs.read()
    if not content:
        return jsonify({'ok': False, 'error': 'Le fichier PDF est vide'}), 400

    try:
        parsed = _extract_enlevement_pdf_preview_low_memory(content)
        enriched = _enlevement_with_article_candidates(parsed)
        return jsonify({
            'ok': True,
            'analysis': enriched,
            'items': enriched.get('items') or [],
            'existing_count': sum(len(x.get('article_candidates') or []) for x in enriched.get('items') or []),
        })
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    except Exception as e:
        print(f'[ENLEVEMENT ARTICLES] Erreur analyse : {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500


def _ensure_articles_for_ticket(ticket, save=True):
    """
    Attribue des ESI-x à toutes les unités des lignes de marchandise d'un ticket.
    Idempotent : une ligne possédant déjà ses esi_ids n'est pas recréée.
    """
    module = _as_text(ticket.get("module")).replace("’", "'").strip()
    if module == "Avis d'arrivée":
        container = dict(ticket.get("avisArrivee") or ticket.get("avis_arrivee") or {})
        items = list(container.get("items") or [])
        container_key = "avisArrivee"
    elif module in ("Demande d'enlèvement", "Demande d'enlevement"):
        container = dict(ticket.get("enlevement") or {})
        items = list(container.get("items") or [])
        container_key = "enlevement"
    else:
        return []

    created = []
    changed = False

    with _ARTICLE_LOCK:
        for index, original in enumerate(items):
            item = dict(original or {})
            qty = _article_quantity(item.get("quantite"), 1)
            existing = [str(x).strip() for x in (item.get("esi_ids") or []) if str(x).strip()]

            # Si l'ancienne structure stocke un seul identifiant.
            if not existing and item.get("esi_id"):
                existing = [_as_text(item.get("esi_id")).strip()]

            while len(existing) < qty:
                unit_index = len(existing) + 1
                payload = _article_payload_from_item(
                    ticket, item, source_module=module, source_index=index, unit_index=unit_index
                )
                article = _create_article_record(payload)
                existing.append(article["esi_id"])
                created.append(article)

            if item.get("esi_ids") != existing:
                item["esi_ids"] = existing
                item["esi_id"] = existing[0] if existing else ""
                items[index] = item
                changed = True

        if changed:
            container["items"] = items
            ticket[container_key] = container
            ticket["updatedAt"] = datetime.now().isoformat()
            if save:
                save_ticket(ticket)

    return created


def _update_article_logistics(esi_ids, lieu_stockage="", statut_logistique="Réceptionné",
                              colis=None, colis_by_esi=None, reception_ref="", receptionne_par=""):
    """Met à jour la fiche globale des articles après une réception, avec un colis précis par ESI."""
    colis_by_esi = dict(colis_by_esi or {})
    fallback = list(colis or [])
    for esi_id in esi_ids or []:
        esi_id = _as_text(esi_id).strip()
        if not esi_id:
            continue
        article_colis = _as_text(colis_by_esi.get(esi_id)).strip()
        article_colis_list = [article_colis] if article_colis else fallback
        safe_esi = urllib.parse.quote(esi_id, safe='-')
        rows = supabase_rest_request("GET", "articles", f"select=*&esi_id=eq.{safe_esi}&limit=1") or []
        if not rows:
            continue
        current = dict(rows[0])
        raw = current.get("raw_json") if isinstance(current.get("raw_json"), dict) else {}
        raw = dict(raw or {})
        history = list(raw.get("receptions") or [])
        history.append({"date": datetime.now().isoformat(), "lieu_stockage": lieu_stockage,
                        "colis": article_colis_list, "reception_ref": reception_ref,
                        "receptionne_par": receptionne_par})
        raw["receptions"] = history
        raw["colis_actuel"] = article_colis or (article_colis_list[0] if len(article_colis_list)==1 else "")
        patch = {"lieu_stockage": _as_text(lieu_stockage).strip(),
                 "statut_logistique": _as_text(statut_logistique).strip(),
                 "dernier_colis": article_colis or ", ".join(article_colis_list),
                 "derniere_reception_ref": _as_text(reception_ref).strip(),
                 "updated_at": datetime.now().isoformat(), "raw_json": raw}
        merged = dict(current); merged.update(patch); patch["search_text"] = _article_search_text(merged)
        supabase_rest_request("PATCH", "articles", f"esi_id=eq.{safe_esi}", patch, prefer="return=minimal")


def _article_ids_for_received_units(item, previous_qty, qty_received):
    """Retourne les ESI-x correspondant précisément aux unités reçues dans cette opération."""
    ids = [str(x).strip() for x in (item.get("esi_ids") or []) if str(x).strip()]
    start = max(0, int(previous_qty))
    end = start + max(0, int(qty_received))
    return ids[start:end]


ARTICLES_MANUAL_CREATE_JS = r"""(function(){
'use strict';

function escManual(v){return String(v??'').replace(/[&<>\"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#039;'}[c]))}

function ensureManualCreateUI(){
  const actions=document.querySelector('.actions');
  if(!actions || document.getElementById('manualArticleCreateBtn')) return;

  const btn=document.createElement('button');
  btn.className='btn primary';
  btn.id='manualArticleCreateBtn';
  btn.type='button';
  btn.textContent='+ Créer un article';
  const excelBtn=document.getElementById('excelImportBtn');
  if(excelBtn) actions.insertBefore(btn, excelBtn); else actions.appendChild(btn);

  const bg=document.createElement('div');
  bg.className='modal-backdrop';
  bg.id='manualArticleModalBackdrop';
  bg.setAttribute('aria-hidden','true');
  bg.innerHTML=`<div class="modal" role="dialog" aria-modal="true" style="width:min(940px,96vw)">
    <div class="modal-head">
      <div>
        <div class="modal-title">Créer un article manuellement</div>
        <div class="modal-sub">Le N° ESI sera généré automatiquement.</div>
      </div>
      <button class="modal-close" id="manualArticleModalClose" type="button">×</button>
    </div>
    <div class="modal-body">
      <div class="edit-grid" id="manualArticleGrid">
        <div class="edit-field"><label>Type</label><select id="manualType"><option value="PRODUIT">PRODUIT</option><option value="CONTENANT">CONTENANT</option></select></div>
        <div class="edit-field"><label>N° dossier *</label><input id="manualDossier" autocomplete="off" placeholder="Ex. 101129"></div>
        <div class="edit-field"><label>Référence / N° inventaire</label><input id="manualReference" autocomplete="off"></div>
        <div class="edit-field"><label>Client</label><input id="manualClient" autocomplete="off"></div>
        <div class="edit-field"><label>Projet / exposition</label><input id="manualProjet" autocomplete="off"></div>
        <div class="edit-field"><label>Réf. caisse</label><input id="manualRefCaisse" autocomplete="off" placeholder="Ex. 101129-01"></div>
        <div class="edit-field"><label>Réf. transporteur</label><input id="manualTransporteurRef" autocomplete="off"></div>
        <div class="edit-field"><label>Lieu de stockage</label><input id="manualLieuStockage" autocomplete="off"></div>
        <div class="edit-field" style="grid-column:1/-1"><label>Description / désignation</label><textarea id="manualDescription"></textarea></div>
        <div class="edit-field"><label>Longueur (cm)</label><input id="manualLongueur" inputmode="decimal"></div>
        <div class="edit-field"><label>Largeur (cm)</label><input id="manualLargeur" inputmode="decimal"></div>
        <div class="edit-field"><label>Hauteur (cm)</label><input id="manualHauteur" inputmode="decimal"></div>
        <div class="edit-field"><label>Poids (kg)</label><input id="manualPoids" inputmode="decimal"></div>
        <div class="edit-field"><label>Volume (m³)</label><input id="manualVolume" inputmode="decimal"></div>
        <div class="edit-field"><label>Surface (m²)</label><input id="manualSurface" inputmode="decimal"></div>
        <div class="edit-field"><label>Statut logistique</label><input id="manualStatut" value="Créé"></div>
      </div>
      <div id="manualArticleHint" class="muted" style="margin-top:12px;font-size:12px">Si ce N° de dossier existe déjà, Client et Projet seront repris automatiquement depuis la base.</div>
      <div class="modal-actions">
        <button class="btn" id="manualArticleCancel" type="button">Annuler</button>
        <button class="btn primary" id="manualArticleSave" type="button">Créer l'article</button>
      </div>
    </div>
  </div>`;
  document.body.appendChild(bg);

  const ids=['manualType','manualDossier','manualReference','manualClient','manualProjet','manualRefCaisse','manualTransporteurRef','manualLieuStockage','manualDescription','manualLongueur','manualLargeur','manualHauteur','manualPoids','manualVolume','manualSurface','manualStatut'];
  function el(id){return document.getElementById(id)}
  function clearForm(){
    ids.forEach(id=>{const node=el(id); if(!node)return; if(id==='manualType')node.value='PRODUIT'; else if(id==='manualStatut')node.value='Créé'; else node.value='';});
    el('manualArticleHint').textContent='Si ce N° de dossier existe déjà, Client et Projet seront repris automatiquement depuis la base.';
  }
  function openModal(){clearForm();bg.classList.add('open');bg.setAttribute('aria-hidden','false');setTimeout(()=>el('manualDossier').focus(),50)}
  function closeModal(){bg.classList.remove('open');bg.setAttribute('aria-hidden','true')}

  btn.onclick=openModal;
  el('manualArticleModalClose').onclick=closeModal;
  el('manualArticleCancel').onclick=closeModal;
  bg.addEventListener('click',e=>{if(e.target===bg)closeModal()});

  let identitySeq=0;
  async function autofillDossier(){
    const dossier=String(el('manualDossier').value||'').trim();
    const seq=++identitySeq;
    if(!dossier)return;
    try{
      const r=await fetch('/api/articles/by-dossier?dossier='+encodeURIComponent(dossier),{cache:'no-store'});
      const d=await r.json();
      if(seq!==identitySeq||!r.ok)return;
      const rows=d.articles||[];
      const first=rows.find(x=>String(x.client||'').trim()||String(x.projet||'').trim());
      if(first){
        if(String(first.client||'').trim())el('manualClient').value=first.client;
        if(String(first.projet||'').trim())el('manualProjet').value=first.projet;
        el('manualArticleHint').textContent='Dossier existant : Client et Projet ont été repris automatiquement.';
      }else{
        el('manualArticleHint').textContent='Nouveau dossier ou dossier sans identité connue : renseigne Client et Projet si nécessaire.';
      }
    }catch(e){}
  }
  el('manualDossier').addEventListener('change',autofillDossier);
  el('manualDossier').addEventListener('blur',autofillDossier);

  el('manualArticleSave').onclick=async()=>{
    const dossier=String(el('manualDossier').value||'').trim();
    const reference=String(el('manualReference').value||'').trim();
    const description=String(el('manualDescription').value||'').trim();
    if(!dossier){alert('Le N° de dossier est obligatoire.');el('manualDossier').focus();return}
    if(!reference&&!description){alert('Renseigne au minimum une référence ou une description.');el('manualReference').focus();return}

    const payload={
      type_objet:el('manualType').value,
      dossier,
      reference,
      description,
      client:String(el('manualClient').value||'').trim(),
      projet:String(el('manualProjet').value||'').trim(),
      ref_caisse:String(el('manualRefCaisse').value||'').trim(),
      transporteur_ref:String(el('manualTransporteurRef').value||'').trim(),
      lieu_stockage:String(el('manualLieuStockage').value||'').trim(),
      longueur_cm:String(el('manualLongueur').value||'').trim(),
      largeur_cm:String(el('manualLargeur').value||'').trim(),
      hauteur_cm:String(el('manualHauteur').value||'').trim(),
      poids_kg:String(el('manualPoids').value||'').trim(),
      volume_m3:String(el('manualVolume').value||'').trim(),
      surface_m2:String(el('manualSurface').value||'').trim(),
      statut_logistique:String(el('manualStatut').value||'').trim()||'Créé'
    };

    const save=el('manualArticleSave');
    save.disabled=true;save.textContent='Création…';
    try{
      const r=await fetch('/api/articles/manual',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
      const text=await r.text();let d={};try{d=text?JSON.parse(text):{}}catch(e){}
      if(!r.ok)throw new Error(d.error||'Impossible de créer l’article');
      const esi=(d.article&&d.article.esi_id)||d.esi_id||'';
      closeModal();
      if(typeof load==='function')await load();
      if(esi&&typeof openArticleDetail==='function')await openArticleDetail(esi);
      else alert('Article créé'+(esi?' : '+esi:''));
    }catch(e){alert(e.message||'Impossible de créer l’article')}
    finally{save.disabled=false;save.textContent="Créer l'article"}
  };
}

if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',ensureManualCreateUI);else ensureManualCreateUI();
})();"""

@app.route('/articles')
def articles_page():
    page = render_template('articles.html')
    inline = '<script>' + ARTICLES_MANUAL_CREATE_JS + '</script>'
    if '</body>' in page:
        page = page.replace('</body>', inline + '\n</body>', 1)
    else:
        page += inline
    response = app.make_response(page)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


@app.route('/api/articles')
def api_articles():
    q = _as_text(request.args.get("q")).strip()
    type_objet = _as_text(request.args.get("type_objet")).strip().upper()
    limit = min(max(int(request.args.get("limit") or 100), 1), 500)

    query = "select=*&order=article_no.desc&limit=" + str(limit)

    if type_objet in ("PRODUIT", "CONTENANT"):
        query += "&type_objet=eq." + urllib.parse.quote(type_objet, safe='')

    if q:
        pattern = "*" + q.replace("*", "") + "*"
        query += "&search_text=ilike." + urllib.parse.quote(pattern, safe='*')

    rows = supabase_rest_request("GET", "articles", query) or []
    return jsonify([_article_row_to_public(row) for row in rows])


@app.route('/api/articles/manual', methods=['POST'])
def api_article_create_manual():
    """Crée un article ou un contenant directement depuis la base Articles."""
    data = request.get_json(silent=True) or {}

    type_objet = _as_text(data.get('type_objet') or 'PRODUIT').strip().upper()
    if type_objet not in ('PRODUIT', 'CONTENANT'):
        return jsonify({'ok': False, 'error': 'Type invalide : PRODUIT ou CONTENANT attendu'}), 400

    dossier = _as_text(data.get('dossier')).strip()
    reference = _as_text(data.get('reference')).strip()
    description = _as_text(data.get('description')).strip()

    if not dossier:
        return jsonify({'ok': False, 'error': 'Le N° de dossier est obligatoire'}), 400
    if not reference and not description:
        return jsonify({'ok': False, 'error': 'Renseigne au minimum une référence ou une description'}), 400

    try:
        identity = _article_dossier_identity(dossier)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Impossible de lire les informations du dossier : {e}'}), 500

    # Pour un dossier déjà connu, l'identité existante reste prioritaire afin
    # d'éviter des Client / Projet différents pour un même N° de dossier.
    client = _as_text(identity.get('client')).strip() or _as_text(data.get('client')).strip()
    projet = _as_text(identity.get('projet')).strip() or _as_text(data.get('projet')).strip()
    now = datetime.now().isoformat()

    payload = {
        'ticket_id': None,
        'source_module': 'Création manuelle',
        'source_index': None,
        'unit_index': 1,
        'type_objet': type_objet,
        'reference': reference,
        'description': description,
        'dossier': dossier,
        'client': client,
        'projet': projet,
        'ref_caisse': _as_text(data.get('ref_caisse')).strip(),
        'transporteur_ref': _as_text(data.get('transporteur_ref')).strip(),
        'longueur_cm': _as_text(data.get('longueur_cm')).strip(),
        'largeur_cm': _as_text(data.get('largeur_cm')).strip(),
        'hauteur_cm': _as_text(data.get('hauteur_cm')).strip(),
        'volume_m3': _as_text(data.get('volume_m3')).strip(),
        'surface_m2': _as_text(data.get('surface_m2')).strip(),
        'poids_kg': _as_text(data.get('poids_kg')).strip(),
        'lieu_stockage': _as_text(data.get('lieu_stockage')).strip(),
        'statut_logistique': _as_text(data.get('statut_logistique')).strip() or 'Créé',
        'created_at': now,
        'updated_at': now,
        'raw_json': {
            'source': 'creation_manuelle',
            'dossier': dossier,
            'created_at': now,
        },
    }
    payload['search_text'] = _article_search_text(payload)

    try:
        with _ARTICLE_LOCK:
            article = _create_article_record(payload)
        return jsonify({
            'ok': True,
            'esi_id': article.get('esi_id'),
            'article': article,
        }), 201
    except Exception as e:
        print(f"[ARTICLE MANUEL] Création impossible : {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/articles/link-search')
def api_articles_link_search():
    """
    Recherche les PRODUITS liables a une fiche de caisse.

    - recherche exacte par N dossier en priorite ;
    - recherche partielle par reference ;
    - secours par dossier partiel ;
    - les CONTENANTS sont toujours exclus.
    """
    q = _as_text(request.args.get('q')).strip()
    if not q:
        return jsonify({'ok': True, 'articles': [], 'count': 0})

    clean_q = q.replace('*', '').strip()
    if not clean_q:
        return jsonify({'ok': True, 'articles': [], 'count': 0})

    rows = []
    seen = set()

    def add_rows(found_rows):
        for row in found_rows or []:
            esi_id = _as_text(row.get('esi_id')).strip()
            if not esi_id or esi_id in seen:
                continue
            type_objet = _as_text(row.get('type_objet') or 'PRODUIT').strip().upper()
            if type_objet == 'CONTENANT':
                continue
            seen.add(esi_id)
            rows.append(row)

    try:
        # 1) N dossier exact : cas principal depuis une fiche de caisse.
        safe_dossier = urllib.parse.quote(clean_q, safe='')
        dossier_rows = supabase_rest_request(
            'GET', 'articles',
            'select=esi_id,dossier,reference,type_objet,article_no'
            '&dossier=eq.' + safe_dossier +
            '&order=article_no.asc&limit=500'
        ) or []
        add_rows(dossier_rows)

        # 2) Reference / numero inventaire partiel.
        safe_ref = urllib.parse.quote('*' + clean_q + '*', safe='*')
        reference_rows = supabase_rest_request(
            'GET', 'articles',
            'select=esi_id,dossier,reference,type_objet,article_no'
            '&reference=ilike.' + safe_ref +
            '&order=article_no.asc&limit=150'
        ) or []
        add_rows(reference_rows)

        # 3) Secours : dossier partiel.
        safe_dossier_like = urllib.parse.quote('*' + clean_q + '*', safe='*')
        dossier_like_rows = supabase_rest_request(
            'GET', 'articles',
            'select=esi_id,dossier,reference,type_objet,article_no'
            '&dossier=ilike.' + safe_dossier_like +
            '&order=article_no.asc&limit=150'
        ) or []
        add_rows(dossier_like_rows)

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'articles': []}), 500

    articles = [{
        'esi_id': _as_text(row.get('esi_id')).strip(),
        'dossier': _as_text(row.get('dossier')).strip(),
        'reference': _as_text(row.get('reference')).strip(),
    } for row in rows]

    articles.sort(key=lambda a: (
        _as_text(a.get('dossier')).casefold(),
        _as_text(a.get('reference')).casefold(),
        _as_text(a.get('esi_id')).casefold(),
    ))

    return jsonify({'ok': True, 'articles': articles, 'count': len(articles)})


@app.route('/api/articles/by-dossier')
def api_articles_by_dossier():
    """Retourne uniquement les articles déjà enregistrés pour un N° de dossier donné."""
    dossier = _as_text(request.args.get("dossier")).strip()
    if not dossier:
        return jsonify({"ok": False, "error": "Le N° de dossier est obligatoire", "articles": []}), 400

    safe_dossier = urllib.parse.quote(dossier, safe='')
    rows = supabase_rest_request(
        "GET",
        "articles",
        f"select=*&dossier=eq.{safe_dossier}&order=article_no.asc&limit=5000"
    ) or []

    articles = [_article_row_to_public(row) for row in rows]
    return jsonify({
        "ok": True,
        "dossier": dossier,
        "count": len(articles),
        "articles": articles,
    })


def _article_file_link(ticket_id, file_info, kind):
    if not isinstance(file_info, dict) or not file_info.get("name"):
        return None
    filename = _as_text(file_info.get("name")).strip()
    if not filename:
        return None
    base = "download-sheet" if kind == "gestionnaire" else "download"
    return {
        "name": filename,
        "kind": kind,
        "url": f"/api/tickets/{urllib.parse.quote(ticket_id, safe='')}/{base}/{urllib.parse.quote(filename, safe='')}",
    }


def _article_reception_history_from_ticket(ticket, article):
    esi_id = _as_text(article.get("esi_id")).strip()
    history = []

    module = _as_text(ticket.get("module")).replace("’", "'").strip()
    if module == "Avis d'arrivée":
        receptions = ticket.get("receptionsAvisArrivee") or []
    else:
        receptions = (ticket.get("enlevement") or {}).get("bons_livraison") or []

    for reception in receptions:
        items = reception.get("items") or []
        linked_item = None
        for item in items:
            if esi_id in [str(x).strip() for x in (item.get("esi_ids") or [])]:
                linked_item = item
                break
        if not linked_item:
            continue

        files = []
        candidates = [
            reception.get("bon_reception_filename"),
            reception.get("filename"),
            reception.get("etiquettes_articles_filename"),
            reception.get("etiquettes_colis_filename"),
        ]
        seen = set()
        for filename in candidates:
            filename = _as_text(filename).strip()
            if not filename or filename in seen:
                continue
            seen.add(filename)
            files.append({
                "name": filename,
                "kind": "gestionnaire",
                "url": f"/api/tickets/{urllib.parse.quote(ticket.get('id') or '', safe='')}/download-sheet/{urllib.parse.quote(filename, safe='')}",
            })

        history.append({
            "type": "Réception",
            "reference": reception.get("reference") or "",
            "date": reception.get("receptionnee_le") or reception.get("created_at") or reception.get("date_reception") or "",
            "date_affichee": reception.get("date_reception") or "",
            "receptionne_par": reception.get("receptionne_par") or "",
            "lieu_stockage": reception.get("lieu_stockage") or linked_item.get("lieu_stockage") or "",
            "numero_dossier": reception.get("numero_dossier") or article.get("dossier") or "",
            "nombre_colis": reception.get("nombre_colis") or "",
            "colis": reception.get("colis") or linked_item.get("colis") or [],
            "quantite": linked_item.get("quantite") or "",
            "files": files,
        })

    return history




_ARTICLE_EDITABLE_FIELDS = {
    "type_objet", "reference", "description", "dossier", "client", "projet",
    "ref_caisse", "transporteur_ref",
    "longueur_cm", "largeur_cm", "hauteur_cm",
    "volume_m3", "surface_m2", "poids_kg",
    "lieu_stockage", "statut_logistique", "dernier_colis",
}


def _article_has_history(article):
    """Protège de la suppression les articles déjà liés à l'historique métier."""
    if _as_text(article.get("ticket_id")).strip():
        return True
    raw = article.get("raw_json") if isinstance(article.get("raw_json"), dict) else {}
    if raw.get("receptions"):
        return True
    if _as_text(article.get("derniere_reception_ref")).strip():
        return True
    if _as_text(article.get("dernier_colis")).strip():
        return True
    return False


def _article_dossier_identity(dossier):
    """Retourne le couple Client / Projet le plus fréquent pour un N° de dossier existant."""
    dossier = _as_text(dossier).strip()
    if not dossier:
        return {"client": "", "projet": ""}

    safe_dossier = urllib.parse.quote(dossier, safe='')
    rows = supabase_rest_request(
        "GET", "articles",
        f"select=client,projet&dossier=eq.{safe_dossier}&limit=10000"
    ) or []

    counts = {}
    for row in rows:
        client = _as_text(row.get("client")).strip()
        projet = _as_text(row.get("projet")).strip()
        if not client and not projet:
            continue
        key = (client, projet)
        counts[key] = counts.get(key, 0) + 1

    if not counts:
        return {"client": "", "projet": ""}

    client, projet = max(counts.items(), key=lambda kv: (kv[1], bool(kv[0][0]), bool(kv[0][1])))[0]
    return {"client": client, "projet": projet}


def _article_sync_dossier_identity(dossier, client=None, projet=None):
    """Uniformise Client / Projet sur tous les articles portant le même N° de dossier."""
    dossier = _as_text(dossier).strip()
    if not dossier:
        return {"updated_count": 0, "client": "", "projet": ""}

    identity = _article_dossier_identity(dossier)
    final_client = _as_text(client).strip() if client is not None else identity.get("client", "")
    final_projet = _as_text(projet).strip() if projet is not None else identity.get("projet", "")

    if not final_client and not final_projet:
        return {"updated_count": 0, "client": "", "projet": ""}

    safe_dossier = urllib.parse.quote(dossier, safe='')
    rows = supabase_rest_request(
        "GET", "articles", f"select=*&dossier=eq.{safe_dossier}&limit=10000"
    ) or []

    updated = 0
    now = datetime.now().isoformat()
    for row in rows:
        patch = {}
        if final_client and _as_text(row.get("client")).strip() != final_client:
            patch["client"] = final_client
        if final_projet and _as_text(row.get("projet")).strip() != final_projet:
            patch["projet"] = final_projet
        if not patch:
            continue

        merged = dict(row)
        merged.update(patch)
        patch["updated_at"] = now
        patch["search_text"] = _article_search_text(merged)
        safe_esi = urllib.parse.quote(_as_text(row.get("esi_id")).strip(), safe='-')
        supabase_rest_request(
            "PATCH", "articles", f"esi_id=eq.{safe_esi}", patch, prefer="return=minimal"
        )
        updated += 1

    return {"updated_count": updated, "client": final_client, "projet": final_projet}


def _article_duplicate_key(article):
    """Clé conservative : ne confond jamais deux unités physiques légitimes."""
    raw = article.get("raw_json") if isinstance(article.get("raw_json"), dict) else {}
    source_module = _as_text(article.get("source_module")).strip()
    ticket_id = _as_text(article.get("ticket_id")).strip()
    source_index = article.get("source_index")
    unit_index = article.get("unit_index")

    if ticket_id and source_index is not None and unit_index is not None:
        return ("ticket", ticket_id, str(source_index), str(unit_index))

    if source_module == "Import Excel":
        filename = _as_text(raw.get("filename")).strip()
        dossier = _as_text(article.get("dossier") or raw.get("dossier")).strip()
        excel_row = raw.get("excel_row") if raw.get("excel_row") is not None else source_index
        if filename and dossier and excel_row is not None and unit_index is not None:
            return ("excel", filename.casefold(), dossier.casefold(), str(excel_row), str(unit_index))

    return None


def _article_duplicate_groups():
    rows = supabase_rest_request(
        "GET", "articles", "select=*&order=article_no.asc&limit=10000"
    ) or []
    grouped = {}
    for row in rows:
        key = _article_duplicate_key(row)
        if key is None:
            continue
        grouped.setdefault(key, []).append(dict(row))
    return [items for items in grouped.values() if len(items) > 1]


@app.route('/api/articles/bulk-update', methods=['PATCH'])
def api_articles_bulk_update():
    data = request.get_json(silent=True) or {}
    esi_ids = data.get('esi_ids') or []
    changes = data.get('changes') or {}

    if not isinstance(esi_ids, list) or not esi_ids:
        return jsonify({'ok': False, 'error': 'Aucun article sélectionné'}), 400
    if not isinstance(changes, dict) or not changes:
        return jsonify({'ok': False, 'error': 'Aucune modification demandée'}), 400

    clean_changes = {}
    for field, value in changes.items():
        if field not in _ARTICLE_EDITABLE_FIELDS:
            continue
        clean_changes[field] = _as_text(value).strip()

    if not clean_changes:
        return jsonify({'ok': False, 'error': 'Aucun champ modifiable fourni'}), 400

    # Si le dossier existe déjà dans la base, Client et Projet sont repris automatiquement.
    target_dossier = _as_text(clean_changes.get('dossier')).strip() if 'dossier' in clean_changes else ''
    if target_dossier:
        identity = _article_dossier_identity(target_dossier)
        if 'client' not in clean_changes and identity.get('client'):
            clean_changes['client'] = identity['client']
        if 'projet' not in clean_changes and identity.get('projet'):
            clean_changes['projet'] = identity['projet']

    updated = []
    errors = []
    now = datetime.now().isoformat()
    affected_dossiers = set()

    with _ARTICLE_LOCK:
        for esi_id in dict.fromkeys(_as_text(x).strip() for x in esi_ids if _as_text(x).strip()):
            safe_esi = urllib.parse.quote(esi_id, safe='-')
            try:
                rows = supabase_rest_request(
                    'GET', 'articles', f'select=*&esi_id=eq.{safe_esi}&limit=1'
                ) or []
                if not rows:
                    errors.append({'esi_id': esi_id, 'error': 'Article introuvable'})
                    continue

                current = dict(rows[0])
                patch = dict(clean_changes)
                patch['updated_at'] = now
                merged = dict(current)
                merged.update(patch)
                patch['search_text'] = _article_search_text(merged)

                supabase_rest_request(
                    'PATCH', 'articles', f'esi_id=eq.{safe_esi}',
                    patch, prefer='return=minimal'
                )
                updated.append(esi_id)
                dossier_after = _as_text(merged.get('dossier')).strip()
                if dossier_after:
                    affected_dossiers.add(dossier_after)
            except Exception as e:
                errors.append({'esi_id': esi_id, 'error': str(e)})

        # Tous les articles d'un même dossier héritent du même Client / Projet.
        dossier_sync = []
        for dossier in sorted(affected_dossiers):
            try:
                sync_client = clean_changes.get('client') if 'client' in clean_changes else None
                sync_projet = clean_changes.get('projet') if 'projet' in clean_changes else None
                result = _article_sync_dossier_identity(dossier, sync_client, sync_projet)
                dossier_sync.append({'dossier': dossier, **result})
            except Exception as e:
                errors.append({'dossier': dossier, 'error': f'Synchronisation Client/Projet : {e}'})

    return jsonify({
        'ok': not errors,
        'updated_count': len(updated),
        'updated': updated,
        'dossier_sync': dossier_sync,
        'errors': errors,
    }), (200 if not errors else 207)


@app.route('/api/articles/bulk-delete', methods=['POST'])
def api_articles_bulk_delete():
    data = request.get_json(silent=True) or {}
    esi_ids = list(dict.fromkeys(
        _as_text(x).strip() for x in (data.get('esi_ids') or []) if _as_text(x).strip()
    ))

    if not esi_ids:
        return jsonify({'ok': False, 'error': 'Aucun article sélectionné'}), 400

    deleted = []
    protected = []
    errors = []

    with _ARTICLE_LOCK:
        for offset in range(0, len(esi_ids), 100):
            part = esi_ids[offset:offset + 100]
            encoded = urllib.parse.quote(','.join(part), safe=',-_')
            try:
                rows = supabase_rest_request(
                    'GET', 'articles', f'select=*&esi_id=in.({encoded})&limit=100'
                ) or []
                by_id = {_as_text(r.get('esi_id')).strip(): dict(r) for r in rows}
                deletable = []
                for esi_id in part:
                    article = by_id.get(esi_id)
                    if not article:
                        errors.append({'esi_id': esi_id, 'error': 'Article introuvable'})
                        continue
                    if _article_has_history(article):
                        protected.append({
                            'esi_id': esi_id,
                            'reason': 'Article lié à un ticket ou à un historique de réception'
                        })
                    else:
                        deletable.append(esi_id)

                if deletable:
                    encoded_delete = urllib.parse.quote(','.join(deletable), safe=',-_')
                    supabase_rest_request(
                        'DELETE', 'articles', f'esi_id=in.({encoded_delete})', prefer='return=minimal'
                    )
                    deleted.extend(deletable)
            except Exception as e:
                errors.append({'esi_ids': part, 'error': str(e)})

    return jsonify({
        'ok': not errors and not protected,
        'deleted_count': len(deleted),
        'deleted': deleted,
        'protected': protected,
        'errors': errors,
    }), (200 if not errors and not protected else 207)


@app.route('/api/articles/duplicates', methods=['GET'])
def api_articles_duplicates():
    try:
        groups = _article_duplicate_groups()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

    public_groups = []
    duplicate_count = 0
    deletable_count = 0
    protected_count = 0
    for items in groups:
        ordered = sorted(items, key=lambda a: int(a.get('article_no') or 0))
        keep = ordered[0]
        extras = ordered[1:]
        deletable = [x for x in extras if not _article_has_history(x)]
        protected = [x for x in extras if _article_has_history(x)]
        duplicate_count += len(extras)
        deletable_count += len(deletable)
        protected_count += len(protected)
        public_groups.append({
            'keep': _article_row_to_public(keep),
            'duplicates': [_article_row_to_public(x) for x in extras],
            'deletable_esi_ids': [_as_text(x.get('esi_id')).strip() for x in deletable],
            'protected_esi_ids': [_as_text(x.get('esi_id')).strip() for x in protected],
        })

    return jsonify({
        'ok': True,
        'group_count': len(groups),
        'duplicate_count': duplicate_count,
        'deletable_count': deletable_count,
        'protected_count': protected_count,
        'groups': public_groups,
    })


@app.route('/api/articles/delete-duplicates', methods=['POST'])
def api_articles_delete_duplicates():
    try:
        groups = _article_duplicate_groups()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

    to_delete = []
    protected = []
    for items in groups:
        ordered = sorted(items, key=lambda a: int(a.get('article_no') or 0))
        for article in ordered[1:]:
            esi_id = _as_text(article.get('esi_id')).strip()
            if _article_has_history(article):
                protected.append(esi_id)
            elif esi_id:
                to_delete.append(esi_id)

    deleted = []
    errors = []
    with _ARTICLE_LOCK:
        for offset in range(0, len(to_delete), 100):
            part = to_delete[offset:offset + 100]
            encoded = urllib.parse.quote(','.join(part), safe=',-_')
            try:
                supabase_rest_request(
                    'DELETE', 'articles', f'esi_id=in.({encoded})', prefer='return=minimal'
                )
                deleted.extend(part)
            except Exception as e:
                errors.append({'esi_ids': part, 'error': str(e)})

    return jsonify({
        'ok': not errors,
        'deleted_count': len(deleted),
        'deleted': deleted,
        'protected_count': len(protected),
        'protected': protected,
        'errors': errors,
    }), (200 if not errors else 207)


@app.route('/api/articles/colis-by-dossier')
def api_articles_colis_by_dossier():
    dossier=_as_text(request.args.get('dossier')).strip()
    if not dossier: return jsonify({'ok':False,'error':'N° dossier obligatoire','colis':[]}),400
    nums=sorted(_existing_colis_numbers(dossier))
    return jsonify({'ok':True,'dossier':dossier,'colis':[f"{dossier}-{n:03d}" for n in nums]})

@app.route('/api/articles/<esi_id>/photo', methods=['GET', 'POST', 'DELETE'])
def api_article_photo(esi_id):
    """Affiche, ajoute/remplace ou retire la photo principale d'un article."""
    esi_id = _as_text(esi_id).strip()
    safe_esi = urllib.parse.quote(esi_id, safe='-')
    rows = supabase_rest_request(
        'GET', 'articles', f'select=*&esi_id=eq.{safe_esi}&limit=1'
    ) or []
    if not rows:
        return jsonify({'ok': False, 'error': 'Article introuvable'}), 404

    article = dict(rows[0])
    raw = article.get('raw_json') if isinstance(article.get('raw_json'), dict) else {}
    raw = dict(raw or {})

    if request.method == 'GET':
        storage_path = _as_text(raw.get('photo_storage_path')).strip()
        if not storage_path:
            abort(404)
        try:
            return redirect(supabase_signed_download_url(storage_path, expires_in=900))
        except Exception as e:
            print(f'[ARTICLE PHOTO] URL signee impossible pour {esi_id}: {e}')
            try:
                data = supabase_download_bytes(storage_path)
            except Exception:
                abort(404)
            return send_file(
                BytesIO(data),
                mimetype=_as_text(raw.get('photo_content_type')).strip() or 'image/jpeg',
                download_name=_as_text(raw.get('photo_filename')).strip() or f'{esi_id}.jpg',
            )

    if request.method == 'DELETE':
        for key in (
            'photo_storage_path', 'photo_filename', 'photo_content_type',
            'photo_updated_at', 'photo_url', 'photo', 'image_url'
        ):
            raw.pop(key, None)
        raw['photo_deleted_at'] = datetime.now().isoformat()
        patch = {
            'raw_json': raw,
            'updated_at': datetime.now().isoformat(),
        }
        merged = dict(article)
        merged.update(patch)
        patch['search_text'] = _article_search_text(merged)
        supabase_rest_request(
            'PATCH', 'articles', f'esi_id=eq.{safe_esi}', patch, prefer='return=minimal'
        )
        return jsonify({'ok': True, 'esi_id': esi_id, 'photo_url': ''})

    fs = request.files.get('photo') or request.files.get('file')
    if not fs or not fs.filename:
        return jsonify({'ok': False, 'error': 'Photo manquante'}), 400

    filename = _as_text(fs.filename).strip()
    ext = Path(filename).suffix.lower()
    allowed_ext = {'.jpg', '.jpeg', '.png', '.webp'}
    if ext not in allowed_ext:
        return jsonify({
            'ok': False,
            'error': 'Format photo non pris en charge. Utilise JPG, PNG ou WEBP.'
        }), 400

    content = fs.read()
    if not content:
        return jsonify({'ok': False, 'error': 'Le fichier image est vide'}), 400
    if len(content) > 15 * 1024 * 1024:
        return jsonify({'ok': False, 'error': 'La photo dépasse la limite de 15 Mo'}), 400

    # Vérification simple de la signature du fichier pour éviter qu'un autre type
    # de contenu soit envoyé avec une extension d'image.
    valid_signature = False
    if ext in ('.jpg', '.jpeg'):
        valid_signature = content.startswith(b'\xff\xd8\xff')
    elif ext == '.png':
        valid_signature = content.startswith(b'\x89PNG\r\n\x1a\n')
    elif ext == '.webp':
        valid_signature = len(content) >= 12 and content[:4] == b'RIFF' and content[8:12] == b'WEBP'
    if not valid_signature:
        return jsonify({'ok': False, 'error': 'Le fichier ne semble pas être une image valide'}), 400

    safe_name = safe_filename(filename) or ('photo' + ext)
    stamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
    storage_path = f'article_photos/{safe_filename(esi_id)}/{stamp}_{safe_name}'
    content_type = _as_text(fs.content_type).strip().lower()
    if content_type not in {'image/jpeg', 'image/png', 'image/webp'}:
        content_type = {
            '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
            '.png': 'image/png', '.webp': 'image/webp'
        }[ext]

    try:
        supabase_upload_bytes(storage_path, content, content_type)
    except Exception as e:
        return jsonify({'ok': False, 'error': f"Impossible d'enregistrer la photo : {e}"}), 500

    raw['photo_storage_path'] = storage_path
    raw['photo_filename'] = filename
    raw['photo_content_type'] = content_type
    raw['photo_updated_at'] = datetime.now().isoformat()
    raw.pop('photo_deleted_at', None)

    patch = {
        'raw_json': raw,
        'updated_at': datetime.now().isoformat(),
    }
    merged = dict(article)
    merged.update(patch)
    patch['search_text'] = _article_search_text(merged)
    supabase_rest_request(
        'PATCH', 'articles', f'esi_id=eq.{safe_esi}', patch, prefer='return=minimal'
    )

    photo_url = f'/api/articles/{urllib.parse.quote(esi_id, safe="-")}/photo?v={urllib.parse.quote(raw["photo_updated_at"], safe="")}'
    return jsonify({
        'ok': True,
        'esi_id': esi_id,
        'photo_url': photo_url,
        'filename': filename,
    })


@app.route('/api/articles/<esi_id>/colis', methods=['PATCH'])
def api_article_update_colis(esi_id):
    safe_esi=urllib.parse.quote(_as_text(esi_id).strip(),safe='-')
    rows=supabase_rest_request('GET','articles',f'select=*&esi_id=eq.{safe_esi}&limit=1') or []
    if not rows: return jsonify({'ok':False,'error':'Article introuvable'}),404
    article=dict(rows[0]); dossier=_as_text(article.get('dossier')).strip()
    colis=_as_text((request.get_json(silent=True) or {}).get('colis')).strip()
    if colis:
        allowed={f"{dossier}-{n:03d}" for n in _existing_colis_numbers(dossier)}
        if colis not in allowed: return jsonify({'ok':False,'error':"Ce colis n'existe pas pour ce dossier"}),400
    raw=article.get('raw_json') if isinstance(article.get('raw_json'),dict) else {}; raw=dict(raw or {})
    raw['colis_actuel']=colis; mods=list(raw.get('modifications_colis') or []); mods.append({'date':datetime.now().isoformat(),'colis':colis}); raw['modifications_colis']=mods
    patch={'dernier_colis':colis,'raw_json':raw,'updated_at':datetime.now().isoformat()}
    merged=dict(article); merged.update(patch); patch['search_text']=_article_search_text(merged)
    supabase_rest_request('PATCH','articles',f'esi_id=eq.{safe_esi}',patch,prefer='return=minimal')
    return jsonify({'ok':True,'esi_id':esi_id,'colis':colis})

@app.route('/api/articles/<esi_id>')
def api_article_detail(esi_id):
    esi_id = _as_text(esi_id).strip()
    safe_esi = urllib.parse.quote(esi_id, safe='-')
    rows = supabase_rest_request(
        "GET", "articles", f"select=*&esi_id=eq.{safe_esi}&limit=1"
    ) or []
    if not rows:
        return jsonify({"error": "Article introuvable"}), 404

    article = _article_row_to_public(rows[0])
    article_raw = article.get("raw_json") if isinstance(article.get("raw_json"), dict) else {}
    article_raw = dict(article_raw or {})
    photo_storage_path = _as_text(article_raw.get("photo_storage_path")).strip()
    if photo_storage_path:
        photo_version = _as_text(article_raw.get("photo_updated_at") or article.get("updated_at")).strip()
        article["photo_url"] = (
            f"/api/articles/{urllib.parse.quote(esi_id, safe='-')}/photo"
            + ("?v=" + urllib.parse.quote(photo_version, safe='') if photo_version else "")
        )
        article["photo_filename"] = _as_text(article_raw.get("photo_filename")).strip()

    ticket_id = _as_text(article.get("ticket_id")).strip()
    ticket = load_ticket(ticket_id) if ticket_id else None

    detail = {
        "article": article,
        "ticket": None,
        "avis_arrivee": None,
        "demande_enlevement": None,
        "receptions": [],
        "documents_source": [],
    }

    if ticket:
        detail["ticket"] = {
            "id": ticket.get("id"),
            "module": ticket.get("module"),
            "status": ticket.get("status"),
            "created_at": ticket.get("createdAt"),
            "updated_at": ticket.get("updatedAt"),
            "dossier": ticket.get("dossier"),
            "ref": ticket.get("ref"),
            "client": ticket.get("dossier"),
            "projet": ticket.get("expo") or ticket.get("objet"),
            "charge_projet": ticket.get("chargeProjet"),
        }

        for f in ticket.get("files") or []:
            link = _article_file_link(ticket_id, f, "demandeur")
            if link:
                detail["documents_source"].append(link)

        module = _as_text(ticket.get("module")).replace("’", "'").strip()
        if module == "Avis d'arrivée":
            avis = ticket.get("avisArrivee") or ticket.get("avis_arrivee") or {}
            detail["avis_arrivee"] = {
                "dossier_ref": avis.get("dossier_ref") or "",
                "client": avis.get("client") or "",
                "projet": avis.get("projet") or "",
                "date_reception_prevue": avis.get("date_reception_prevue") or "",
                "coordinateur": avis.get("coordinateur") or "",
                "commentaire": avis.get("commentaire") or ticket.get("commentaire") or "",
                "expediteur": avis.get("expediteur") or {},
                "transporteur": avis.get("transporteur") or {},
            }
        else:
            enl = ticket.get("enlevement") or {}
            detail["demande_enlevement"] = {
                "numero_bon": enl.get("numero_bon") or ticket.get("ref") or "",
                "client": enl.get("client") or ticket.get("dossier") or "",
                "projet": enl.get("exhibition") or ticket.get("expo") or ticket.get("objet") or "",
                "coordinateur": enl.get("coordinateur") or ticket.get("chargeProjet") or "",
                "date_enlevement": enl.get("date_enlevement") or "",
                "adresse_depart": enl.get("adresse_depart") or "",
                "adresse_destination": enl.get("adresse_destination") or "",
                "notes": enl.get("notes") or "",
                "instructions": enl.get("instructions") or "",
            }

        detail["receptions"] = _article_reception_history_from_ticket(ticket, article)

    raw = article_raw
    raw_receptions = raw.get("receptions") or []
    if raw_receptions:
        known_refs = {str(x.get("reference") or "") for x in detail["receptions"]}
        for r in raw_receptions:
            ref = _as_text(r.get("reception_ref")).strip()
            if ref and ref in known_refs:
                continue
            detail["receptions"].append({
                "type": "Réception",
                "reference": ref,
                "date": r.get("date") or "",
                "date_affichee": "",
                "receptionne_par": r.get("receptionne_par") or "",
                "lieu_stockage": r.get("lieu_stockage") or "",
                "numero_dossier": article.get("dossier") or "",
                "nombre_colis": "",
                "colis": r.get("colis") or [],
                "quantite": "1",
                "files": [],
            })

    return jsonify(detail)




def _article_import_text(value):
    """Normalise une valeur Excel sans transformer 12 en '12.0'."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return (f"{value:.10f}").rstrip("0").rstrip(".")
    return re.sub(r"\s+", " ", _as_text(value).replace("\u00a0", " ")).strip()


def _article_import_norm(value):
    """Valeur canonique utilisée uniquement pour la détection de doublons."""
    text = _article_import_text(value).casefold().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _article_import_signature(values):
    """Signature métier d'un article importé pour éviter les créations en double."""
    return "|".join([
        _article_import_norm(values.get("reference")),
        _article_import_norm(values.get("description")),
        _article_import_norm(values.get("longueur_cm")),
        _article_import_norm(values.get("largeur_cm")),
        _article_import_norm(values.get("hauteur_cm")),
        _article_import_norm(values.get("poids_kg")),
    ])


def _article_import_headers(ws):
    """Repère les colonnes du format Excel normalisé ESI TICKETS."""
    aliases = {
        "quantite": {"QUANTITE", "QTE", "QTY"},
        "longueur_cm": {"LONGUEUR_CM", "LONGUEUR (CM)", "LONGUEUR"},
        "largeur_cm": {"LARGEUR_CM", "LARGEUR (CM)", "LARGEUR"},
        "hauteur_cm": {"HAUTEUR_CM", "HAUTEUR (CM)", "HAUTEUR"},
        "poids_kg": {"POIDS_BRUT_KG", "POIDS_KG", "POIDS (KG)", "POIDS"},
        "reference": {"REFERENCE_PINTO", "REFERENCE", "REF", "REF ARTICLE"},
        "description": {"DESIGNATION", "DESCRIPTION"},
    }

    found = {}
    for col in range(1, ws.max_column + 1):
        raw = _article_import_text(ws.cell(row=1, column=col).value)
        header = re.sub(r"\s+", " ", raw.upper()).strip()
        if not header:
            continue
        for field, names in aliases.items():
            if header in names and field not in found:
                found[field] = col
                break
    return found


def _article_import_existing_counts():
    """Compte les articles existants par signature métier."""
    rows = supabase_rest_request(
        "GET",
        "articles",
        "select=reference,description,longueur_cm,largeur_cm,hauteur_cm,poids_kg&limit=10000"
    ) or []
    counts = {}
    for row in rows:
        sig = _article_import_signature(row)
        if sig.strip("|"):
            counts[sig] = counts.get(sig, 0) + 1
    return counts


@app.route('/api/articles/import-excel', methods=['POST'])
def api_articles_import_excel():
    """
    Importe le fichier Excel normalisé ESI TICKETS.

    - 1 unité physique = 1 numéro ESI unique ;
    - la colonne QUANTITE peut donc créer plusieurs ESI pour une même ligne ;
    - les doublons sont contrôlés par référence + désignation + dimensions + poids ;
    - les photos sont volontairement ignorées pour le moment.
    """
    fs = request.files.get('file')
    if not fs or not fs.filename:
        return jsonify({'ok': False, 'error': 'Fichier Excel manquant'}), 400

    filename = _as_text(fs.filename).strip()
    dossier = _as_text(request.form.get('dossier')).strip()
    if not dossier:
        return jsonify({'ok': False, 'error': 'Le N° de dossier est obligatoire pour importer les articles.'}), 400

    if not filename.lower().endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'Le fichier doit être au format .xlsx'}), 400

    content = fs.read()
    if not content:
        return jsonify({'ok': False, 'error': 'Le fichier Excel est vide'}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Fichier Excel illisible : {e}'}), 400

    headers = _article_import_headers(ws)
    if 'reference' not in headers and 'description' not in headers:
        return jsonify({
            'ok': False,
            'error': "Colonnes non reconnues. Le fichier doit contenir au minimum REFERENCE_PINTO/REFERENCE ou DESIGNATION/DESCRIPTION."
        }), 400

    stats = {
        'lignes_lues': 0,
        'articles_demandes': 0,
        'articles_crees': 0,
        'doublons_ignores': 0,
        'lignes_ignorees': 0,
        'errors': [],
        'esi_ids': [],
    }

    try:
        existing_counts = _article_import_existing_counts()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Impossible de vérifier les doublons : {e}'}), 500

    now = datetime.now().isoformat()

    # Si ce dossier existe déjà, tous les nouveaux articles reprennent automatiquement
    # le même Client et le même Projet.
    try:
        dossier_identity = _article_dossier_identity(dossier)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Impossible de lire les informations du dossier : {e}'}), 500

    with _ARTICLE_LOCK:
        for row_num in range(2, ws.max_row + 1):
            def cell(field):
                col = headers.get(field)
                return ws.cell(row=row_num, column=col).value if col else None

            values = {
                'reference': _article_import_text(cell('reference')),
                'description': _article_import_text(cell('description')),
                'longueur_cm': _article_import_text(cell('longueur_cm')),
                'largeur_cm': _article_import_text(cell('largeur_cm')),
                'hauteur_cm': _article_import_text(cell('hauteur_cm')),
                'poids_kg': _article_import_text(cell('poids_kg')),
            }

            # Une ligne totalement vide n'est pas une erreur.
            if not any(values.values()) and not _article_import_text(cell('quantite')):
                continue

            stats['lignes_lues'] += 1

            if not values['reference'] and not values['description']:
                stats['lignes_ignorees'] += 1
                stats['errors'].append({
                    'ligne': row_num,
                    'error': 'Référence et désignation toutes les deux vides'
                })
                continue

            qty = _article_quantity(cell('quantite'), 1)
            stats['articles_demandes'] += qty

            signature = _article_import_signature(values)
            already = existing_counts.get(signature, 0)
            to_create = max(0, qty - already)
            skipped = qty - to_create
            stats['doublons_ignores'] += skipped

            created_for_row = 0
            for unit_index in range(already + 1, already + to_create + 1):
                payload = {
                    # Un import Excel dans la base Articles n'est pas rattache a un ticket.
                    # ticket_id reste a NULL : en PostgreSQL, la contrainte UNIQUE
                    # (ticket_id, source_index, unit_index) ne bloque pas plusieurs lignes
                    # lorsque ticket_id est NULL. source_index et unit_index restent renseignes
                    # car la table impose notamment unit_index NOT NULL.
                    'ticket_id': None,
                    'source_module': 'Import Excel',
                    'source_index': row_num,
                    'unit_index': unit_index,
                    'reference': values['reference'],
                    'description': values['description'],
                    'dossier': dossier,
                    'client': dossier_identity.get('client', ''),
                    'projet': dossier_identity.get('projet', ''),
                    'ref_caisse': '',
                    'transporteur_ref': '',
                    'longueur_cm': values['longueur_cm'],
                    'largeur_cm': values['largeur_cm'],
                    'hauteur_cm': values['hauteur_cm'],
                    'volume_m3': '',
                    'surface_m2': '',
                    'poids_kg': values['poids_kg'],
                    'lieu_stockage': '',
                    'statut_logistique': 'Créé',
                    'created_at': now,
                    'updated_at': now,
                    'raw_json': {
                        'source': 'import_excel',
                        'filename': filename,
                        'dossier': dossier,
                        'excel_row': row_num,
                        'quantity_source': qty,
                        'unit_index': unit_index,
                    },
                }
                payload['search_text'] = _article_search_text(payload)

                try:
                    article = _create_article_record(payload)
                    stats['articles_crees'] += 1
                    created_for_row += 1
                    if article.get('esi_id'):
                        stats['esi_ids'].append(article['esi_id'])
                except Exception as e:
                    stats['errors'].append({'ligne': row_num, 'error': str(e)})
                    break

            # Seules les créations réellement réussies deviennent des doublons pour les lignes suivantes.
            existing_counts[signature] = already + created_for_row

    try:
        wb.close()
    except Exception:
        pass

    # Uniformise également les anciens articles du dossier si une identité Client/Projet existe.
    try:
        if dossier_identity.get('client') or dossier_identity.get('projet'):
            _article_sync_dossier_identity(
                dossier, dossier_identity.get('client') or None, dossier_identity.get('projet') or None
            )
    except Exception as e:
        stats['errors'].append({'ligne': 0, 'error': f'Synchronisation dossier : {e}'})

    return jsonify({
        'ok': len(stats['errors']) == 0,
        'dossier': dossier,
        **stats,
    }), (200 if not stats['errors'] else 207)


@app.route('/api/articles/migrate', methods=['POST'])
def api_articles_migrate():
    """Importe les articles déjà présents dans les tickets et leur attribue un ESI-x."""
    stats = {"tickets": 0, "articles_crees": 0, "errors": []}
    try:
        tickets = list_tickets()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    for ticket in tickets:
        module = _as_text(ticket.get("module")).replace("’", "'").strip()
        if module not in ("Avis d'arrivée", "Demande d'enlèvement", "Demande d'enlevement"):
            continue
        try:
            created = _ensure_articles_for_ticket(ticket, save=True)
            stats["tickets"] += 1
            stats["articles_crees"] += len(created)
        except Exception as e:
            stats["errors"].append({"ticket_id": ticket.get("id"), "error": str(e)})

    return jsonify({
        "ok": not stats["errors"],
        **stats
    }), (200 if not stats["errors"] else 207)


def supabase_rest_request(method, table, query='', payload=None, prefer=None):
    """Appelle l'API REST Supabase Database sans dépendre du SDK Python."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if query:
        url += "?" + query.lstrip('?')

    data = None
    headers = {
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "apikey": SUPABASE_KEY,
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    elif method.upper() in ("POST", "PATCH", "DELETE"):
        headers["Prefer"] = "return=representation"

    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(url, data=data, method=method.upper(), headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            if not body:
                return None
            return json.loads(body)
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        print(f"[SUPABASE DB ERROR] {method} {url} -> HTTP {e.code} {e.reason} {body}")
        raise RuntimeError(f"Erreur Supabase DB HTTP {e.code}: {body or e.reason}")


def init_db():
    """Vérifie simplement que les tables Supabase répondent."""
    try:
        supabase_rest_request("GET", "tickets", "select=id&limit=1")
        print("[SUPABASE DB] Connexion OK")
    except Exception as e:
        print(f"[SUPABASE DB] Connexion impossible : {e}")


def _ticket_to_db_row(ticket):
    return {
        "id": _as_text(ticket.get("id")),
        "module": _as_text(ticket.get("module")),
        "status": _as_text(ticket.get("status")),
        "created_at": _as_text(ticket.get("createdAt")),
        "updated_at": _as_text(ticket.get("updatedAt")),
        "dossier": _as_text(ticket.get("dossier")),
        "ref": _as_text(ticket.get("ref")),
        "preteur": _as_text(ticket.get("preteur")),
        "expo": _as_text(ticket.get("expo")),
        "objet": _as_text(ticket.get("objet")),
        "charge_projet": _as_text(ticket.get("chargeProjet")),
        "type_caisse": _as_text(ticket.get("typeCaisse")),
        "dimensions": _as_text(ticket.get("dimensions")),
        "date_emballage": _as_text(ticket.get("dateEmballage")),
        "prix_devis": _as_text(ticket.get("prixDevis")),
        "date_rdv": _as_text(ticket.get("dateRdv")),
        "heure_rdv": _as_text(ticket.get("heureRdv")),
        "lieu_rdv": _as_text(ticket.get("lieuRdv")),
        "contact_rdv": _as_text(ticket.get("contactRdv")),
        "commentaire": _as_text(ticket.get("commentaire")),
        "validated_at": _as_text(ticket.get("validatedAt")),
        "raw_json": ticket,
    }


def _ticket_from_db_row(row):
    # raw_json permet de conserver des donnees metier additionnelles sans ajouter
    # immediatement de nouvelles colonnes Supabase (ex. analyse d'un bon d'enlevement).
    #
    # Selon la reponse PostgREST / le type reel de la colonne Supabase, raw_json
    # peut arriver soit deja decode en dict, soit sous forme de chaine JSON.
    raw = row.get("raw_json")
    if isinstance(raw, dict):
        ticket = dict(raw)
    elif isinstance(raw, str) and raw.strip():
        try:
            decoded = json.loads(raw)
            ticket = dict(decoded) if isinstance(decoded, dict) else {}
        except (json.JSONDecodeError, TypeError, ValueError):
            print(f"[SUPABASE DB] raw_json invalide pour ticket {row.get('id')}: {type(raw).__name__}")
            ticket = {}
    else:
        ticket = {}

    # Les colonnes principales restent la source de verite pour les champs historiques.
    ticket.update({
        "id": row.get("id") or "",
        "module": row.get("module") or "",
        "status": row.get("status") or "",
        "createdAt": row.get("created_at") or "",
        "updatedAt": row.get("updated_at") or "",
        "dossier": row.get("dossier") or "",
        "ref": row.get("ref") or "",
        "preteur": row.get("preteur") or "-",
        "expo": row.get("expo") or "-",
        "objet": row.get("objet") or "-",
        "chargeProjet": row.get("charge_projet") or "-",
        "typeCaisse": row.get("type_caisse") or "-",
        "dimensions": row.get("dimensions") or "-",
        "dateEmballage": row.get("date_emballage") or "-",
        "prixDevis": row.get("prix_devis") or "-",
        "dateRdv": row.get("date_rdv") or "-",
        "heureRdv": row.get("heure_rdv") or "-",
        "lieuRdv": row.get("lieu_rdv") or "-",
        "contactRdv": row.get("contact_rdv") or "-",
        "commentaire": row.get("commentaire") or "",
        "validatedAt": row.get("validated_at") or "",
    })
    return ticket


def _fiche_to_db_row(ticket_id, fiche):
    return {
        "ticket_id": ticket_id,
        "longueur": _as_text(fiche.get("longueur")),
        "largeur": _as_text(fiche.get("largeur")),
        "hauteur": _as_text(fiche.get("hauteur")),
        "dimensions_ext": _as_text(fiche.get("dimensionsExt")),
        "prix_achat": _as_text(fiche.get("prixAchat")),
        "prix_cession": _as_text(fiche.get("prixCession")),
        "type_caisse_fiche": _as_text(fiche.get("typeCaisseFiche")),
        "bilan_carbone": _as_text(fiche.get("bilanCarbone")),
        "poids": _as_text(fiche.get("poids")),
        "choix_caissier": _as_text(fiche.get("choixCaissier")),
        "localisation": _as_text(fiche.get("localisation")),
    }


def _fiche_from_db_row(row):
    return {
        "longueur": row.get("longueur") or "",
        "largeur": row.get("largeur") or "",
        "hauteur": row.get("hauteur") or "",
        "dimensionsExt": row.get("dimensions_ext") or "",
        "prixAchat": row.get("prix_achat") or "",
        "prixCession": row.get("prix_cession") or "",
        "typeCaisseFiche": row.get("type_caisse_fiche") or "",
        "bilanCarbone": row.get("bilan_carbone") or "",
        "poids": row.get("poids") or "",
        "choixCaissier": row.get("choix_caissier") or "",
        "localisation": row.get("localisation") or "",
    }


def _add_file_to_ticket(ticket, f):
    item = {
        "name": f.get("filename") or "",
        "size": f.get("size") or 0,
        "path": f.get("storage_path") or ""
    }
    if f.get("kind") == "gestionnaire":
        ticket.setdefault("managerSheets", []).append(item)
    else:
        ticket.setdefault("files", []).append(item)


def _attach_children(ticket):
    """Charge les enfants d'un seul ticket. Utilisé pour les actions ciblées."""
    tid = ticket.get("id")
    if not tid:
        return ticket

    safe_tid = urllib.parse.quote(tid, safe='')

    fiches = supabase_rest_request(
        "GET",
        "fiches",
        f"select=*&ticket_id=eq.{safe_tid}&limit=1"
    ) or []
    if fiches:
        ticket["fiche"] = _fiche_from_db_row(fiches[0])

    rows = supabase_rest_request(
        "GET",
        "ticket_files",
        f"select=*&ticket_id=eq.{safe_tid}&order=uploaded_at.asc"
    ) or []

    ticket["files"] = []
    ticket["managerSheets"] = []
    for f in rows:
        _add_file_to_ticket(ticket, f)

    return ticket


def _chunks(values, size=100):
    for i in range(0, len(values), size):
        yield values[i:i + size]


def _in_filter(values):
    # Format PostgREST : in.(DEM-001,DEM-002). Les ids internes ne contiennent pas de virgule.
    return urllib.parse.quote(",".join(values), safe=",-_")


def list_tickets(status=None, limit=None):
    query = "select=*&order=created_at.desc"
    if status:
        query += "&status=eq." + urllib.parse.quote(status, safe='')
    if limit:
        query += "&limit=" + str(int(limit))

    rows = supabase_rest_request("GET", "tickets", query) or []
    tickets = [_ticket_from_db_row(row) for row in rows]

    by_id = {t.get("id"): t for t in tickets if t.get("id")}
    ids = list(by_id.keys())
    if not ids:
        return tickets

    # Initialisation des listes pour éviter les champs absents côté interface
    for t in tickets:
        t["files"] = []
        t["managerSheets"] = []

    # Chargement groupé des fiches : au lieu de 1 requête par ticket
    for part in _chunks(ids):
        fiches = supabase_rest_request(
            "GET",
            "fiches",
            "select=*&ticket_id=in.(" + _in_filter(part) + ")"
        ) or []
        for f in fiches:
            tid = f.get("ticket_id")
            if tid in by_id:
                by_id[tid]["fiche"] = _fiche_from_db_row(f)

    # Chargement groupé des fichiers : au lieu de 1 requête par ticket
    for part in _chunks(ids):
        rows_files = supabase_rest_request(
            "GET",
            "ticket_files",
            "select=*&ticket_id=in.(" + _in_filter(part) + ")&order=uploaded_at.asc"
        ) or []
        for f in rows_files:
            tid = f.get("ticket_id")
            if tid in by_id:
                _add_file_to_ticket(by_id[tid], f)

    return tickets



_RECEPTION_MIGRATION_LOCK = threading.Lock()
_RECEPTION_MIGRATION_DONE = False


def _is_caisse_receptionnee(ticket):
    """
    Retourne True si la caisse a été réceptionnée.
    La réception est une information logistique indépendante du statut du ticket.
    """
    reception = ticket.get('reception') or {}
    if reception.get('receptionnee') is True:
        return True

    # Compatibilité avec les réceptions déjà enregistrées avant l'ajout du booléen.
    if reception.get('receptionnee_le'):
        return True

    return False


def migrate_caisses_avant_18_aout_2026():
    """
    Corrige l'ancienne logique qui utilisait 'Réceptionnée' comme statut de ticket.

    Règles :
      - les vrais statuts de fiche de caisse restent Demande créée / En cours / Terminé ;
      - toute fiche de caisse anciennement mise au statut 'Réceptionnée' repasse à 'Terminé' ;
      - les fiches terminées dont la date de mise à dispo est strictement antérieure
        au 18/08/2026 sont marquées comme réceptionnées dans ticket['reception'],
        sans modifier leur statut métier.
    """
    global _RECEPTION_MIGRATION_DONE

    with _RECEPTION_MIGRATION_LOCK:
        if _RECEPTION_MIGRATION_DONE:
            return

        cutoff = datetime(2026, 8, 18)
        updated = 0
        repaired_status = 0
        errors = 0

        try:
            all_tickets = list_tickets()
        except Exception as e:
            print(f"[RECEPTION MIGRATION] Lecture des tickets impossible : {e}")
            return

        for ticket in all_tickets:
            try:
                if ticket.get('module') != 'Fiche de caisse':
                    continue

                changed = False
                current_status = _as_text(ticket.get('status')).strip()

                # Répare les tickets affectés par l'ancienne logique.
                if current_status == 'Réceptionnée':
                    ticket['status'] = 'Terminé'
                    current_status = 'Terminé'
                    repaired_status += 1
                    changed = True

                raw_date = _as_text(ticket.get('dateEmballage')).strip()
                date_emballage = None
                if raw_date and raw_date != '-':
                    try:
                        date_emballage = datetime.fromisoformat(raw_date[:10])
                    except Exception:
                        date_emballage = None

                # Historique demandé : uniquement les caisses terminées avant le 18/08/2026.
                if (
                    current_status == 'Terminé'
                    and date_emballage is not None
                    and date_emballage < cutoff
                ):
                    reception = dict(ticket.get('reception') or {})
                    if not reception.get('receptionnee'):
                        reception['receptionnee'] = True
                        reception.setdefault('receptionnee_le', date_emballage.isoformat())
                        reception.setdefault('mode', 'migration_2026_08_18')
                        ticket['reception'] = reception
                        changed = True

                if not changed:
                    continue

                ticket['updatedAt'] = datetime.now().isoformat()

                supabase_rest_request(
                    "PATCH",
                    "tickets",
                    "id=eq." + urllib.parse.quote(ticket.get('id'), safe=''),
                    {
                        "status": ticket['status'],
                        "updated_at": ticket['updatedAt'],
                        "raw_json": ticket
                    },
                    prefer="return=minimal"
                )
                updated += 1

            except Exception as e:
                errors += 1
                print(f"[RECEPTION MIGRATION] Erreur {ticket.get('id')}: {e}")

        _RECEPTION_MIGRATION_DONE = True
        print(
            f"[RECEPTION MIGRATION] {updated} ticket(s) corrigé(s), "
            f"{repaired_status} statut(s) Réceptionnée -> Terminé, "
            f"{errors} erreur(s)"
        )


def next_id(prefix):
    safe_prefix = urllib.parse.quote(prefix + '-*', safe='*-')
    rows = supabase_rest_request(
        "GET",
        "tickets",
        f"select=id&id=like.{safe_prefix}&order=id.desc&limit=5000"
    ) or []
    nums = []
    for row in rows:
        try:
            nums.append(int(str(row.get("id", "")).split('-')[1]))
        except Exception:
            pass
    mx = max(nums) if nums else 0
    return f"{prefix}-{mx+1:03d}"


def save_ticket(ticket):
    if not ticket.get("id"):
        raise RuntimeError("Ticket sans ID")

    ticket.setdefault("updatedAt", datetime.now().isoformat())

    # Upsert du ticket principal
    supabase_rest_request(
        "POST",
        "tickets",
        "on_conflict=id",
        [_ticket_to_db_row(ticket)],
        prefer="resolution=merge-duplicates,return=minimal"
    )

    ticket_id = ticket.get("id")
    safe_tid = urllib.parse.quote(ticket_id, safe='')

    # Fiche gestionnaire
    fiche = ticket.get("fiche") or {}
    if fiche:
        supabase_rest_request(
            "POST",
            "fiches",
            "on_conflict=ticket_id",
            [_fiche_to_db_row(ticket_id, fiche)],
            prefer="resolution=merge-duplicates,return=minimal"
        )
    else:
        supabase_rest_request("DELETE", "fiches", f"ticket_id=eq.{safe_tid}", prefer="return=minimal")

    # Fichiers : on remplace la liste associée au ticket
    supabase_rest_request("DELETE", "ticket_files", f"ticket_id=eq.{safe_tid}", prefer="return=minimal")

    file_rows = []
    for fs in ticket.get("files") or []:
        if fs and fs.get("name"):
            file_rows.append({
                "ticket_id": ticket_id,
                "kind": "demandeur",
                "filename": _as_text(fs.get("name")),
                "size": fs.get("size") or 0,
                "storage_path": _as_text(fs.get("path")),
            })

    manager_sheets = list(ticket.get("managerSheets") or [])
    legacy = ticket.get("managerSheet")
    if legacy and isinstance(legacy, dict) and legacy.get("name"):
        if not any(x.get("name") == legacy.get("name") for x in manager_sheets):
            manager_sheets.append(legacy)

    for fs in manager_sheets:
        if fs and fs.get("name"):
            file_rows.append({
                "ticket_id": ticket_id,
                "kind": "gestionnaire",
                "filename": _as_text(fs.get("name")),
                "size": fs.get("size") or 0,
                "storage_path": _as_text(fs.get("path")),
            })

    if file_rows:
        supabase_rest_request("POST", "ticket_files", "", file_rows, prefer="return=minimal")

    print("[SUPABASE DB] Ticket sauvegardé", ticket_id)


def load_ticket(ticket_id):
    safe_tid = urllib.parse.quote(ticket_id, safe='')
    rows = supabase_rest_request("GET", "tickets", f"select=*&id=eq.{safe_tid}&limit=1") or []
    if not rows:
        return None
    return _attach_children(_ticket_from_db_row(rows[0]))


# -----------------------------------------------------------------------------
# Demande d'enlevement - analyse automatique des bons PDF
# -----------------------------------------------------------------------------
def _extract_enlevement_pdf_text(pdf_bytes):
    """Extrait le texte d'un bon d'enlevement natif ou scanne."""
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise RuntimeError(
            "Le module pypdf n'est pas installe. Ajoute pypdf a requirements.txt."
        ) from e

    try:
        reader = PdfReader(BytesIO(pdf_bytes))
    except Exception as e:
        raise ValueError(f"PDF illisible : {e}")

    pages_text = []
    for page in reader.pages:
        try:
            pages_text.append(page.extract_text() or "")
        except Exception:
            pages_text.append("")

    native_text = "\n".join(pages_text).strip()
    text = native_text
    ocr_used = False

    if len(native_text) < 120:
        print("[ENLEVEMENT OCR] Texte natif insuffisant, lancement de l'OCR")
        try:
            from pdf2image import convert_from_bytes
            import pytesseract
        except Exception as e:
            raise RuntimeError(
                "OCR indisponible. Verifie pdf2image, pytesseract, Pillow, "
                "tesseract-ocr et poppler-utils."
            ) from e

        with _RECEPTION_OCR_LOCK:
            try:
                images = convert_from_bytes(
                    pdf_bytes,
                    dpi=300,
                    grayscale=True,
                    thread_count=1
                )
            except Exception as e:
                raise RuntimeError(
                    "Impossible de convertir le PDF en image pour l'OCR."
                ) from e

            ocr_pages = []
            total_pages = len(images)

            for index, image in enumerate(images, start=1):
                try:
                    # PSM 4 respecte mieux les blocs/colonnes du bon interne.
                    page_text = pytesseract.image_to_string(
                        image,
                        lang="fra",
                        config="--psm 4"
                    )
                    # Secours pour les scans atypiques.
                    if len((page_text or "").strip()) < 120:
                        page_text = pytesseract.image_to_string(
                            image,
                            lang="fra",
                            config="--psm 11"
                        )
                except Exception as e:
                    raise RuntimeError("Echec OCR Tesseract.") from e

                print(f"[ENLEVEMENT OCR] Page {index}/{total_pages} analysee")
                ocr_pages.append(page_text or "")
                try:
                    image.close()
                except Exception:
                    pass

            images.clear()
            text = "\n".join(ocr_pages).strip()

        ocr_used = True

    if not text:
        raise ValueError("Aucun texte exploitable trouve dans le bon d'enlevement.")

    print(f"[ENLEVEMENT OCR] {len(text)} caracteres exploitables")
    return text, len(reader.pages), ocr_used


def _clean_ocr_line(value):
    value = _as_text(value)
    value = value.replace("\u00a0", " ")
    value = value.replace("–", "-").replace("—", "-")
    value = value.replace("：", ":")
    return re.sub(r"\s+", " ", value).strip(" \t|")


def _enlevement_lines(text):
    return [
        _clean_ocr_line(line)
        for line in _as_text(text).replace("\r", "").splitlines()
        if _clean_ocr_line(line)
    ]


def _value_after_label(lines, labels, stop_labels=None):
    stop_labels = stop_labels or []
    all_labels = list(labels) + list(stop_labels)

    for i, line in enumerate(lines):
        for label in labels:
            m = re.search(label, line, re.I)
            if not m:
                continue

            rest = line[m.end():].strip(" :-|")
            if rest:
                cuts = []
                for other in all_labels:
                    mo = re.search(other, rest, re.I)
                    if mo and mo.start() > 0:
                        cuts.append(mo.start())
                if cuts:
                    rest = rest[:min(cuts)].strip(" :-|")
                if rest:
                    return rest

            for offset in (1, 2):
                if i + offset < len(lines):
                    candidate = lines[i + offset].strip()
                    if candidate and not any(re.search(x, candidate, re.I) for x in all_labels):
                        return candidate

    return ""


def _normalise_enlevement_date(value):
    value = _clean_ocr_line(value)
    m = re.search(r"\b([0-3]?\d)[/.\-]([01]?\d)[/.\-](\d{2}|\d{4})\b", value)
    if not m:
        return ""
    day = int(m.group(1))
    month = int(m.group(2))
    year = int(m.group(3))
    if year < 100:
        year += 2000
    try:
        return datetime(year, month, day).strftime("%d/%m/%Y")
    except ValueError:
        return ""



def _extract_enlevement_items(instructions_text):
    """
    Extrait les articles de la zone Instructions avec une logique générique.

    Objectif :
      - ne plus dépendre d'un format précis de référence ;
      - accepter LDV_1047, ABC-123, 750012MW02, 960130M 04, etc. ;
      - utiliser le reste de la ligne comme désignation ;
      - éviter les faux positifs (dates, téléphones, dimensions, phrases générales).
    """
    lines = _enlevement_lines(instructions_text)
    items = []
    seen_refs = set()

    noise_patterns = [
        r"^\s*$",
        r"^\s*(?:merci|rappel|vous\s+pouvez|merci\s+de|storage|instruction|assur[eé]|valeur|observation|service)\b",
        r"^\s*(?:date|heure|notes?|assign[eé]|v[eé]hicules?)\s*:?\s*$",
    ]

    def is_noise(line):
        return any(re.search(p, line, re.I) for p in noise_patterns)

    def is_date_like(value):
        return bool(re.fullmatch(r"[0-3]?\d[/.\-][01]?\d[/.\-](?:\d{2}|\d{4})", value.strip()))

    def is_phone_like(value):
        compact = re.sub(r"[\s.\-()]+", "", value)
        return bool(re.fullmatch(r"(?:\+33|0)\d{9,10}", compact))

    def is_dimension_like(value):
        return bool(re.fullmatch(
            r"\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?"
            r"(?:\s*[xX×]\s*\d+(?:[.,]\d+)?)?\s*(?:cm|mm|m)?",
            value.strip(),
            re.I
        ))

    def normalise_reference(ref):
        ref = _clean_ocr_line(ref).strip(" :;,.|")
        ref = re.sub(r"\s+", " ", ref)

        # Cas fréquent OCR : "960130M 04" -> "960130M04".
        # On fusionne uniquement si les deux blocs ressemblent à une même référence.
        parts = ref.split()
        if len(parts) == 2:
            a, b = parts
            if (
                re.search(r"[A-Za-z]", a)
                and re.search(r"\d", a)
                and re.fullmatch(r"[A-Za-z0-9]{1,4}", b)
            ):
                ref = a + b

        # Les références métier ne doivent pas garder les espaces internes.
        ref = ref.replace(" ", "")
        ref = re.sub(r"^[^A-Za-z0-9]+|[^A-Za-z0-9_-]+$", "", ref)
        return ref

    def extract_dimensions(line):
        m = re.search(
            r"(?:Dims?\.?\s*[:.-]?\s*)?"
            r"(\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?"
            r"(?:\s*[xX×]\s*\d+(?:[.,]\d+)?)?\s*(?:cm|mm|m)?)",
            line,
            re.I
        )
        return _clean_ocr_line(m.group(1)) if m else ""

    def add_item(reference, description="", dimensions="", qty=""):
        ref = normalise_reference(reference)
        if len(ref) < 3:
            return
        if is_date_like(ref) or is_phone_like(ref) or is_dimension_like(ref):
            return

        # Une vraie référence doit contenir au moins une lettre et un chiffre,
        # ou bien comporter un séparateur structurant (_ / -).
        if not (
            (re.search(r"[A-Za-z]", ref) and re.search(r"\d", ref))
            or "_" in ref
            or "-" in ref
        ):
            return

        key = ref.upper()
        if key in seen_refs:
            return
        seen_refs.add(key)

        designation = _clean_ocr_line(description).strip(" -:;,.")
        quantity = _clean_ocr_line(qty)

        # Si aucune quantité n'est donnée explicitement, on considère 1 article.
        if not quantity:
            quantity = "1"

        items.append({
            "reference": ref,
            "quantite": quantity,
            "designation": designation,
            "dimensions": _clean_ocr_line(dimensions),
        })

    # 1) Cas explicite REF / REFERENCE.
    for line in lines:
        m = re.search(
            r"\bREF(?:ERENCE)?\s*[:.=\-]?\s*([A-Za-z0-9][A-Za-z0-9 _/-]{2,30})",
            line,
            re.I
        )
        if not m:
            continue

        raw = _clean_ocr_line(m.group(1))
        # Coupe la référence au premier gros séparateur ou début clair de désignation.
        tokens = raw.split()
        ref_tokens = []
        for token in tokens[:3]:
            if re.fullmatch(r"[A-Za-z0-9_-]+", token):
                ref_tokens.append(token)
            else:
                break
        raw_ref = " ".join(ref_tokens) if ref_tokens else raw

        designation = line[m.end():].strip(" :-|")
        add_item(raw_ref, designation, extract_dimensions(line))

    # 2) Références structurées avec _ ou - n'importe où dans la ligne.
    for line in lines:
        if is_noise(line):
            continue
        for m in re.finditer(
            r"\b([A-Za-z0-9]{1,20}[_-][A-Za-z0-9_-]{2,30})\b",
            line
        ):
            ref = m.group(1)
            designation = (line[:m.start()] + " " + line[m.end():]).strip(" :-|")
            # Si la ligne commence par la référence, la suite est la désignation.
            if not line[:m.start()].strip():
                designation = line[m.end():].strip(" :-|")
            add_item(ref, designation, extract_dimensions(line))

    # 3) Détection générique du début de ligne.
    #    On examine les 1 à 3 premiers blocs et on choisit le meilleur candidat.
    for line in lines:
        if is_noise(line):
            continue

        # Retire une éventuelle quantité en début de ligne : "1 750012MW02 ..."
        qty = ""
        work = line
        mq = re.match(r"^\s*(\d{1,3})\s+[xX]?\s+(.+)$", work)
        if mq:
            qty = mq.group(1)
            work = mq.group(2).strip()

        tokens = work.split()
        if len(tokens) < 2:
            continue

        candidates = []

        # Candidat 1 token : 750012MW02, LDV_1047, ABC-123
        candidates.append((tokens[0], 1))

        # Candidat 2 tokens : 960130M 04
        if len(tokens) >= 3:
            candidates.append((tokens[0] + " " + tokens[1], 2))

        # Candidat 3 tokens, très rare mais toléré.
        if len(tokens) >= 4:
            candidates.append((tokens[0] + " " + tokens[1] + " " + tokens[2], 3))

        best = None
        best_score = -999

        for cand, used in candidates:
            ref = normalise_reference(cand)
            if len(ref) < 4 or len(ref) > 30:
                continue
            if is_date_like(ref) or is_phone_like(ref) or is_dimension_like(ref):
                continue

            score = 0

            has_letter = bool(re.search(r"[A-Za-z]", ref))
            has_digit = bool(re.search(r"\d", ref))

            if has_letter and has_digit:
                score += 5
            if "_" in ref or "-" in ref:
                score += 3
            if 5 <= len(ref) <= 18:
                score += 2
            if re.match(r"^[A-Za-z0-9]", ref):
                score += 1
            if used == 1:
                score += 1

            # Pénalise les mots ordinaires.
            if ref.isalpha():
                score -= 8
            if re.fullmatch(r"\d+", ref):
                score -= 8

            # Il doit rester une désignation crédible après la référence.
            remaining = tokens[used:]
            designation = " ".join(remaining).strip()
            if len(designation) >= 4 and re.search(r"[A-Za-zÀ-ÿ]", designation):
                score += 4
            else:
                score -= 4

            if score > best_score:
                best_score = score
                best = (ref, used, designation)

        if best and best_score >= 7:
            ref, used, designation = best
            dims = extract_dimensions(line)
            add_item(ref, designation, dims, qty)

    print(f"[ENLEVEMENT ITEMS] {len(items)} article(s) detecte(s): {[x['reference'] for x in items]}")
    return items

def _extract_instructions_block(clean_text):
    """Isole la zone Instructions, y compris si le titre est légèrement déformé par l'OCR."""
    m = re.search(r"\bInstr(?:uctions?|uctions|uction)\b", clean_text, re.I)
    if not m:
        return ""

    tail = clean_text[m.end():]
    stop = re.search(
        r"\b(?:Assur[eé]\s+par|OBSERVATIONS?(?:\s+ou\s+R[EÉ]SERVES?)?|"
        r"Valeur\s+assur[eé]e|Signature|Heure\s+d['’]?arriv[eé]e|Heure\s+de\s+d[eé]part)\b",
        tail,
        re.I
    )
    block = tail[:stop.start()] if stop else tail[:3000]
    return block.strip()


def _slice_columns_by_headers(raw_lines, header_index, headers):
    """
    Découpe les lignes sous une ligne d'en-têtes en colonnes selon la position
    horizontale de chaque libellé. Utile pour les tableaux OCR du bon interne.
    """
    header_line = raw_lines[header_index]
    positions = []
    for name, pattern in headers:
        m = re.search(pattern, header_line, re.I)
        if m:
            positions.append((name, m.start()))
    positions.sort(key=lambda x: x[1])

    if len(positions) < 2:
        return {}

    bounds = {}
    for i, (name, pos) in enumerate(positions):
        if i == 0:
            left = 0
        else:
            prev_pos = positions[i - 1][1]
            left = (prev_pos + pos) // 2
        if i + 1 < len(positions):
            right = (pos + positions[i + 1][1]) // 2
        else:
            right = None
        bounds[name] = (left, right)

    columns = {name: [] for name, _ in positions}
    for line in raw_lines[header_index + 1:]:
        for name, (left, right) in bounds.items():
            piece = line[left:right].strip() if right is not None else line[left:].strip()
            if piece:
                columns[name].append(_clean_ocr_line(piece))
    return columns


def _extract_programme_chantier(clean_text):
    """
    Lit la zone 'Programme du chantier' comme un tableau vertical :
    les valeurs sont sous Date:, Heure:, Service:, Notes:, Assigné à:, Véhicules:.
    """
    result = {
        "date_enlevement": "",
        "heure_enlevement": "",
        "service": "",
        "notes": "",
        "assigne_a": "",
        "vehicules": "",
    }

    raw_lines = _as_text(clean_text).replace("\r", "").splitlines()
    start_idx = None
    end_idx = len(raw_lines)

    for i, line in enumerate(raw_lines):
        if re.search(r"Programme\s+du\s+chantier", line, re.I):
            start_idx = i
            break
    if start_idx is None:
        return result

    for i in range(start_idx + 1, len(raw_lines)):
        if re.search(r"\bInstr(?:uctions?|uction)\b", raw_lines[i], re.I):
            end_idx = i
            break

    zone = raw_lines[start_idx + 1:end_idx]

    header_idx_local = None
    for i, line in enumerate(zone):
        # On cherche la vraie ligne d'en-têtes du tableau.
        if re.search(r"\bDate\s*:", line, re.I) and re.search(r"\bNotes?\s*:", line, re.I):
            header_idx_local = i
            break

    if header_idx_local is not None:
        headers = [
            ("date", r"\bDate\s*:"),
            ("heure", r"\bHeure\s*:"),
            ("service", r"\bService\s*:"),
            ("notes", r"\bNotes?\s*:"),
            ("assigne", r"\bAssign[eé]\s+[aà]\s*:"),
            ("vehicules", r"\bV[eé]hicules?\s*:"),
        ]
        cols = _slice_columns_by_headers(zone, header_idx_local, headers)

        def first_value(name):
            vals = [v for v in cols.get(name, []) if v and not re.search(r"^[|:_-]+$", v)]
            return vals[0] if vals else ""

        result["date_enlevement"] = _normalise_enlevement_date(first_value("date"))
        result["heure_enlevement"] = first_value("heure")
        result["service"] = first_value("service")
        result["notes"] = first_value("notes")
        result["assigne_a"] = first_value("assigne")
        result["vehicules"] = first_value("vehicules")

    # Secours ciblé uniquement dans la zone Programme du chantier.
    zone_text = "\n".join(zone)
    if not result["date_enlevement"]:
        result["date_enlevement"] = _normalise_enlevement_date(zone_text)

    # Si les colonnes OCR sont mal alignées, essaie les valeurs situées
    # immédiatement sous les libellés sur des lignes séparées.
    cleaned_zone = _enlevement_lines(zone_text)
    if not result["notes"]:
        result["notes"] = _value_after_label(
            cleaned_zone,
            [r"\bNotes?\b"],
            [r"Assign[eé]\s+[aà]", r"V[eé]hicules?", r"Service", r"Instructions?"]
        )
    if not result["heure_enlevement"]:
        result["heure_enlevement"] = _value_after_label(
            cleaned_zone, [r"\bHeure\b"], [r"Service", r"Notes?", r"Assign[eé]\s+[aà]"]
        )
    if not result["service"]:
        result["service"] = _value_after_label(
            cleaned_zone, [r"\bService\b"], [r"Notes?", r"Assign[eé]\s+[aà]", r"V[eé]hicules?"]
        )

    return result


def _extract_contact_blocks(clean_text):
    """Extrait les blocs Depuis / À en conservant toutes les informations dans l'adresse."""
    result = {
        "adresse_depart": "",
        "adresse_destination": "",
    }

    raw_lines = _as_text(clean_text).replace("\r", "").splitlines()
    header_idx = None
    for i, line in enumerate(raw_lines):
        if re.search(r"\bDepuis\s*:", line, re.I) and re.search(r"(?:\bA\s*:|\bÀ\s*:)", line, re.I):
            header_idx = i
            break
    if header_idx is None:
        return result

    end_idx = len(raw_lines)
    for i in range(header_idx + 1, len(raw_lines)):
        if re.search(r"Programme\s+du\s+chantier", raw_lines[i], re.I):
            end_idx = i
            break

    cols = _slice_columns_by_headers(
        raw_lines[:end_idx], header_idx,
        [("depart", r"\bDepuis\s*:"), ("destination", r"(?:\bA\s*:|\bÀ\s*:)")]
    )

    def full_block(values):
        vals = [_clean_ocr_line(v) for v in values if _clean_ocr_line(v)]
        return "\n".join(vals)

    result["adresse_depart"] = full_block(cols.get("depart", []))
    result["adresse_destination"] = full_block(cols.get("destination", []))
    return result



def _spatial_group_lines(words, y_tolerance=14):
    """Regroupe des mots OCR par lignes en conservant l'ordre horizontal."""
    if not words:
        return []
    words = sorted(words, key=lambda w: (w["cy"], w["left"]))
    lines = []
    for word in words:
        target = None
        for line in lines:
            if abs(word["cy"] - line["cy"]) <= y_tolerance:
                target = line
                break
        if target is None:
            target = {"cy": word["cy"], "words": []}
            lines.append(target)
        target["words"].append(word)
        target["cy"] = sum(w["cy"] for w in target["words"]) / len(target["words"])

    result = []
    for line in sorted(lines, key=lambda x: x["cy"]):
        ordered = sorted(line["words"], key=lambda w: w["left"])
        text = _clean_ocr_line(" ".join(w["text"] for w in ordered))
        if text:
            result.append({"cy": line["cy"], "text": text, "words": ordered})
    return result


def _extract_enlevement_spatial(pdf_bytes):
    """
    Lecture spatiale de la première page du bon d'enlèvement.

    Contrairement à la lecture texte classique, cette méthode utilise les vraies
    coordonnées OCR des mots. Elle sert uniquement aux zones dont la mise en page
    est en colonnes : Adresses (Depuis / À) et Programme du chantier.
    """
    result = {
        "adresse_depart": "",
        "adresse_destination": "",
        "date_enlevement": "",
        "notes": "",
    }

    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except Exception as e:
        print(f"[ENLEVEMENT SPATIAL] OCR spatial indisponible: {e}")
        return result

    try:
        # 200 dpi suffit pour localiser correctement les colonnes et limite la charge mémoire.
        images = convert_from_bytes(
            pdf_bytes,
            dpi=200,
            grayscale=True,
            first_page=1,
            last_page=1,
            thread_count=1,
        )
        if not images:
            return result
        image = images[0]
        data = pytesseract.image_to_data(
            image,
            lang="fra",
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )
    except Exception as e:
        print(f"[ENLEVEMENT SPATIAL] Echec lecture spatiale: {e}")
        return result
    finally:
        try:
            for img in locals().get('images', []) or []:
                img.close()
        except Exception:
            pass

    words = []
    count = len(data.get("text", []))
    for i in range(count):
        text = _clean_ocr_line(data["text"][i])
        if not text:
            continue
        try:
            conf = float(data.get("conf", [0] * count)[i])
        except Exception:
            conf = 0
        if conf < 20:
            continue
        left = int(data["left"][i])
        top = int(data["top"][i])
        width = int(data["width"][i])
        height = int(data["height"][i])
        words.append({
            "text": text,
            "left": left,
            "top": top,
            "right": left + width,
            "bottom": top + height,
            "cx": left + width / 2,
            "cy": top + height / 2,
        })

    if not words:
        return result

    page_width = max(w["right"] for w in words)
    page_mid = page_width * 0.50

    def find_word(pattern, y_min=0, y_max=10**9, x_min=0, x_max=10**9):
        for w in sorted(words, key=lambda z: (z["top"], z["left"])):
            if not (y_min <= w["cy"] <= y_max and x_min <= w["cx"] <= x_max):
                continue
            if re.search(pattern, w["text"], re.I):
                return w
        return None

    # ------------------------------------------------------------------
    # Zone Adresses : sépare physiquement la moitié gauche (Depuis) et droite (À).
    # ------------------------------------------------------------------
    depuis = find_word(r"^Depuis:?$")
    programme = find_word(r"^Programme$", y_min=(depuis["cy"] if depuis else 0))
    if depuis and programme:
        y_start = depuis["bottom"] + 8
        y_end = programme["top"] - 10
        zone_words = [w for w in words if y_start <= w["cy"] <= y_end]

        left_lines = _spatial_group_lines([w for w in zone_words if w["cx"] < page_mid])
        right_lines = _spatial_group_lines([w for w in zone_words if w["cx"] >= page_mid])

        def clean_address_lines(lines):
            cleaned = []
            for line in lines:
                txt = _clean_ocr_line(line["text"])
                if not txt:
                    continue
                if re.fullmatch(r"(?:Depuis|A|À)\s*:?", txt, re.I):
                    continue
                cleaned.append(txt)
            return cleaned

        left_clean = clean_address_lines(left_lines)
        right_clean = clean_address_lines(right_lines)

        if left_clean:
            result["adresse_depart"] = "\n".join(left_clean)
        if right_clean:
            result["adresse_destination"] = "\n".join(right_clean)

    # ------------------------------------------------------------------
    # Programme du chantier : valeur située SOUS le libellé.
    # ------------------------------------------------------------------
    programme = programme or find_word(r"^Programme$")
    instructions = find_word(r"^Instructions?$", y_min=(programme["cy"] if programme else 0))
    if programme and instructions:
        y_top = programme["bottom"] + 8
        y_bottom = instructions["top"] - 8
        prog_words = [w for w in words if y_top <= w["cy"] <= y_bottom]

        # Repère les colonnes à partir des libellés Date / Heure / Service / Notes / Assigné.
        labels = {}
        for name, pattern in [
            ("date", r"^Date:?$"),
            ("heure", r"^Heure:?$"),
            ("service", r"^Service:?$"),
            ("notes", r"^Notes?:?$"),
            ("assigne", r"^Assign[eé]$"),
        ]:
            matches = [w for w in prog_words if re.search(pattern, w["text"], re.I)]
            if matches:
                labels[name] = min(matches, key=lambda w: w["top"])

        ordered = sorted((w["left"], name, w) for name, w in labels.items())
        bounds = {}
        for idx, (left, name, word) in enumerate(ordered):
            x0 = 0 if idx == 0 else (ordered[idx - 1][0] + left) / 2
            x1 = page_width if idx + 1 == len(ordered) else (left + ordered[idx + 1][0]) / 2
            bounds[name] = (x0, x1, word["bottom"] + 4)

        if "date" in bounds:
            x0, x1, y0 = bounds["date"]
            vals = [w for w in prog_words if x0 <= w["cx"] < x1 and w["cy"] >= y0]
            date_text = " ".join(x["text"] for x in sorted(vals, key=lambda z: (z["top"], z["left"])))
            result["date_enlevement"] = _normalise_enlevement_date(date_text)

        if "notes" in bounds:
            x0, x1, y0 = bounds["notes"]
            vals = [w for w in prog_words if x0 <= w["cx"] < x1 and w["cy"] >= y0]
            note_lines = _spatial_group_lines(vals)
            if note_lines:
                note = note_lines[0]["text"]
                # Sur certains scans, le trait oblique sous la note transforme SS# en SSH#.
                note = re.sub(r"\bSSH#(\d+)\b", r"SS#\1", note, flags=re.I)
                result["notes"] = _clean_ocr_line(note)

    print(
        "[ENLEVEMENT SPATIAL] "
        f"depart={result.get('adresse_depart') or '-'} | "
        f"destination={result.get('adresse_destination') or '-'} | "
        f"date={result.get('date_enlevement') or '-'} | notes={result.get('notes') or '-'}"
    )
    return result


def _extract_enlevement_pdf(pdf_bytes):
    """Analyse un bon d'enlevement et retourne les donnees utiles au planning reception."""
    text, page_count, ocr_used = _extract_enlevement_pdf_text(pdf_bytes)
    clean_text = text.replace("\r", "")

    # Toute la partie du bon à partir de "Assuré par" est administrative :
    # elle ne doit jamais alimenter les champs, instructions ou articles.
    cutoff = re.search(r"\bAssur[eé]\s+par\b", clean_text, re.I)
    if cutoff:
        clean_text = clean_text[:cutoff.start()].rstrip()

    lines = _enlevement_lines(clean_text)

    label_numero = [
        r"Num[eé]ro\s+de\s+r[eé]f[eé]r(?:ence)?",
        r"N[°ºo]\s*de\s*r[eé]f[eé]rence",
        r"R[eé]f[eé]rence\s+du\s+bon",
    ]
    common_stops = [
        r"Client", r"Coordinateur", r"Exhibition", r"Programme\s+du\s+chantier",
        r"Instructions?", r"Adresse", r"Service"
    ]

    numero_bon = _value_after_label(lines, label_numero, common_stops)
    if numero_bon:
        m = re.search(r"\b([A-Za-z0-9][A-Za-z0-9_-]{3,})\b", numero_bon)
        numero_bon = m.group(1) if m else ""

    client = _value_after_label(
        lines,
        [r"\bClient\b"],
        [r"Coordinateur", r"Exhibition", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )
    coordinateur = _value_after_label(
        lines,
        [r"Coordinateur"],
        [r"Client", r"Exhibition", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )
    exhibition = _value_after_label(
        lines,
        [r"Exhibition"],
        [r"Client", r"Coordinateur", r"Programme\s+du\s+chantier", r"Adresse", r"Service"]
    )

    if client:
        client = re.split(r"\bExhibition\b", client, maxsplit=1, flags=re.I)[0].strip(" :-|")
    if coordinateur:
        coordinateur = re.split(r"\b(?:Client|Exhibition)\b", coordinateur, maxsplit=1, flags=re.I)[0].strip(" :-|")
    if exhibition:
        exhibition = re.split(r"\b(?:Client|Coordinateur)\b", exhibition, maxsplit=1, flags=re.I)[0].strip(" :-|")

    programme = _extract_programme_chantier(clean_text)
    spatial = _extract_enlevement_spatial(pdf_bytes)
    date_enlevement = spatial.get("date_enlevement") or programme.get("date_enlevement", "")

    instructions = _extract_instructions_block(clean_text)
    # Les articles doivent provenir uniquement de la zone Instructions.
    items = _extract_enlevement_items(instructions) if instructions else []

    contact_data = _extract_contact_blocks(clean_text)
    # Les coordonnées OCR réelles sont prioritaires pour les deux colonnes d'adresses.
    for key in ("adresse_depart", "adresse_destination"):
        if spatial.get(key):
            contact_data[key] = spatial[key]
    # Ces informations ne sont pas utilisées dans la fiche réception.
    contact_data.pop("telephone_depart", None)
    contact_data.pop("telephone_destination", None)
    notes = spatial.get("notes") or programme.get("notes", "")

    display_name = " - ".join(
        x for x in [_clean_ocr_line(client), _clean_ocr_line(numero_bon)] if x
    )

    result = {
        "numero_bon": numero_bon,
        "client": client,
        "display_name": display_name,
        "coordinateur": coordinateur,
        "exhibition": exhibition,
        "date_enlevement": date_enlevement,
        "service": programme.get("service", ""),
        "assigne_a": programme.get("assigne_a", ""),
        "vehicules": programme.get("vehicules", ""),
        "notes": notes,
        "instructions": instructions,
        "items": items,
        "references": [x.get("reference") for x in items if x.get("reference")],
        "page_count": page_count,
        "ocr_used": ocr_used,
        "raw_text": clean_text,
        **contact_data,
    }

    print(
        "[ENLEVEMENT EXTRACTION] "
        f"bon={result.get('numero_bon') or '-'} "
        f"client={result.get('client') or '-'} "
        f"date={result.get('date_enlevement') or '-'} "
        f"refs={result.get('references') or []}"
    )
    return result


def _analyse_enlevement_ticket_background(ticket_id, pdf_bytes):
    """Analyse le bon apres creation du ticket, sans bloquer le demandeur."""
    try:
        print(f"[ENLEVEMENT] Analyse asynchrone demarree pour {ticket_id}")
        parsed = _extract_enlevement_pdf(pdf_bytes)
        ticket = load_ticket(ticket_id)
        if not ticket:
            print(f"[ENLEVEMENT] Ticket introuvable apres creation : {ticket_id}")
            return

        ticket["enlevement"] = {
            **parsed,
            # Nom visible côté réception. L'id ENL-xxx reste uniquement technique.
            "display_name": parsed.get("display_name") or " - ".join(
                x for x in [parsed.get("client"), parsed.get("numero_bon")] if x
            ),
            "analysis_status": "ready",
            "analysis_error": "",
            "analysed_at": datetime.now().isoformat(),
        }

        if parsed.get("client"):
            ticket["dossier"] = parsed["client"]
        if parsed.get("numero_bon"):
            ticket["ref"] = parsed["numero_bon"]
        if parsed.get("coordinateur"):
            ticket["chargeProjet"] = parsed["coordinateur"]
        if parsed.get("exhibition"):
            ticket["expo"] = parsed["exhibition"]
            ticket["objet"] = parsed["exhibition"]
        if parsed.get("date_enlevement"):
            try:
                dt = datetime.strptime(parsed["date_enlevement"], "%d/%m/%Y")
                ticket["dateRdv"] = dt.strftime("%Y-%m-%d")
            except ValueError:
                pass

        ticket["updatedAt"] = datetime.now().isoformat()
        save_ticket(ticket)
        try:
            _ensure_articles_for_ticket(ticket, save=True)
        except Exception as article_error:
            print(f"[ARTICLES] Attribution ESI impossible pour {ticket_id}: {article_error}")
        print(
            f"[ENLEVEMENT] Analyse terminee pour {ticket_id}: "
            f"{len(parsed.get('items') or [])} item(s), OCR={parsed.get('ocr_used')}"
        )
    except Exception as e:
        print(f"[ENLEVEMENT] Erreur analyse {ticket_id}: {e}")
        try:
            ticket = load_ticket(ticket_id)
            if ticket:
                current = ticket.get("enlevement") or {}
                current.update({
                    "analysis_status": "error",
                    "analysis_error": str(e),
                    "analysed_at": datetime.now().isoformat(),
                })
                ticket["enlevement"] = current
                ticket["updatedAt"] = datetime.now().isoformat()
                save_ticket(ticket)
        except Exception as save_error:
            print(f"[ENLEVEMENT] Impossible d'enregistrer l'erreur {ticket_id}: {save_error}")


# -----------------------------------------------------------------------------
# Réception caisserie - lecture des bordereaux PDF fournisseur
# -----------------------------------------------------------------------------
def _normalise_numero_caisse(value):
    """Normalise 01, 1, 1.0 -> 1 pour fiabiliser les rapprochements."""
    txt = _as_text(value).strip()
    if not txt:
        return ""
    try:
        return str(int(float(txt.replace(",", "."))))
    except Exception:
        return txt.lstrip("0") or "0"


def _extract_reception_pdf(pdf_bytes):
    """
    Extrait les informations utiles d'un bordereau PDF.

    1) Essaie d'abord l'extraction texte native avec pypdf.
    2) Si le PDF contient trop peu de texte, lance automatiquement un OCR.

    Format SECO actuellement reconnu :
      - BORDEREAU D'EXPEDITION N° 26400467 du 17/08/2026
      - V/Cde : 101138/01
    """
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise RuntimeError(
            "Le module pypdf n'est pas installé. Ajoute pypdf à requirements.txt."
        ) from e

    try:
        reader = PdfReader(BytesIO(pdf_bytes))
    except Exception as e:
        raise ValueError(f"PDF illisible : {e}")

    pages_text = []
    for page in reader.pages:
        try:
            pages_text.append(page.extract_text() or "")
        except Exception:
            pages_text.append("")

    text = "\n".join(pages_text).strip()
    ocr_used = False

    # Un scan image peut renvoyer une chaine vide ou quelques caracteres inutilisables.
    if len(text) < 50:
        print("[RECEPTION OCR] Texte natif insuffisant, lancement de l'OCR")
        try:
            from pdf2image import convert_from_bytes
            import pytesseract
        except Exception as e:
            raise RuntimeError(
                "OCR indisponible. Ajoute pdf2image, pytesseract et Pillow a requirements.txt, "
                "puis installe tesseract-ocr et poppler-utils sur Render."
            ) from e

        # Un seul OCR lourd à la fois par worker. Cela évite les pics mémoire
        # si le navigateur/proxy soumet plusieurs fois le même PDF.
        with _RECEPTION_OCR_LOCK:
            try:
                images = convert_from_bytes(
                    pdf_bytes,
                    dpi=200,
                    grayscale=True,
                    thread_count=1
                )
            except Exception as e:
                raise RuntimeError(
                    "Impossible de convertir le PDF en image pour l'OCR. "
                    "Verifie que poppler-utils est installe sur Render."
                ) from e

            ocr_pages = []
            total_pages = len(images)
            for index, image in enumerate(images, start=1):
                try:
                    page_text = pytesseract.image_to_string(
                        image,
                        lang="fra",
                        config="--psm 6"
                    )
                except Exception as e:
                    raise RuntimeError(
                        "Echec OCR Tesseract. Verifie que tesseract-ocr et la langue francaise sont installes."
                    ) from e
                print(f"[RECEPTION OCR] Page {index}/{total_pages} analysee")
                ocr_pages.append(page_text or "")
                try:
                    image.close()
                except Exception:
                    pass

            # Libère explicitement les images avant le rapprochement Supabase.
            images.clear()
            text = "\n".join(ocr_pages).strip()
        ocr_used = True
        print(f"[RECEPTION OCR] OCR termine, {len(text)} caracteres detectes")

    if not text:
        raise ValueError(
            "Aucun texte exploitable trouve dans le PDF, meme apres OCR."
        )

    # Numéro et date du bordereau.
    bl_numero = ""
    bl_date = ""
    m = re.search(
        r"BORDEREAU\s+D['’]EXPEDITION.*?N\s*[°ºo]?\s*([0-9]+)\s+du\s+([0-9]{2}/[0-9]{2}/[0-9]{4})",
        text,
        flags=re.IGNORECASE | re.DOTALL
    )
    if m:
        bl_numero = m.group(1)
        bl_date = m.group(2)
    else:
        # Secours, plus tolérant.
        m = re.search(r"N\s*[°ºo]?\s*([0-9]{6,})\s+du\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", text)
        if m:
            bl_numero = m.group(1)
            bl_date = m.group(2)

    # Extrait toutes les références V/Cde : dossier/numero.
    refs = []
    seen = set()
    # OCR peut lire "V/Cde" comme "ViCde", "VICde", "V Cde", etc.
    # On tolère donc un séparateur imparfait entre V et Cde.
    for dossier, numero in re.findall(
        r"V\s*[/|Il1i\-]?\s*Cde\s*:\s*([A-Za-z0-9_-]+)\s*/\s*([0-9]+)",
        text,
        flags=re.IGNORECASE
    ):
        dossier = dossier.strip()
        numero_norm = _normalise_numero_caisse(numero)
        key = (dossier, numero_norm)
        if key not in seen:
            seen.add(key)
            refs.append({
                "dossier": dossier,
                "numero": numero_norm,
                "numero_pdf": numero.strip()
            })

    if not refs:
        raise ValueError(
            "Aucune référence de caisse de type 'V/Cde : dossier/numéro' n'a été détectée, même après OCR."
        )

    return {
        "bl_numero": bl_numero,
        "bl_date": bl_date,
        "references": refs,
        "page_count": len(reader.pages),
        "ocr_used": ocr_used,
    }


def _match_reception_refs_to_tickets(references):
    """Rapproche les références du PDF avec les fiches de caisse ESI TICKETS."""
    all_tickets = list_tickets()
    candidates = [
        t for t in all_tickets
        if t.get("module") == "Fiche de caisse"
    ]

    by_key = {}
    for t in candidates:
        dossier = _as_text(t.get("dossier")).strip()
        numero = _normalise_numero_caisse(t.get("ref"))
        if dossier and numero:
            by_key.setdefault((dossier, numero), []).append(t)

    results = []
    for ref in references:
        key = (ref["dossier"], ref["numero"])
        matches = by_key.get(key, [])

        if len(matches) == 1:
            t = matches[0]
            fiche = t.get("fiche") or {}
            results.append({
                "found": True,
                "ambiguous": False,
                "ticket_id": t.get("id"),
                "dossier": t.get("dossier") or "",
                "ref": t.get("ref") or "",
                "charge_projet": t.get("chargeProjet") or "",
                "date_emballage": t.get("dateEmballage") or "",
                "localisation": fiche.get("localisation") or "",
                "status": t.get("status") or "",
                "receptionnee": _is_caisse_receptionnee(t),
            })
        elif len(matches) > 1:
            results.append({
                "found": False,
                "ambiguous": True,
                "dossier": ref["dossier"],
                "ref": ref["numero_pdf"],
                "error": f"{len(matches)} tickets correspondent à cette référence"
            })
        else:
            results.append({
                "found": False,
                "ambiguous": False,
                "dossier": ref["dossier"],
                "ref": ref["numero_pdf"],
                "error": "Ticket introuvable"
            })

    return results


# -----------------------------------------------------------------------------
# Référentiels métier : chargés de projet, clients, contacts
# -----------------------------------------------------------------------------
REFERENTIELS = {
    "project-managers": {
        "table": "project_managers",
        "allowed": ["nom", "email", "telephone", "actif"],
        "search": ["nom", "email", "telephone"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    },
    "clients": {
        "table": "clients",
        "allowed": ["nom", "adresse", "contact_nom", "contact_email", "contact_telephone", "actif"],
        "search": ["nom", "adresse", "contact_nom", "contact_email", "contact_telephone"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    },
    "contacts": {
        "table": "contacts",
        "allowed": ["nom", "email", "telephone", "client_nom", "fonction", "actif"],
        "search": ["nom", "email", "telephone", "client_nom", "fonction"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    }
}


def _referentiel_config(kind):
    cfg = REFERENTIELS.get(kind)
    if not cfg:
        abort(404)
    return cfg


def _clean_referentiel_payload(kind, data, partial=False):
    cfg = _referentiel_config(kind)
    data = data or {}
    payload = {}

    for field in cfg["allowed"]:
        if field in data:
            if field == "actif":
                payload[field] = bool(data.get(field))
            else:
                payload[field] = _as_text(data.get(field)).strip()

    if not partial:
        for field, value in cfg.get("defaults", {}).items():
            payload.setdefault(field, value)

        missing = [field for field in cfg.get("required", []) if not payload.get(field)]
        if missing:
            raise ValueError("Champ obligatoire manquant : " + ", ".join(missing))

    return payload


@app.route('/api/referentiels/<kind>', methods=['GET'])
def api_list_referentiel(kind):
    cfg = _referentiel_config(kind)
    q = (request.args.get('q') or '').strip()
    include_inactive = request.args.get('include_inactive') == '1'
    limit = request.args.get('limit') or '100'

    query = "select=*"
    if not include_inactive:
        query += "&actif=eq.true"
    if q:
        pattern = "*" + q.replace("*", "") + "*"
        parts = []
        for field in cfg["search"]:
            parts.append(f"{field}.ilike.{urllib.parse.quote(pattern, safe='*')}")
        query += "&or=(" + ",".join(parts) + ")"
    query += "&order=" + urllib.parse.quote(cfg.get("order", "nom.asc"), safe='.,')
    query += "&limit=" + urllib.parse.quote(str(limit), safe='')

    rows = supabase_rest_request("GET", cfg["table"], query) or []
    return jsonify(rows)


@app.route('/api/referentiels/<kind>', methods=['POST'])
def api_create_referentiel(kind):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    try:
        payload = _clean_referentiel_payload(kind, data, partial=False)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    try:
        rows = supabase_rest_request("POST", cfg["table"], "", [payload], prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/referentiels/<kind>/<item_id>', methods=['PUT'])
def api_update_referentiel(kind, item_id):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    try:
        payload = _clean_referentiel_payload(kind, data, partial=True)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    if not payload:
        return jsonify({'ok': False, 'error': 'Aucune donnée à modifier'}), 400

    safe_id = urllib.parse.quote(str(item_id), safe='')
    try:
        rows = supabase_rest_request("PATCH", cfg["table"], f"id=eq.{safe_id}", payload, prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/referentiels/<kind>/<item_id>/toggle', methods=['PATCH'])
def api_toggle_referentiel(kind, item_id):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    actif = bool(data.get('actif'))
    safe_id = urllib.parse.quote(str(item_id), safe='')
    try:
        rows = supabase_rest_request("PATCH", cfg["table"], f"id=eq.{safe_id}", {"actif": actif}, prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else {'id': item_id, 'actif': actif}})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# -----------------------------------------------------------------------------
# Notifications email Outlook / Microsoft 365
# -----------------------------------------------------------------------------
def _format_ticket_notification_subject(ticket):
    """Prépare l'objet métier du mail de notification, sans numéro interne de ticket."""
    module = ticket.get("module") or "Ticket"
    dossier = (ticket.get("dossier") or "").strip()
    ref = (ticket.get("ref") or "").strip()
    preteur = (ticket.get("preteur") or "").strip()
    projet = (ticket.get("expo") or ticket.get("objet") or "").strip()
    lieu_rdv = (ticket.get("lieuRdv") or "").strip()

    if module == "Fiche de caisse":
        suffix = " ".join([x for x in [dossier, ref] if x]).strip()
        return f"[ESI Tickets] Fiche de caisse terminée - {suffix}".strip()

    if module == "Demande de devis":
        suffix = " ".join([x for x in [dossier, projet] if x]).strip()
        return f"[ESI Tickets] Demande de devis terminée - {suffix}".strip()

    if module == "Demande Aller voir":
        suffix = " ".join([x for x in [dossier, lieu_rdv or projet] if x]).strip()
        return f"[ESI Tickets] Aller voir finalisé - {suffix}".strip()

    if module == "Avis d'arrivée":
        suffix = " ".join([x for x in [dossier, projet] if x]).strip()
        return f"[ESI Tickets] Avis d'arrivée terminé - {suffix}".strip()

    suffix = dossier or projet or ref
    return f"[ESI Tickets] Ticket terminé - {suffix}".strip()


def _find_project_manager_email(charge_projet):
    """Retrouve l'email du chargé de projet depuis le référentiel Supabase."""
    charge_projet = (charge_projet or "").strip()
    if not charge_projet or charge_projet == "-":
        return ""

    # Recherche exacte sur le nom enregistré dans le ticket.
    nom_encode = urllib.parse.quote(charge_projet, safe='')
    rows = supabase_rest_request(
        "GET",
        "project_managers",
        f"select=nom,email&nom=eq.{nom_encode}&limit=1"
    ) or []

    if rows and rows[0].get("email"):
        return (rows[0].get("email") or "").strip()

    # Secours : recherche souple si le nom contient une différence d'espace ou de casse.
    pattern = "*" + charge_projet.replace("*", "") + "*"
    pattern_encode = urllib.parse.quote(pattern, safe='*')
    rows = supabase_rest_request(
        "GET",
        "project_managers",
        f"select=nom,email&nom=ilike.{pattern_encode}&limit=1"
    ) or []

    if rows and rows[0].get("email"):
        return (rows[0].get("email") or "").strip()

    return ""


def envoyer_notification_fin_ticket(ticket):
    """Envoie un email au chargé de projet quand un ticket passe au statut Terminé."""
    try:
        print("[MAIL] Début envoi notification ticket terminé")

        charge_projet = (ticket.get("chargeProjet") or "").strip()
        print("[MAIL] Chargé projet =", charge_projet)

        email_dest = _find_project_manager_email(charge_projet)
        print("[MAIL] Destinataire =", email_dest or "NON TROUVÉ")

        if not email_dest:
            print(f"[MAIL] Aucun email trouvé pour le chargé de projet : {charge_projet}")
            return False

        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        smtp_user = os.getenv("SMTP_USER")
        smtp_password = os.getenv("SMTP_PASSWORD")

        if not smtp_host or not smtp_user or not smtp_password:
            print("[MAIL] Configuration SMTP manquante : vérifier SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD dans Render")
            return False

        ticket_id = ticket.get("id", "")
        module = ticket.get("module", "")
        dossier = ticket.get("dossier", "")
        ref = ticket.get("ref", "")
        projet = ticket.get("expo") or ticket.get("objet") or ""
        lieu_rdv = ticket.get("lieuRdv", "")
        date_rdv = ticket.get("dateRdv", "")
        heure_rdv = ticket.get("heureRdv", "")
        commentaire = ticket.get("commentaire", "")

        sujet = _format_ticket_notification_subject(ticket)

        corps = f"""Bonjour,

Le ticket suivant vient d'être terminé :

Numéro ticket : {ticket_id}
Type : {module}
Dossier / Client : {dossier}
Référence / N° caisse : {ref}
Projet / Expo : {projet}
Chargé de projet : {charge_projet}
Lieu RDV : {lieu_rdv}
Date RDV : {date_rdv} {heure_rdv}

Commentaire :
{commentaire or '-'}

Le document est disponible dans ESI Tickets.

Cordialement,
ESI Tickets
"""

        msg = MIMEText(corps, "plain", "utf-8")
        msg["Subject"] = sujet
        msg["From"] = smtp_user
        msg["To"] = email_dest

        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)

        print(f"[MAIL] Notification envoyée à {email_dest}")
        return True

    except Exception as e:
        print(f"[MAIL] Erreur envoi notification : {e}")
        return False


@app.route('/')
def index():
    return redirect(url_for('demandeur'))

@app.route('/demandeur')
def demandeur():
    return render_template('demandeur.html')

GESTIONNAIRE_ARTICLES_LIES_JS = r"""(function(){
'use strict';

const STYLE_ID='esi-caisse-articles-v2-style';
if(!document.getElementById(STYLE_ID)){
  const style=document.createElement('style');
  style.id=STYLE_ID;
  style.textContent=`
    .caisse-articles-panel{display:none;margin:14px 0 0;border:1px solid #b9dbea;background:linear-gradient(180deg,#f0f9ff,#fff);border-radius:16px;padding:14px}
    .caisse-articles-panel.show{display:block}
    .caisse-articles-title{font-size:11px;font-weight:900;color:#0284c7;text-transform:uppercase;letter-spacing:.055em;margin-bottom:9px}
    .caisse-articles-selected{display:flex;gap:7px;flex-wrap:wrap;align-items:center;min-height:30px}
    .caisse-article-chip{appearance:none;border:1px solid #7dd3fc;background:#e0f2fe;color:#075985;border-radius:999px;padding:7px 10px;font:inherit;font-size:12px;font-weight:800;cursor:pointer;display:inline-flex;gap:7px;align-items:center;max-width:100%}
    .caisse-article-chip:hover{background:#bae6fd}.caisse-article-chip .remove{font-size:15px;line-height:1;color:#0f2f4f}
    .caisse-articles-empty{font-size:12px;color:#64748b}
    .caisse-articles-search{display:grid;grid-template-columns:minmax(0,1fr) auto;gap:8px;margin-top:12px}
    .caisse-articles-search input{width:100%;border:1px solid #9bcbe7;background:#fff;border-radius:12px;padding:10px 12px;font:inherit;box-sizing:border-box}
    .caisse-articles-results{margin-top:10px;border:1px solid #dbe7f0;border-radius:12px;background:#fff;max-height:260px;overflow:auto}
    .caisse-articles-result-head,.caisse-articles-result{display:grid;grid-template-columns:38px minmax(110px,.75fr) minmax(165px,1.25fr);gap:8px;align-items:center;padding:9px 10px}
    .caisse-articles-result-head{position:sticky;top:0;background:#f8fafc;border-bottom:1px solid #dbe7f0;font-size:10px;font-weight:900;text-transform:uppercase;color:#64748b;z-index:1}
    .caisse-articles-result{border-bottom:1px solid #eef2f7;font-size:12px}.caisse-articles-result:last-child{border-bottom:0}.caisse-articles-result:hover{background:#f0f9ff}
    .caisse-articles-result input{width:17px;height:17px;cursor:pointer}.caisse-articles-result .ref{font-weight:850;color:#0f2f4f;overflow-wrap:anywhere}.caisse-articles-result .dos{color:#475569;overflow-wrap:anywhere}
    .caisse-articles-hint{font-size:10px;color:#64748b;margin-top:7px;line-height:1.35}
    .caisse-article-modal-backdrop{position:fixed;inset:0;background:rgba(15,23,42,.55);z-index:1600;display:none;align-items:center;justify-content:center;padding:18px}.caisse-article-modal-backdrop.show{display:flex}
    .caisse-article-modal{background:#f8fbfd;width:min(1000px,97vw);max-height:93vh;overflow:auto;border-radius:22px;box-shadow:0 26px 80px rgba(15,23,42,.28);border:1px solid #dbe7f0}
    .caisse-article-modal-head{position:sticky;top:0;z-index:3;background:#fff;border-bottom:1px solid #dbe7f0;padding:16px 18px;display:flex;justify-content:space-between;gap:14px;align-items:center}.caisse-article-modal-title{font-size:19px;font-weight:900;color:#0f2f4f}.caisse-article-modal-sub{font-size:11px;color:#64748b;margin-top:3px}.caisse-article-modal-close{border:0;background:#eaf6ff;color:#0f2f4f;border-radius:10px;width:36px;height:36px;font-size:21px;cursor:pointer}
    .caisse-article-modal-body{padding:18px}.caisse-article-hero{background:linear-gradient(135deg,#0f2f4f,#16476f);color:#fff;border-radius:16px;padding:18px;display:flex;justify-content:space-between;gap:16px;align-items:flex-start;flex-wrap:wrap}.caisse-article-hero .ref{font-size:25px;font-weight:950;overflow-wrap:anywhere}.caisse-article-hero .desc{font-size:12px;color:#dbeafe;margin-top:5px;max-width:650px}.caisse-article-hero .esi{font-size:11px;font-weight:900;background:rgba(255,255,255,.13);border:1px solid rgba(255,255,255,.25);padding:6px 9px;border-radius:999px}
    .caisse-article-detail-layout{display:grid;grid-template-columns:minmax(0,1fr) 240px;gap:14px;margin-top:14px}.caisse-article-fields{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:9px}.caisse-article-field{border:1px solid #c9e0ed;background:#fff;border-radius:12px;padding:10px}.caisse-article-field.wide{grid-column:1/-1}.caisse-article-field .k{font-size:9px;font-weight:900;text-transform:uppercase;color:#64748b;letter-spacing:.04em}.caisse-article-field .v{font-size:13px;font-weight:750;color:#16324a;margin-top:5px;white-space:pre-wrap;overflow-wrap:anywhere}.caisse-article-photo{border:1px solid #c9e0ed;background:#fff;border-radius:14px;min-height:220px;display:flex;align-items:center;justify-content:center;overflow:hidden}.caisse-article-photo img{width:100%;height:100%;min-height:220px;object-fit:contain}.caisse-article-no-photo{font-size:11px;color:#64748b;text-align:center;padding:18px}
    .caisse-article-section{margin-top:14px;border:1px solid #dbe7f0;background:#fff;border-radius:14px;padding:13px}.caisse-article-section-title{font-size:11px;font-weight:900;text-transform:uppercase;color:#0369a1;margin-bottom:9px}.caisse-article-history{display:grid;gap:8px}.caisse-article-history-item{border-left:4px solid #0ea5e9;background:#f8fafc;border-radius:10px;padding:9px 10px;font-size:11px;color:#334155}.caisse-article-history-item strong{color:#0f2f4f}
    @media(max-width:760px){.caisse-articles-search{grid-template-columns:1fr}.caisse-article-detail-layout{grid-template-columns:1fr}.caisse-article-fields{grid-template-columns:1fr}.caisse-article-field.wide{grid-column:auto}.caisse-articles-result-head,.caisse-articles-result{grid-template-columns:34px minmax(90px,.7fr) minmax(120px,1.3fr)}}
  `;
  document.head.appendChild(style);
}

let linkedTicketId='';
let linkedDraft=new Map();
let linkedSearchResults=[];
let linkedLoadSeq=0;

function currentTicket(){
  try{return state.tickets.find(t=>t.id===state.selectedTicketId)||null}catch(e){return null}
}
function isFicheCaisse(ticket){return !!ticket && ticket.module==='Fiche de caisse'}
function refLabel(article){return String((article&&article.reference)||'').trim()||String((article&&article.esi_id)||'').trim()||'Sans référence'}

function ensurePanel(){
  let panel=document.getElementById('caisseArticlesPanel');
  if(panel)return panel;
  panel=document.createElement('div');
  panel.id='caisseArticlesPanel';
  panel.className='caisse-articles-panel';
  const grid=document.getElementById('detailGrid');
  if(grid)grid.insertAdjacentElement('afterend',panel);
  return panel;
}

function selectedHtml(editing){
  const values=[...linkedDraft.values()];
  if(!values.length)return '<div class="caisse-articles-empty">Aucun article lié à cette caisse.</div>';
  return values.map(a=>`<button type="button" class="caisse-article-chip" data-open-esi="${escapeHtml(a.esi_id)}" title="Ouvrir la fiche détaillée">${escapeHtml(refLabel(a))}${editing?`<span class="remove" data-remove-esi="${escapeHtml(a.esi_id)}" title="Retirer">×</span>`:''}</button>`).join('');
}

function renderPanel(){
  const panel=ensurePanel(),ticket=currentTicket();
  if(!panel)return;
  if(!isFicheCaisse(ticket)){panel.classList.remove('show');panel.innerHTML='';return}
  panel.classList.add('show');
  const editing=!!state.editMode;
  panel.innerHTML=`
    <div class="caisse-articles-title">Articles liés à cette caisse</div>
    <div class="caisse-articles-selected" id="caisseArticlesSelected">${selectedHtml(editing)}</div>
    ${editing?`<div class="caisse-articles-search"><input id="caisseArticlesSearchInput" autocomplete="off" placeholder="Rechercher par N° dossier ou N° inventaire / référence"><button class="btn secondary" id="caisseArticlesSearchBtn" type="button">Rechercher</button></div><div class="caisse-articles-hint">Résultats : N° dossier + N° inventaire / référence uniquement.</div><div class="caisse-articles-results" id="caisseArticlesResults" style="display:none"></div>`:''}
  `;
  bindPanelEvents();
}

async function loadLinks(force=false){
  const ticket=currentTicket();
  if(!isFicheCaisse(ticket)){linkedTicketId='';linkedDraft.clear();renderPanel();return}
  if(!force && linkedTicketId===ticket.id){renderPanel();return}
  const seq=++linkedLoadSeq;
  linkedTicketId=ticket.id;linkedDraft.clear();linkedSearchResults=[];
  try{
    const r=await fetch('/api/tickets/'+encodeURIComponent(ticket.id)+'/articles-lies',{cache:'no-store'});
    const d=await r.json();
    if(seq!==linkedLoadSeq)return;
    if(!r.ok)throw new Error(d.error||'Impossible de charger les articles liés');
    (d.articles||[]).forEach(a=>linkedDraft.set(String(a.esi_id),a));
  }catch(e){if(seq===linkedLoadSeq)showNotice(e.message||'Impossible de charger les articles liés')}
  if(seq===linkedLoadSeq)renderPanel();
}

function bindSelectedEvents(){
  const selected=document.getElementById('caisseArticlesSelected');if(!selected)return;
  selected.querySelectorAll('[data-open-esi]').forEach(btn=>btn.addEventListener('click',e=>{
    if(e.target&&e.target.matches('[data-remove-esi]'))return;
    openLinkedArticleDetail(btn.dataset.openEsi);
  }));
  selected.querySelectorAll('[data-remove-esi]').forEach(x=>x.addEventListener('click',e=>{
    e.preventDefault();e.stopPropagation();linkedDraft.delete(String(x.dataset.removeEsi));renderPanel();
  }));
}

function bindPanelEvents(){
  bindSelectedEvents();
  const input=document.getElementById('caisseArticlesSearchInput');
  const button=document.getElementById('caisseArticlesSearchBtn');
  if(button)button.addEventListener('click',searchArticles);
  if(input)input.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();searchArticles()}});
}

async function searchArticles(){
  const input=document.getElementById('caisseArticlesSearchInput'),box=document.getElementById('caisseArticlesResults');
  const q=String(input&&input.value||'').trim();
  if(!q){showNotice('Saisis un N° dossier ou une référence.');return}
  if(box){box.style.display='block';box.innerHTML='<div class="small" style="padding:12px">Recherche…</div>'}
  try{
    const r=await fetch('/api/articles/link-search?q='+encodeURIComponent(q),{cache:'no-store'}),d=await r.json();
    if(!r.ok)throw new Error(d.error||'Recherche impossible');
    linkedSearchResults=d.articles||[];renderSearchResults();
  }catch(e){if(box)box.innerHTML='<div class="small" style="padding:12px">'+escapeHtml(e.message||'Recherche impossible')+'</div>'}
}

function renderSearchResults(){
  const box=document.getElementById('caisseArticlesResults');if(!box)return;
  box.style.display='block';
  if(!linkedSearchResults.length){box.innerHTML='<div class="small" style="padding:12px">Aucun article trouvé.</div>';return}
  box.innerHTML='<div class="caisse-articles-result-head"><div></div><div>N° dossier</div><div>N° inventaire / référence</div></div>'+linkedSearchResults.map(a=>`<label class="caisse-articles-result"><div><input type="checkbox" data-link-esi="${escapeHtml(a.esi_id)}" ${linkedDraft.has(String(a.esi_id))?'checked':''}></div><div class="dos">${escapeHtml(a.dossier||'-')}</div><div class="ref">${escapeHtml(a.reference||'-')}</div></label>`).join('');
  box.querySelectorAll('[data-link-esi]').forEach(cb=>cb.addEventListener('change',()=>{
    const a=linkedSearchResults.find(x=>String(x.esi_id)===String(cb.dataset.linkEsi));
    if(!a)return;
    if(cb.checked)linkedDraft.set(String(a.esi_id),a);else linkedDraft.delete(String(a.esi_id));
    const selected=document.getElementById('caisseArticlesSelected');if(selected)selected.innerHTML=selectedHtml(true);
    bindSelectedEvents();
  }));
}

async function saveLinks(){
  const ticket=currentTicket();if(!isFicheCaisse(ticket))return true;
  const r=await fetch('/api/tickets/'+encodeURIComponent(ticket.id)+'/articles-lies',{
    method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({esi_ids:[...linkedDraft.keys()]})
  });
  const d=await r.json();if(!r.ok)throw new Error(d.error||'Impossible d’enregistrer les articles liés');
  linkedDraft.clear();(d.articles||[]).forEach(a=>linkedDraft.set(String(a.esi_id),a));
  return true;
}

function ensureDetailModal(){
  let bg=document.getElementById('linkedArticleDetailModal');if(bg)return bg;
  bg=document.createElement('div');bg.id='linkedArticleDetailModal';bg.className='caisse-article-modal-backdrop';
  bg.innerHTML=`<div class="caisse-article-modal"><div class="caisse-article-modal-head"><div><div class="caisse-article-modal-title">CARTE D'IDENTITÉ DE L'ARTICLE</div><div class="caisse-article-modal-sub" id="linkedArticleModalSub"></div></div><button class="caisse-article-modal-close" id="linkedArticleModalClose" type="button">×</button></div><div class="caisse-article-modal-body" id="linkedArticleModalBody"></div></div>`;
  document.body.appendChild(bg);
  document.getElementById('linkedArticleModalClose').onclick=()=>bg.classList.remove('show');
  bg.addEventListener('click',e=>{if(e.target===bg)bg.classList.remove('show')});
  return bg;
}
function detailField(k,v,wide=false){const val=String(v??'').trim()||'-';return `<div class="caisse-article-field ${wide?'wide':''}"><div class="k">${escapeHtml(k)}</div><div class="v">${escapeHtml(val)}</div></div>`}
function dateText(v){if(!v)return '-';const d=new Date(v);return isNaN(d.getTime())?String(v):d.toLocaleString('fr-FR')}
async function openLinkedArticleDetail(esi){
  const bg=ensureDetailModal(),body=document.getElementById('linkedArticleModalBody'),sub=document.getElementById('linkedArticleModalSub');
  bg.classList.add('show');sub.textContent=esi;body.innerHTML='<div class="small">Chargement…</div>';
  try{
    const r=await fetch('/api/articles/'+encodeURIComponent(esi),{cache:'no-store'}),d=await r.json();if(!r.ok)throw new Error(d.error||'Article introuvable');
    const a=d.article||{},dims=[a.longueur_cm,a.largeur_cm,a.hauteur_cm].filter(x=>String(x??'').trim()).join(' × '),photo=String(a.photo_url||'').trim();
    sub.textContent=(a.esi_id||esi)+' • Fiche détaillée et historique';
    let html=`<div class="caisse-article-hero"><div><div class="ref">${escapeHtml(a.reference||a.esi_id||esi)}</div>${a.description?`<div class="desc">${escapeHtml(a.description)}</div>`:''}</div><div class="esi">${escapeHtml(a.esi_id||esi)}</div></div>`;
    html+=`<div class="caisse-article-detail-layout"><div class="caisse-article-fields">${detailField('N° ESI',a.esi_id)}${detailField('N° dossier',a.dossier)}${detailField('Référence / inventaire',a.reference)}${detailField('Client',a.client)}${detailField('Projet / exposition',a.projet,true)}${detailField('Description / désignation',a.description,true)}${detailField('Dimensions',dims?dims+' cm':'-')}${detailField('Poids',a.poids_kg?String(a.poids_kg)+' kg':'-')}${detailField('Stockage actuel',a.lieu_stockage)}${detailField('Statut logistique',a.statut_logistique)}${detailField('N° colis',a.dernier_colis)}${detailField('Dernière réception',a.derniere_reception_ref)}</div><div class="caisse-article-photo">${photo?`<img src="${escapeHtml(photo)}" alt="Photo article">`:'<div class="caisse-article-no-photo">Aucune photo enregistrée</div>'}</div></div>`;
    html+='<div class="caisse-article-section"><div class="caisse-article-section-title">Historique des réceptions</div>';
    if((d.receptions||[]).length){
      html+='<div class="caisse-article-history">'+d.receptions.map(x=>`<div class="caisse-article-history-item"><strong>${escapeHtml(x.reference||'Réception')}</strong> · ${escapeHtml(x.date_affichee||dateText(x.date))}<br>Stockage : ${escapeHtml(x.lieu_stockage||'-')} · Colis : ${escapeHtml((x.colis||[]).join(', ')||'-')}</div>`).join('')+'</div>';
    }else html+='<div class="small">Aucune réception enregistrée.</div>';
    html+='</div>';body.innerHTML=html;
  }catch(e){body.innerHTML='<div class="small">'+escapeHtml(e.message||'Impossible de charger la fiche article')+'</div>'}
}

// Rend le bloc a chaque affichage / passage en modification.
if(typeof renderDetail==='function'){
  const originalRenderDetail=renderDetail;
  renderDetail=async function(){const result=await originalRenderDetail.apply(this,arguments);await loadLinks(false);return result};
}

// Sauvegarde d'abord la selection, puis le ticket avec le bouton Enregistrer deja present.
const saveButton=document.getElementById('saveTicketBtn');
if(saveButton){
  let saving=false;
  saveButton.addEventListener('click',async function(e){
    const ticket=currentTicket();if(!state.editMode||!isFicheCaisse(ticket))return;
    e.preventDefault();e.stopImmediatePropagation();if(saving)return;saving=true;
    const oldText=saveButton.textContent;saveButton.disabled=true;saveButton.textContent='Enregistrement…';
    try{await saveLinks();await saveEditedTicket()}catch(err){showNotice(err.message||'Erreur enregistrement des articles liés')}finally{saving=false;saveButton.disabled=false;saveButton.textContent=oldText}
  },true);
}

// Cas ou la page a fini son premier rendu avant l'injection.
setTimeout(()=>{try{loadLinks(true)}catch(e){}},100);
})();"""


@app.route('/gestionnaire')
def gestionnaire():
    from flask import request, redirect, url_for
    if request.args.get('pwd') != '1234':
        return redirect(url_for('login'))

    # Injection inline volontaire : pas de fichier JS externe, donc pas de probleme de cache/404.
    page = render_template('gestionnaire.html')
    inline = '<script>' + GESTIONNAIRE_ARTICLES_LIES_JS + '</script>'
    if '</body>' in page:
        page = page.replace('</body>', inline + '\n</body>', 1)
    else:
        page += inline

    response = app.make_response(page)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response

@app.route('/login', methods=['GET','POST'])
def login():
    from flask import request, redirect, render_template_string
    error = ''
    if request.method == 'POST':
        if request.form.get('password') == '1234':
            return redirect('/gestionnaire?pwd=1234')
        error = 'Mot de passe incorrect'
    return render_template_string("""<!DOCTYPE html>
<html lang='fr'>
<head>
<meta charset='UTF-8'>
<meta name='viewport' content='width=device-width, initial-scale=1.0'>
<title>Connexion gestionnaire</title>
<style>
body{font-family:Arial,Helvetica,sans-serif;background:linear-gradient(180deg,#eef6fb 0%,#f6f8fb 100%);margin:0;display:flex;align-items:center;justify-content:center;height:100vh;color:#1e293b}
.card{background:#fff;border:1px solid #dbe7f0;border-radius:20px;padding:28px;box-shadow:0 12px 30px rgba(15,23,42,.08);width:360px}
h1{margin:0 0 10px;font-size:28px} p{margin:0 0 18px;color:#64748b}
input{width:100%;padding:12px 14px;border:1px solid #dbe7f0;border-radius:14px;font-size:15px;box-sizing:border-box}
button{margin-top:14px;width:100%;padding:12px 14px;border:none;border-radius:14px;background:linear-gradient(135deg,#0ea5e9 0%, #0284c7 100%);color:#fff;font-weight:700;cursor:pointer}
.err{margin-top:12px;color:#b91c1c;font-size:13px}
</style>
</head>
<body>
  <form class='card' method='post'>
    <h1>Gestion Tickets</h1>
    <p>Accès protégé par mot de passe</p>
    <input type='password' name='password' placeholder='Mot de passe' autofocus />
    <button type='submit'>Entrer</button>
    {% if error %}<div class='err'>{{ error }}</div>{% endif %}
  </form>
</body>
</html>""", error=error)


@app.route('/reception')
def reception():
    return render_template('reception.html')


@app.route('/gestion-reception')
def gestion_reception():
    return render_template('gestion_reception.html')

@app.route('/api/status')
def api_status():
    root = ensure_shared_root()
    return jsonify({'shared_path': str(root), 'mode': 'automatic_app_folder'})

@app.route('/api/tickets')
def api_tickets():
    status = request.args.get('status')
    limit = request.args.get('limit')

    # Migration historique : toutes les caisses antérieures au 18/08/2026
    # sont considérées comme réceptionnées.
    migrate_caisses_avant_18_aout_2026()

    tickets = list_tickets(status=status, limit=limit)

    # Expose un indicateur logistique normalisé pour le planning.
    # Cela ne modifie jamais le statut métier du ticket.
    for ticket in tickets:
        if ticket.get('module') == 'Fiche de caisse':
            reception = dict(ticket.get('reception') or {})
            reception['receptionnee'] = _is_caisse_receptionnee(ticket)
            ticket['reception'] = reception

    return jsonify(tickets)


@app.route('/api/caisse-status/<path:caisse_ref>')
def api_caisse_status(caisse_ref):
    """Retourne l'état fournisseur d'une caisse depuis le Google Sheet public."""
    result = get_caisse_fournisseur_status(caisse_ref)
    if result.get("success"):
        return jsonify(result)

    # On garde un HTTP 200 pour que la fenêtre puisse afficher proprement
    # "Caisse introuvable" sans traiter cela comme une panne réseau.
    return jsonify(result)


@app.route('/api/tickets/<ticket_id>/localisation', methods=['PATCH'])
def api_update_localisation(ticket_id):
    """Met à jour uniquement la localisation d'une fiche de caisse."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404

    fiche = ticket.get('fiche') or {}
    if not fiche:
        return jsonify({'ok': False, 'error': 'Fiche de caisse introuvable'}), 404

    data = request.get_json(silent=True) or {}
    localisation = _as_text(data.get('localisation')).strip()

    fiche['localisation'] = localisation
    ticket['fiche'] = fiche
    ticket['updatedAt'] = datetime.now().isoformat()

    # Une localisation renseignée acte une réception manuelle.
    # IMPORTANT : la réception est indépendante du statut métier du ticket.
    reception = dict(ticket.get('reception') or {})
    if localisation:
        reception['receptionnee'] = True
        reception.setdefault('receptionnee_le', datetime.now().isoformat())
        reception.setdefault('mode', 'manuel')
        reception['localisation'] = localisation
        ticket['reception'] = reception
    else:
        # Effacer la localisation annule uniquement l'information de réception.
        # Le statut du ticket (Demande créée / En cours / Terminé) n'est jamais modifié ici.
        ticket.pop('reception', None)
        reception = {}

    try:
        supabase_rest_request(
            "PATCH",
            "fiches",
            "ticket_id=eq." + urllib.parse.quote(ticket_id, safe=''),
            {"localisation": localisation},
            prefer="return=minimal"
        )
        # raw_json conserve les métadonnées de réception sans nouvelle colonne Supabase.
        supabase_rest_request(
            "PATCH",
            "tickets",
            "id=eq." + urllib.parse.quote(ticket_id, safe=''),
            {"updated_at": ticket['updatedAt'], "raw_json": ticket},
            prefer="return=minimal"
        )
        return jsonify({
            'ok': True,
            'localisation': localisation,
            'reception': reception,
            'receptionnee': _is_caisse_receptionnee(ticket),
            'status': ticket.get('status')
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/reception/analyse-bl', methods=['POST'])
def api_reception_analyse_bl():
    """
    Analyse un bordereau PDF sans modifier la base.
    Retourne les caisses détectées et leur correspondance avec ESI TICKETS.
    """
    fs = request.files.get('file')
    if not fs or not fs.filename:
        return jsonify({'ok': False, 'error': 'Fichier PDF manquant'}), 400

    if not fs.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'error': 'Le fichier doit être un PDF'}), 400

    content = fs.read()
    if not content:
        return jsonify({'ok': False, 'error': 'Le fichier PDF est vide'}), 400

    try:
        cache_key, parsed = _reception_cache_get(content)
        if parsed is not None:
            print("[RECEPTION OCR] Resultat reutilise depuis le cache PDF")
        else:
            parsed = _extract_reception_pdf(content)
            _reception_cache_set(cache_key, parsed)

        print(
            f"[RECEPTION PDF] Analyse terminee: BL={parsed.get('bl_numero', '') or '-'} "
            f"refs={len(parsed.get('references') or [])} ocr={parsed.get('ocr_used', False)}"
        )
        matches = _match_reception_refs_to_tickets(parsed["references"])
        print(f"[RECEPTION PDF] Rapprochement termine: {len(matches)} ligne(s)")

        # Le PDF analysé est conservé dans Supabase afin de pouvoir le rouvrir
        # depuis la fiche de la caisse après validation de la réception.
        pdf_hash = hashlib.sha256(content).hexdigest()[:16]
        bl_numero = parsed.get('bl_numero', '') or 'sans_numero'
        bl_filename = safe_filename(fs.filename or f"BL_{bl_numero}.pdf")
        bl_storage_path = f"reception_bls/{safe_filename(bl_numero)}_{pdf_hash}_{bl_filename}"
        supabase_upload_bytes(bl_storage_path, content, "application/pdf")

        return jsonify({
            'ok': True,
            'bl_numero': parsed.get('bl_numero', ''),
            'bl_date': parsed.get('bl_date', ''),
            'bl_storage_path': bl_storage_path,
            'bl_filename': fs.filename or bl_filename,
            'page_count': parsed.get('page_count', 0),
            'ocr_used': parsed.get('ocr_used', False),
            'items': matches,
            'found_count': sum(1 for x in matches if x.get('found')),
            'missing_count': sum(1 for x in matches if not x.get('found')),
        })
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    except Exception as e:
        print(f"[RECEPTION PDF] Erreur analyse : {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/reception/valider-bl', methods=['POST'])
def api_reception_valider_bl():
    """Valide en lot la réception et rattache le BL fournisseur à chaque caisse."""
    data = request.get_json(silent=True) or {}
    ticket_ids = data.get('ticket_ids') or []
    localisation = _as_text(data.get('localisation')).strip()
    bl_numero = _as_text(data.get('bl_numero')).strip()
    bl_date = _as_text(data.get('bl_date')).strip()
    bl_storage_path = _as_text(data.get('bl_storage_path')).strip()
    bl_filename = _as_text(data.get('bl_filename')).strip()

    if not localisation:
        return jsonify({
            'ok': False,
            'error': 'La localisation est obligatoire pour valider la réception.'
        }), 400

    if not isinstance(ticket_ids, list) or not ticket_ids:
        return jsonify({'ok': False, 'error': 'Aucune caisse sélectionnée'}), 400

    if not bl_storage_path:
        return jsonify({'ok': False, 'error': 'Le PDF du BL analysé est manquant.'}), 400

    updated = []
    errors = []
    receptionnee_le = datetime.now().isoformat()

    for ticket_id in dict.fromkeys(_as_text(x).strip() for x in ticket_ids if _as_text(x).strip()):
        try:
            ticket = load_ticket(ticket_id)
            if not ticket:
                errors.append({'ticket_id': ticket_id, 'error': 'Ticket introuvable'})
                continue

            if ticket.get('module') != 'Fiche de caisse':
                errors.append({'ticket_id': ticket_id, 'error': 'Ce ticket n’est pas une fiche de caisse'})
                continue

            fiche = ticket.get('fiche') or {}
            if not fiche:
                errors.append({'ticket_id': ticket_id, 'error': 'Fiche de caisse introuvable'})
                continue

            fiche['localisation'] = localisation
            ticket['fiche'] = fiche
            ticket['reception'] = {
                'receptionnee': True,
                'receptionnee_le': receptionnee_le,
                'localisation': localisation,
                'mode': 'bl_fournisseur',
                'bl_numero': bl_numero,
                'bl_date': bl_date,
                'bl_storage_path': bl_storage_path,
                'bl_filename': bl_filename,
            }
            # La réception logistique ne modifie pas le statut métier du ticket.
            ticket['updatedAt'] = datetime.now().isoformat()

            supabase_rest_request(
                "PATCH",
                "fiches",
                "ticket_id=eq." + urllib.parse.quote(ticket_id, safe=''),
                {"localisation": localisation},
                prefer="return=minimal"
            )

            supabase_rest_request(
                "PATCH",
                "tickets",
                "id=eq." + urllib.parse.quote(ticket_id, safe=''),
                {"updated_at": ticket['updatedAt'], "raw_json": ticket},
                prefer="return=minimal"
            )

            updated.append(ticket_id)

        except Exception as e:
            print(f"[RECEPTION PDF] Erreur validation {ticket_id}: {e}")
            errors.append({'ticket_id': ticket_id, 'error': str(e)})

    return jsonify({
        'ok': len(errors) == 0,
        'updated_count': len(updated),
        'updated': updated,
        'errors': errors,
        'localisation': localisation,
        'bl_numero': bl_numero,
        'receptionnee_le': receptionnee_le,
    }), (200 if not errors else 207)


@app.route('/api/tickets/<ticket_id>/reception-bl')
def api_ticket_reception_bl(ticket_id):
    """Ouvre le BL fournisseur associé à la réception d'une caisse."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        abort(404)

    reception = ticket.get('reception') or {}
    storage_path = _as_text(reception.get('bl_storage_path')).strip()
    filename = _as_text(reception.get('bl_filename')).strip() or (
        f"BL_{_as_text(reception.get('bl_numero')).strip() or ticket_id}.pdf"
    )
    if not storage_path:
        abort(404)

    try:
        signed_url = supabase_signed_download_url(storage_path, expires_in=300)
        return redirect(signed_url)
    except Exception as e:
        print(f"[RECEPTION BL] URL signée impossible, fallback Render : {e}")
        try:
            data = supabase_download_bytes(storage_path)
        except Exception:
            abort(404)
        return send_file(BytesIO(data), mimetype='application/pdf', download_name=filename)


@app.route('/api/tickets', methods=['POST'])
def api_create_ticket():
    form = request.form
    module = form.get('module','')
    # Normalise l'apostrophe typographique envoyee par certains navigateurs/pages HTML.
    # Ainsi "Demande d’enlèvement" et "Demande d'enlèvement" sont traites de la meme facon.
    module = module.replace("’", "'").strip()

    prefixes = {
        'Fiche de caisse': 'DEM',
        'Demande de devis': 'DEV',
        'Demande Aller voir': 'AV',
        'Demande d\'enlèvement': 'ENL',
        'Demande d\'enlevement': 'ENL',
        "Avis d'arrivée": 'ARR',
    }
    prefix = prefixes.get(module, 'AV')

    incoming_files = [fs for fs in request.files.getlist('files') if fs and fs.filename]
    is_enlevement = module in ("Demande d'enlèvement", "Demande d'enlevement")
    is_avis_arrivee = module == "Avis d'arrivée"
    avis_arrivee = None
    enlevement_analyse = None
    article_selections = []
    if is_enlevement:
        raw_analysis = form.get('enlevementAnalyse', '')
        raw_selections = form.get('articleSelections', '')
        if raw_analysis:
            try:
                enlevement_analyse = json.loads(raw_analysis)
            except (json.JSONDecodeError, TypeError, ValueError):
                return jsonify({'ok': False, 'error': "Analyse du bon d'enlèvement invalide. Relance l'analyse du PDF."}), 400
            if not isinstance(enlevement_analyse, dict):
                return jsonify({'ok': False, 'error': "Analyse du bon d'enlèvement invalide."}), 400
        if raw_selections:
            try:
                article_selections = json.loads(raw_selections)
            except (json.JSONDecodeError, TypeError, ValueError):
                return jsonify({'ok': False, 'error': 'Sélection des articles invalide.'}), 400
            if not isinstance(article_selections, list):
                return jsonify({'ok': False, 'error': 'Sélection des articles invalide.'}), 400
    if is_avis_arrivee:
        raw_avis = form.get('avisArrivee', '')
        try:
            avis_arrivee = json.loads(raw_avis) if raw_avis else {}
        except (json.JSONDecodeError, TypeError, ValueError):
            return jsonify({'ok': False, 'error': "Données de l'avis d'arrivée invalides."}), 400
        if not isinstance(avis_arrivee, dict):
            return jsonify({'ok': False, 'error': "Données de l'avis d'arrivée invalides."}), 400

        required = {
            'dossier_ref': 'Réf. dossier',
            'client': 'Nom du client',
            'projet': 'Projet ou expo',
            'date_reception_prevue': 'Date de réception prévue',
            'coordinateur': 'Nom du coordinateur',
        }
        missing = [label for key, label in required.items() if not _as_text(avis_arrivee.get(key)).strip()]
        if missing:
            return jsonify({'ok': False, 'error': 'Champ(s) obligatoire(s) manquant(s) : ' + ', '.join(missing)}), 400
        if not isinstance(avis_arrivee.get('items'), list) or not avis_arrivee.get('items'):
            return jsonify({'ok': False, 'error': "Ajoute au moins une ligne de marchandise."}), 400

    # Pour une demande d'enlevement, le demandeur ne fait qu'une chose : deposer un PDF.
    if is_enlevement:
        if len(incoming_files) != 1:
            return jsonify({
                'ok': False,
                'error': "La demande d'enlèvement doit contenir un seul bon PDF."
            }), 400
        if not incoming_files[0].filename.lower().endswith('.pdf'):
            return jsonify({
                'ok': False,
                'error': "Le bon d'enlèvement doit être un fichier PDF."
            }), 400

    ticket_id = next_id(prefix)
    ticket = {
        'id': ticket_id,
        'module': module,
        'status': 'Demande créée',
        'createdAt': datetime.now().isoformat(),
        'dossier': form.get('dossier',''),
        'ref': form.get('ref',''),
        'preteur': form.get('preteur','-') or '-',
        'expo': form.get('expo','-') or '-',
        'objet': form.get('objet','-') or '-',
        'chargeProjet': form.get('chargeProjet','-') or '-',
        'typeCaisse': form.get('typeCaisse','-') or '-',
        'dimensions': form.get('dimensions','-') or '-',
        'dateEmballage': form.get('dateEmballage','-') or '-',
        'prixDevis': form.get('prixDevis','-') or '-',
        'dateRdv': form.get('dateRdv','-') or '-',
        'heureRdv': form.get('heureRdv','-') or '-',
        'lieuRdv': form.get('lieuRdv','-') or '-',
        'contactRdv': form.get('contactRdv','-') or '-',
        'commentaire': form.get('commentaire',''),
        'files': [],
        'managerSheets': []
    }

    if is_enlevement:
        if enlevement_analyse:
            clean_analysis = dict(enlevement_analyse)
            for item in clean_analysis.get('items') or []:
                if isinstance(item, dict):
                    item.pop('article_candidates', None)
                    item.pop('source_index', None)
            clean_analysis['analysis_status'] = 'ready'
            clean_analysis['analysis_error'] = ''
            clean_analysis['analysed_at'] = datetime.now().isoformat()
            ticket['enlevement'] = clean_analysis

            if clean_analysis.get('client'):
                ticket['dossier'] = clean_analysis['client']
            if clean_analysis.get('numero_bon'):
                ticket['ref'] = clean_analysis['numero_bon']
            if clean_analysis.get('coordinateur'):
                ticket['chargeProjet'] = clean_analysis['coordinateur']
            if clean_analysis.get('exhibition'):
                ticket['expo'] = clean_analysis['exhibition']
                ticket['objet'] = clean_analysis['exhibition']
            if clean_analysis.get('date_enlevement'):
                try:
                    dt = datetime.strptime(clean_analysis['date_enlevement'], '%d/%m/%Y')
                    ticket['dateRdv'] = dt.strftime('%Y-%m-%d')
                except ValueError:
                    pass

            _apply_enlevement_article_selections(ticket, article_selections)
        else:
            ticket['enlevement'] = {
                'analysis_status': 'pending',
                'analysis_error': '',
                'items': [],
                'references': [],
            }

    if is_avis_arrivee:
        # Les données spécifiques restent dans raw_json : aucune nouvelle colonne Supabase n'est nécessaire.
        # Les champs historiques ci-dessus restent remplis pour que l'avis apparaisse dans les listes existantes.
        ticket['avisArrivee'] = avis_arrivee

    ticket_folder(ticket_id)  # conserve la création du dossier local historique

    enlevement_pdf_bytes = None
    for fs in incoming_files:
        content = fs.read()
        if is_enlevement:
            enlevement_pdf_bytes = content

        clean_name = safe_filename(fs.filename)
        storage_path = f"{ticket_id}/{datetime.now().strftime('%Y%m%d%H%M%S')}_{clean_name}"

        try:
            supabase_upload_bytes(
                storage_path,
                content,
                fs.content_type
            )
        except Exception as e:
            print(f"[SUPABASE UPLOAD] Erreur : {e}")
            return jsonify({'ok': False, 'error': f'Erreur upload Supabase : {e}'}), 500

        ticket['files'].append({
            'name': fs.filename,
            'size': len(content),
            'path': storage_path
        })

    # Attribue immédiatement un identifiant ESI à chaque unité d'un avis d'arrivée.
    # Le ticket est d'abord sauvegardé afin que les références Supabase soient cohérentes.
    save_ticket(ticket)
    if is_avis_arrivee:
        try:
            _ensure_articles_for_ticket(ticket, save=True)
        except Exception as e:
            print(f"[ARTICLES] Attribution ESI impossible pour {ticket_id}: {e}")
            return jsonify({
                'ok': False,
                'error': "Le ticket a été créé, mais l'attribution des numéros ESI a échoué. "
                         "Vérifie que la table Supabase 'articles' a bien été créée."
            }), 500

    if is_enlevement and enlevement_analyse:
        try:
            _ensure_articles_for_ticket(ticket, save=True)
        except Exception as e:
            print(f"[ARTICLES] Rattachement/création ESI impossible pour {ticket_id}: {e}")
            return jsonify({
                'ok': False,
                'error': "Le ticket a été créé, mais le rattachement des articles a échoué. " + str(e)
            }), 500
    elif is_enlevement and enlevement_pdf_bytes:
        # Compatibilité avec les anciennes pages : sans pré-analyse, on conserve l'ancien fonctionnement.
        worker = threading.Thread(
            target=_analyse_enlevement_ticket_background,
            args=(ticket_id, enlevement_pdf_bytes),
            daemon=True,
            name=f"enlevement-{ticket_id}"
        )
        worker.start()

    return jsonify({
        'ok': True,
        'id': ticket_id,
        'analysis_status': ('ready' if is_enlevement and enlevement_analyse else ('pending' if is_enlevement else None))
    })


def _reception_qty_int(value, default=0):
    try:
        return max(0, int(float(str(value).replace(',', '.'))))
    except Exception:
        return default


def _rebuild_ticket_after_reception_cancel(ticket, is_avis):
    """Reconstruit les quantités et états articles uniquement depuis les réceptions encore actives."""
    if is_avis:
        container = dict(ticket.get('avisArrivee') or ticket.get('avis_arrivee') or {})
        active = list(ticket.get('receptionsAvisArrivee') or [])
        container_key = 'avisArrivee'
    else:
        container = dict(ticket.get('enlevement') or {})
        active = list(container.get('bons_livraison') or [])
        container_key = 'enlevement'

    items = [dict(x or {}) for x in (container.get('items') or [])]
    for idx, item in enumerate(items):
        total = 0
        item_history = []
        latest = None
        for rec in active:
            for rec_item in rec.get('items') or []:
                try:
                    rec_idx = int(rec_item.get('index'))
                except Exception:
                    continue
                if rec_idx != idx:
                    continue
                qty = _reception_qty_int(rec_item.get('quantite'), 0)
                total += qty
                latest = rec
                item_history.append({
                    'reference': _as_text(rec.get('reference')).strip(),
                    'date': rec.get('receptionnee_le') or rec.get('created_at') or rec.get('date_reception') or '',
                    'quantite': qty,
                    'receptionne_par': rec.get('receptionne_par') or '',
                    'lieu_stockage': rec.get('lieu_stockage') or '',
                })

        planned = max(1, _reception_qty_int(item.get('quantite') or 1, 1))
        item['quantite_recue_totale'] = total
        item['receptionne'] = total >= planned
        item['receptions'] = item_history
        if latest is not None and total > 0:
            item['receptionne_le'] = latest.get('receptionnee_le') or latest.get('created_at') or ''
            item['receptionne_par'] = latest.get('receptionne_par') or ''
            item['lieu_stockage'] = latest.get('lieu_stockage') or ''
        else:
            item['receptionne_le'] = ''
            item['receptionne_par'] = ''
            item['lieu_stockage'] = ''
        items[idx] = item

    container['items'] = items
    ticket[container_key] = container
    if is_avis:
        ticket['receptionAvisArrivee'] = active[-1] if active else None
    return ticket


def _active_reception_for_esi(ticket, esi_id, is_avis):
    if is_avis:
        active = list(ticket.get('receptionsAvisArrivee') or [])
    else:
        active = list((ticket.get('enlevement') or {}).get('bons_livraison') or [])
    found = None
    for rec in active:
        for rec_item in rec.get('items') or []:
            ids = [_as_text(x).strip() for x in (rec_item.get('esi_ids') or [])]
            if esi_id in ids:
                colis_map = rec_item.get('colis_par_esi') if isinstance(rec_item.get('colis_par_esi'), dict) else {}
                found = (rec, _as_text(colis_map.get(esi_id)).strip())
    return found


def _sync_articles_after_reception_cancel(ticket, cancelled, is_avis):
    """Retire l'historique logistique de la réception annulée sans bloquer l'annulation du ticket."""
    affected = []
    for esi_id in cancelled.get('article_esi_ids') or []:
        esi_id = _as_text(esi_id).strip()
        if esi_id and esi_id not in affected:
            affected.append(esi_id)
    for rec_item in cancelled.get('items') or []:
        for esi_id in rec_item.get('esi_ids') or []:
            esi_id = _as_text(esi_id).strip()
            if esi_id and esi_id not in affected:
                affected.append(esi_id)

    cancelled_ref = _as_text(cancelled.get('reference')).strip()
    for esi_id in affected:
        try:
            safe_esi = urllib.parse.quote(esi_id, safe='-')
            rows = supabase_rest_request('GET', 'articles', f'select=*&esi_id=eq.{safe_esi}&limit=1') or []
            if not rows:
                continue
            article = dict(rows[0])
            raw = article.get('raw_json') if isinstance(article.get('raw_json'), dict) else {}
            raw = dict(raw or {})

            old_history = list(raw.get('receptions') or [])
            removed = [r for r in old_history if _as_text(r.get('reception_ref')).strip() == cancelled_ref]
            raw['receptions'] = [r for r in old_history if _as_text(r.get('reception_ref')).strip() != cancelled_ref]
            cancelled_history = list(raw.get('receptions_annulees') or [])
            for entry in removed:
                copy = dict(entry)
                copy['annulee_le'] = cancelled.get('annulee_le') or datetime.now().isoformat()
                cancelled_history.append(copy)
            raw['receptions_annulees'] = cancelled_history

            latest = _active_reception_for_esi(ticket, esi_id, is_avis)
            if latest:
                rec, colis = latest
                raw['colis_actuel'] = colis
                patch = {
                    'lieu_stockage': _as_text(rec.get('lieu_stockage')).strip(),
                    'statut_logistique': 'Réceptionné',
                    'dernier_colis': colis,
                    'derniere_reception_ref': _as_text(rec.get('reference')).strip(),
                    'updated_at': datetime.now().isoformat(),
                    'raw_json': raw,
                }
            else:
                raw['colis_actuel'] = ''
                patch = {
                    'lieu_stockage': '',
                    'statut_logistique': 'Créé',
                    'dernier_colis': '',
                    'derniere_reception_ref': '',
                    'updated_at': datetime.now().isoformat(),
                    'raw_json': raw,
                }
            merged = dict(article)
            merged.update(patch)
            patch['search_text'] = _article_search_text(merged)
            supabase_rest_request('PATCH', 'articles', f'esi_id=eq.{safe_esi}', patch, prefer='return=minimal')
        except Exception as e:
            print(f'[ANNULATION RECEPTION] Article {esi_id}: {e}')


def _cancel_specific_reception(ticket, reference):
    reference = _as_text(reference).strip()
    if not reference:
        raise ValueError('Référence de réception manquante')

    module = _as_text(ticket.get('module')).replace('’', "'").strip()
    is_avis = module == "Avis d'arrivée" or _as_text(ticket.get('id')).startswith('ARR-')
    cancelled = None

    if is_avis:
        active = list(ticket.get('receptionsAvisArrivee') or [])
        kept = []
        for rec in active:
            if cancelled is None and _as_text(rec.get('reference')).strip() == reference:
                cancelled = dict(rec)
            else:
                kept.append(rec)
        if cancelled is None:
            raise ValueError(f'Réception {reference} introuvable parmi les réceptions actives')
        cancelled['annulee'] = True
        cancelled['annulee_le'] = datetime.now().isoformat()
        ticket['receptionsAvisArrivee'] = kept
        history = list(ticket.get('receptionsAvisArriveeAnnulees') or [])
        history.append(cancelled)
        ticket['receptionsAvisArriveeAnnulees'] = history
    else:
        enl = dict(ticket.get('enlevement') or {})
        active = list(enl.get('bons_livraison') or [])
        kept = []
        for rec in active:
            if cancelled is None and _as_text(rec.get('reference')).strip() == reference:
                cancelled = dict(rec)
            else:
                kept.append(rec)
        if cancelled is None:
            raise ValueError(f'Réception {reference} introuvable parmi les réceptions actives')
        cancelled['annulee'] = True
        cancelled['annulee_le'] = datetime.now().isoformat()
        enl['bons_livraison'] = kept
        history = list(enl.get('bons_livraison_annules') or [])
        history.append(cancelled)
        enl['bons_livraison_annules'] = history
        ticket['enlevement'] = enl

    _rebuild_ticket_after_reception_cancel(ticket, is_avis)
    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)

    # Vérification côté base AVANT de dire OK au navigateur.
    checked = load_ticket(ticket.get('id'))
    if not checked:
        raise RuntimeError('Impossible de relire le ticket après annulation')
    if is_avis:
        still_active = any(
            _as_text(r.get('reference')).strip() == reference
            for r in (checked.get('receptionsAvisArrivee') or [])
        )
    else:
        still_active = any(
            _as_text(r.get('reference')).strip() == reference
            for r in ((checked.get('enlevement') or {}).get('bons_livraison') or [])
        )
    if still_active:
        raise RuntimeError(f'La réception {reference} est encore active après sauvegarde Supabase')

    # La base Articles est synchronisée seulement après validation de la sauvegarde du ticket.
    _sync_articles_after_reception_cancel(checked, cancelled, is_avis)
    return checked


@app.route('/api/tickets/<ticket_id>', methods=['PUT'])
def api_update_ticket(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    data = request.get_json(silent=True) or {}

    if _as_text(data.get('action')).strip() == 'annuler_reception':
        reference = _as_text(data.get('reference')).strip()
        try:
            refreshed = _cancel_specific_reception(ticket, reference)
            module = _as_text(refreshed.get('module')).replace('’', "'").strip()
            is_avis = module == "Avis d'arrivée" or _as_text(refreshed.get('id')).startswith('ARR-')
            active_count = (
                len(refreshed.get('receptionsAvisArrivee') or [])
                if is_avis else
                len((refreshed.get('enlevement') or {}).get('bons_livraison') or [])
            )
            return jsonify({
                'ok': True,
                'action': 'annuler_reception',
                'reference': reference,
                'active_receptions': active_count,
            })
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 404
        except Exception as e:
            print(f'[ANNULATION RECEPTION] {ticket_id} {reference}: {e}')
            return jsonify({'ok': False, 'error': str(e)}), 500

    editable_fields = [
        'dossier',
        'ref',
        'preteur',
        'expo',
        'objet',
        'chargeProjet',
        'typeCaisse',
        'dimensions',
        'dateEmballage',
        'prixDevis',
        'dateRdv',
        'heureRdv',
        'lieuRdv',
        'contactRdv',
        'commentaire'
    ]

    for field in editable_fields:
        if field in data:
            ticket[field] = data.get(field, '')

    if 'expo' in data and 'objet' not in data:
        ticket['objet'] = data.get('expo', '')

    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})


@app.route('/api/tickets/<ticket_id>/reception-avis-arrivee', methods=['POST'])
def api_reception_avis_arrivee(ticket_id):
    """Valide une réception partielle/totale d'un avis d'arrivée et génère les étiquettes."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404

    module_normalise = _as_text(ticket.get('module')).replace("’", "'").strip()
    if module_normalise != "Avis d'arrivée" and not _as_text(ticket_id).startswith('ARR-'):
        return jsonify({'ok': False, 'error': "Ce ticket n'est pas un avis d'arrivée"}), 400

    data = request.get_json(silent=True) or {}
    receptionne_par = _as_text(data.get('receptionne_par')).strip()
    lieu_stockage = _as_text(data.get('lieu_stockage')).strip()
    numero_dossier = _as_text(data.get('numero_dossier')).strip()
    commentaire = _as_text(data.get('commentaire')).strip()

    try:
        nombre_colis = int(data.get('nombre_colis') or 0)
    except (TypeError, ValueError):
        nombre_colis = 0

    items_reception = data.get('items_reception')
    selected_indexes = data.get('selected_indexes') or []
    colis_repartition = data.get('colis_repartition') or []

    if not receptionne_par:
        return jsonify({'ok': False, 'error': 'Nom et prénom du réceptionnaire manquants'}), 400
    if not lieu_stockage:
        return jsonify({'ok': False, 'error': 'Lieu de stockage manquant'}), 400
    if not numero_dossier:
        return jsonify({'ok': False, 'error': 'N° dossier obligatoire pour numéroter les colis'}), 400
    if nombre_colis < 1:
        return jsonify({'ok': False, 'error': 'Le nombre total de colis doit être supérieur ou égal à 1'}), 400

    avis = dict(ticket.get('avisArrivee') or ticket.get('avis_arrivee') or {})
    items = list(avis.get('items') or [])
    if not items:
        return jsonify({'ok': False, 'error': "Aucune marchandise dans cet avis d'arrivée"}), 400

    # Compatibilité ancienne interface
    if not isinstance(items_reception, list):
        items_reception = []
        for raw_idx in selected_indexes if isinstance(selected_indexes, list) else []:
            try:
                items_reception.append({'index': int(raw_idx), 'quantite_recue': None})
            except Exception:
                pass

    if not items_reception:
        return jsonify({'ok': False, 'error': 'Aucune marchandise sélectionnée'}), 400

    def _qty_int(value, default=0):
        try:
            return max(0, int(float(str(value).replace(',', '.'))))
        except Exception:
            return default

    def _item_fully_received(item):
        planned = max(1, _qty_int(item.get('quantite') or 1, 1))
        received = _qty_int(item.get('quantite_recue_totale'), -1)
        if received < 0:
            received = planned if item.get('receptionne') else 0
        return received >= planned

    if items and all(_item_fully_received(dict(x or {})) for x in items):
        return jsonify({
            'ok': False,
            'error': 'Réception clôturée : tous les articles ont déjà été réceptionnés.'
        }), 409

    # Sécurise les anciens avis créés avant l'ajout du référentiel articles.
    try:
        _ensure_articles_for_ticket(ticket, save=True)
        avis = dict(ticket.get('avisArrivee') or {})
        items = list(avis.get('items') or [])
    except Exception as e:
        return jsonify({'ok': False, 'error': f"Impossible d'attribuer les numéros ESI : {e}"}), 500

    now = datetime.now()
    selected = []
    article_labels = []
    reception_esi_ids = []

    for entry in items_reception:
        if not isinstance(entry, dict):
            continue

        try:
            idx = int(entry.get('index'))
        except Exception:
            continue

        if idx < 0 or idx >= len(items):
            continue

        item = dict(items[idx] or {})
        planned = max(1, _qty_int(item.get('quantite') or 1, 1))

        previous = _qty_int(item.get('quantite_recue_totale'), -1)
        if previous < 0:
            previous = planned if item.get('receptionne') else 0

        remaining = max(planned - previous, 0)
        if remaining <= 0:
            continue

        requested = entry.get('quantite_recue')
        qty_received = remaining if requested in (None, '', 0, '0') else _qty_int(requested, 0)

        if qty_received < 1:
            continue

        if qty_received > remaining:
            return jsonify({
                'ok': False,
                'error': f"Quantité reçue trop élevée pour {item.get('reference') or idx} : reste {remaining}"
            }), 400

        new_total = previous + qty_received

        history = list(item.get('receptions') or [])
        history.append({
            'date': now.isoformat(),
            'quantite': qty_received,
            'receptionne_par': receptionne_par,
            'lieu_stockage': lieu_stockage,
        })

        item['quantite_recue_totale'] = new_total
        item['receptionne'] = new_total >= planned
        item['receptionne_le'] = now.isoformat()
        item['receptionne_par'] = receptionne_par
        item['lieu_stockage'] = lieu_stockage
        item['receptions'] = history
        items[idx] = item

        received_esi_ids = _article_ids_for_received_units(item, previous, qty_received)
        reception_esi_ids.extend(received_esi_ids)

        selected_item = {
            'index': idx,
            'quantite': str(qty_received),
            'quantite_prevue': str(planned),
            'quantite_deja_recue': str(previous),
            'quantite_recue_totale': str(new_total),
            'reference': _as_text(item.get('reference')).strip(),
            'designation': _as_text(item.get('description')).strip(),
            'esi_ids': received_esi_ids,
        }
        selected.append(selected_item)

        for unit_no, esi_id in enumerate(received_esi_ids, start=1):
            article_labels.append({
                'titre': 'ARTICLE',
                'principal': esi_id,
                'esi_id': esi_id,
                'dossier': numero_dossier,
                'client': avis.get('client') or ticket.get('dossier') or '',
                'reference': _as_text(item.get('reference')).strip(),
                'designation': _as_text(item.get('description')).strip(),
                'quantite': f"{unit_no}/{qty_received}",
                'lieu': lieu_stockage,
            })

    if not selected:
        return jsonify({'ok': False, 'error': 'Aucune marchandise sélectionnée valide'}), 400

    with _BLR_LOCK:
        # Référence propre aux réceptions d'avis
        existing_refs = []
        try:
            for t in list_tickets():
                for r in list(t.get('receptionsAvisArrivee') or []) + list(t.get('receptionsAvisArriveeAnnulees') or []):
                    ref = _as_text(r.get('reference')).strip()
                    m = re.fullmatch(r"RAR-(\d+)", ref, re.I)
                    if m:
                        existing_refs.append(int(m.group(1)))
        except Exception as e:
            print(f"[RAR] Lecture historique impossible: {e}")

        reception_ref = f"RAR-{(max(existing_refs) if existing_refs else 0) + 1:04d}"
        colis_refs = _allocate_colis_numbers(numero_dossier, nombre_colis)
        try:
            colis_by_esi = _resolve_colis_repartition(selected, colis_repartition, colis_refs)
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        _apply_colis_to_selected_items(selected, colis_by_esi, reception_ref, lieu_stockage)
        for label in article_labels:
            label['colis'] = colis_by_esi.get(label.get('esi_id'), '')

        article_labels_bytes = _build_labels_pdf_bytes(article_labels, kind="article")
        article_labels_filename = f"{reception_ref}_etiquettes_articles.pdf"
        article_labels_path = f"{ticket_id}/receptions_avis/{now.strftime('%Y%m%d%H%M%S')}_{article_labels_filename}"

        colis_labels = [{
            'titre': 'COLIS',
            'principal': colis_ref,
            'dossier': numero_dossier,
            'client': avis.get('client') or ticket.get('dossier') or '',
            'colis': colis_ref,
            'lieu': lieu_stockage,
            'bon': reception_ref,
        } for colis_ref in colis_refs]

        colis_labels_bytes = _build_labels_pdf_bytes(colis_labels, kind="colis")
        colis_labels_filename = f"{reception_ref}_etiquettes_colis.pdf"
        colis_labels_path = f"{ticket_id}/receptions_avis/{now.strftime('%Y%m%d%H%M%S')}_{colis_labels_filename}"

        reception_pdf_data = {
            'reference': reception_ref,
            'ticket_id': ticket_id,
            'numero_dossier': numero_dossier,
            'receptionne_par': receptionne_par,
            'lieu_stockage': lieu_stockage,
            'date_reception': now.strftime("%d/%m/%Y %H:%M"),
            'nombre_colis': nombre_colis,
            'colis': colis_refs,
            'article_esi_ids': list(reception_esi_ids),
            'commentaire': commentaire,
            'items': selected,
        }
        reception_pdf_bytes = _build_reception_form_pdf_bytes(ticket, reception_pdf_data, source_type='avis')
        reception_pdf_filename = f"{reception_ref}_bon_reception.pdf"
        reception_pdf_path = f"{ticket_id}/receptions_avis/{now.strftime('%Y%m%d%H%M%S')}_{reception_pdf_filename}"

        try:
            supabase_upload_bytes(reception_pdf_path, reception_pdf_bytes, "application/pdf")
            supabase_upload_bytes(article_labels_path, article_labels_bytes, "application/pdf")
            supabase_upload_bytes(colis_labels_path, colis_labels_bytes, "application/pdf")
        except Exception as e:
            print(f"[RAR] Erreur upload étiquettes: {e}")
            return jsonify({'ok': False, 'error': f'Impossible d’enregistrer les PDF d’étiquettes : {e}'}), 500

        avis['items'] = items
        ticket['avisArrivee'] = avis

        reception = {
            'reference': reception_ref,
            'ticket_id': ticket_id,
            'receptionnee': True,
            'receptionnee_le': now.isoformat(),
            'date_reception': now.strftime("%d/%m/%Y %H:%M"),
            'receptionne_par': receptionne_par,
            'lieu_stockage': lieu_stockage,
            'numero_dossier': numero_dossier,
            'nombre_colis': nombre_colis,
            'colis': colis_refs,
            'commentaire': commentaire,
            'items': selected,
            'bon_reception_filename': reception_pdf_filename,
            'bon_reception_path': reception_pdf_path,
            'etiquettes_articles_filename': article_labels_filename,
            'etiquettes_articles_path': article_labels_path,
            'etiquettes_colis_filename': colis_labels_filename,
            'etiquettes_colis_path': colis_labels_path,
        }

        receptions = list(ticket.get('receptionsAvisArrivee') or [])
        receptions.append(reception)
        ticket['receptionsAvisArrivee'] = receptions

        # Compatibilité avec l'ancien champ de synthèse
        ticket['receptionAvisArrivee'] = reception

        manager_sheets = list(ticket.get('managerSheets') or [])
        for name, size, path in [
            (reception_pdf_filename, len(reception_pdf_bytes), reception_pdf_path),
            (article_labels_filename, len(article_labels_bytes), article_labels_path),
            (colis_labels_filename, len(colis_labels_bytes), colis_labels_path),
        ]:
            manager_sheets.append({'name': name, 'size': size, 'path': path})

        ticket['managerSheets'] = manager_sheets
        ticket['updatedAt'] = now.isoformat()

        save_ticket(ticket)
        try:
            _update_article_logistics(
                reception_esi_ids,
                lieu_stockage=lieu_stockage,
                statut_logistique="Réceptionné",
                colis_by_esi=colis_by_esi,
                reception_ref=reception_ref,
                receptionne_par=receptionne_par,
            )
        except Exception as e:
            print(f"[ARTICLES] Mise à jour logistique avis impossible: {e}")

    return jsonify({
        'ok': True,
        'ticket_id': ticket_id,
        'reference': reception_ref,
        'reception': reception,
        'colis': colis_refs,
        'bon_reception_filename': reception_pdf_filename,
        'etiquettes_articles_filename': article_labels_filename,
        'etiquettes_colis_filename': colis_labels_filename,
        'status': ticket.get('status')
    })


@app.route('/api/tickets/<ticket_id>/enlevement', methods=['PATCH'])
def api_update_enlevement(ticket_id):
    """Enregistre les corrections manuelles des champs d'une demande d'enlèvement."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404

    module_normalise = _as_text(ticket.get('module')).replace("’", "'").strip()
    if module_normalise not in ("Demande d'enlèvement", "Demande d'enlevement") and not _as_text(ticket_id).startswith('ENL-'):
        return jsonify({'ok': False, 'error': "Ce ticket n'est pas une demande d'enlèvement"}), 400

    data = request.get_json(silent=True) or {}
    current = dict(ticket.get('enlevement') or {})

    editable = [
        'client', 'numero_bon', 'date_enlevement',
        'coordinateur', 'exhibition',
        'adresse_depart',
        'adresse_destination',
        'notes', 'instructions'
    ]
    for field in editable:
        if field in data:
            current[field] = _as_text(data.get(field)).strip()

    if 'items' in data:
        if not isinstance(data.get('items'), list):
            return jsonify({'ok': False, 'error': 'Format des articles invalide'}), 400
        cleaned_items = []
        for item in data.get('items') or []:
            if not isinstance(item, dict):
                continue
            cleaned_items.append({
                'quantite': _as_text(item.get('quantite')).strip(),
                'designation': _as_text(item.get('designation')).strip(),
                'reference': _as_text(item.get('reference')).strip(),
                'dimensions': _as_text(item.get('dimensions')).strip(),
                'esi_ids': list(item.get('esi_ids') or []),
                'esi_id': _as_text(item.get('esi_id')).strip(),
            })
        current['items'] = cleaned_items
        current['references'] = [x['reference'] for x in cleaned_items if x.get('reference')]

    current['display_name'] = " - ".join(
        x for x in [current.get('client', '').strip(), current.get('numero_bon', '').strip()] if x
    )
    current['manually_edited'] = True
    current['manual_updated_at'] = datetime.now().isoformat()
    ticket['enlevement'] = current

    # Synchronisation avec les champs historiques utilisés ailleurs.
    ticket['dossier'] = current.get('client', '')
    ticket['ref'] = current.get('numero_bon', '')
    ticket['chargeProjet'] = current.get('coordinateur', '') or '-'
    ticket['expo'] = current.get('exhibition', '') or '-'
    ticket['objet'] = current.get('exhibition', '') or '-'
    # L'heure n'est pas utilisee dans la gestion reception.
    ticket['heureRdv'] = '-'

    date_fr = current.get('date_enlevement', '')
    if date_fr:
        try:
            ticket['dateRdv'] = datetime.strptime(date_fr, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return jsonify({'ok': False, 'error': 'La date doit être au format JJ/MM/AAAA'}), 400
    else:
        ticket['dateRdv'] = '-'

    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True, 'enlevement': current})




_BLR_LOCK = threading.Lock()


def _next_blr_reference():
    """Retourne une référence BLR-0001, BLR-0002... à partir des bons déjà enregistrés."""
    highest = 0
    try:
        for t in list_tickets():
            enl = t.get("enlevement") or {}
            for bon in list(enl.get("bons_livraison") or []) + list(enl.get("bons_livraison_annules") or []):
                ref = _as_text(bon.get("reference")).strip()
                m = re.fullmatch(r"BLR-(\d+)", ref, re.I)
                if m:
                    highest = max(highest, int(m.group(1)))
    except Exception as e:
        print(f"[BLR] Impossible de lire l'historique des BLR: {e}")
    return f"BLR-{highest + 1:04d}"


def _build_reception_form_pdf_bytes(ticket, bon, source_type="enlevement"):
    """Bon de reception A4 inspire de la trame ESI Avis d'arrivee fournie."""
    import io
    import textwrap as _tw

    PAGE_W, PAGE_H = 595, 842
    NAVY = (0.035, 0.145, 0.235)
    CYAN = (0.05, 0.55, 0.76)
    PALE = (0.965, 0.975, 0.982)
    LINE = (0.62, 0.70, 0.76)
    WHITE = (1, 1, 1)
    TEXT = (0.06, 0.10, 0.16)

    def clean(v, default='-'):
        s = _as_text(v).replace('\r', ' ').replace('\n', ' ').strip()
        return s or default

    def esc(v):
        return clean(v, '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

    def rgb(c, stroke=False):
        return f"{c[0]:.3f} {c[1]:.3f} {c[2]:.3f} {'RG' if stroke else 'rg'}"

    ops = []
    def rect(x, y, w, h, fill=None, stroke=None, lw=0.6):
        if fill is not None:
            ops.append(rgb(fill))
        if stroke is not None:
            ops.append(rgb(stroke, True))
        ops.append(f"{lw:.2f} w")
        mode = 'B' if fill is not None and stroke is not None else ('f' if fill is not None else 'S')
        ops.append(f"{x:.1f} {y:.1f} {w:.1f} {h:.1f} re {mode}")

    def line(x1, y1, x2, y2, color=LINE, lw=0.6):
        ops.append(rgb(color, True)); ops.append(f"{lw:.2f} w")
        ops.append(f"{x1:.1f} {y1:.1f} m {x2:.1f} {y2:.1f} l S")

    def txt(x, y, value, size=8, bold=False, color=TEXT):
        value = esc(value)
        ops.append(rgb(color))
        ops.append('BT')
        ops.append(f"/{'F2' if bold else 'F1'} {size:.1f} Tf")
        ops.append(f"{x:.1f} {y:.1f} Td")
        ops.append(f"({value}) Tj")
        ops.append('ET')

    def fit_txt(x, y, value, width, size=8, bold=False, max_lines=2, leading=10, color=TEXT):
        approx = max(8, int(width / max(size * 0.62, 1)))
        wrapped = _tw.wrap(clean(value, ''), width=approx) or ['']
        for n, part in enumerate(wrapped[:max_lines]):
            txt(x, y - n * leading, part, size=size, bold=bold, color=color)

    avis = ticket.get('avisArrivee') or ticket.get('avis_arrivee') or {}
    enl = ticket.get('enlevement') or {}
    if source_type == 'avis':
        dossier = clean(bon.get('numero_dossier') or avis.get('dossier_ref') or ticket.get('dossier'))
        client = clean(avis.get('client') or ticket.get('dossier'))
        projet = clean(avis.get('projet') or ticket.get('expo') or ticket.get('objet'))
        coordinateur = clean(avis.get('coordinateur') or ticket.get('chargeProjet'))
        exp = avis.get('expediteur') or {}
        tr = avis.get('transporteur') or {}
        left_block = [clean(exp.get('nom')), clean(exp.get('adresse')), clean(exp.get('contact'))]
        right_block = [clean(tr.get('nom')), clean(tr.get('adresse')), clean(tr.get('contact')), clean(tr.get('reference'))]
    else:
        dossier = clean(bon.get('numero_dossier'))
        client = clean(enl.get('client') or ticket.get('dossier'))
        projet = clean(enl.get('exhibition') or ticket.get('expo') or ticket.get('objet'))
        coordinateur = clean(enl.get('coordinateur') or ticket.get('chargeProjet'))
        left_block = [clean(enl.get('adresse_depart')), '-', '-']
        right_block = [clean(enl.get('adresse_destination')), '-', '-', clean(enl.get('numero_bon') or ticket.get('ref'))]

    # Header avec le vrai logo ESI depuis static/logo.png.
    logo_path = APP_DIR / 'static' / 'logo.png'
    logo_image_bytes = None
    logo_w = logo_h = None
    if logo_path.exists():
        try:
            from PIL import Image
            with Image.open(logo_path) as im:
                if im.mode not in ('RGB', 'L'):
                    im = im.convert('RGB')
                logo_w, logo_h = im.size
                buf = io.BytesIO()
                im.save(buf, format='JPEG', quality=92)
                logo_image_bytes = buf.getvalue()
        except Exception as e:
            print(f'[PDF] Logo ESI non charge: {e}')

    bon_reference = clean(bon.get('reference'))
    txt(168, 786, f'BON DE RECEPTION - N° {bon_reference}', 18, True, TEXT)
    txt(168, 767, 'Controle et enregistrement de la marchandise', 10, False, (0.35,0.40,0.45))
    line(30, 744, 565, 744, CYAN, 1.6)

    txt(30, 726, 'INFORMATIONS DOSSIER', 10, True, CYAN)
    rect(30, 695, 535, 24, fill=NAVY)
    txt(40, 704, 'DOSSIER', 9, True, WHITE)

    # Informations dossier : référence, client, projet, chargé de projet et date de réception.
    cols = [
        (30,100,'Ref. dossier',dossier),
        (130,110,'Nom du client',client),
        (240,110,'Projet ou expo',projet),
        (350,110,'Chargé de projet',coordinateur),
        (460,105,'Date de reception',bon.get('date_reception')),
    ]
    for x,w,label,value in cols:
        rect(x, 652, w, 43, fill=PALE, stroke=LINE)
        txt(x+5, 680, label, 6.5, True, (0.25,0.32,0.38))
        fit_txt(x+5, 663, value, w-10, size=8, bold=True, max_lines=2, leading=9)

    # Side blocks, visually matching the source template.
    rect(30, 623, 260, 22, fill=NAVY); txt(38, 631, 'EXPEDITEUR / DEPART', 8, True, WHITE)
    rect(305, 623, 260, 22, fill=NAVY); txt(313, 631, 'TRANSPORTEUR / DESTINATION', 8, True, WHITE)
    left_labels = ['Nom / adresse','Adresse','Contact']
    right_labels = ['Nom / adresse','Adresse','Contact','Reference']
    y = 603
    for i, label in enumerate(left_labels):
        rect(30, y-20*i, 62, 20, fill=PALE, stroke=LINE); txt(35, y+7-20*i, label, 6.2, True)
        rect(92, y-20*i, 198, 20, stroke=LINE); fit_txt(97, y+7-20*i, left_block[i], 188, 6.5, False, 1, 8)
    for i, label in enumerate(right_labels):
        rect(305, y-20*i, 62, 20, fill=PALE, stroke=LINE); txt(310, y+7-20*i, label, 6.2, True)
        rect(367, y-20*i, 198, 20, stroke=LINE); fit_txt(372, y+7-20*i, right_block[i], 188, 6.5, False, 1, 8)

    # Goods table.
    table_top = 517
    widths = [58, 68, 145, 82, 64, 48, 70]
    headers = ['N° ESI','Ref. item','Description de la marchandise','N° COLIS','Dimensions','Qte recue','Stockage']
    x = 30
    for w, h in zip(widths, headers):
        rect(x, table_top-38, w, 38, fill=NAVY, stroke=WHITE, lw=0.4)
        fit_txt(x+4, table_top-17, h, w-8, 6.4, True, 2, 8, WHITE)
        x += w

    rows = bon.get('items') or []
    row_h = 29
    max_rows = 8
    for ridx in range(max_rows):
        y0 = table_top - 38 - (ridx+1)*row_h
        x = 30
        fill = (0.99, 0.985, 0.955)
        row = rows[ridx] if ridx < len(rows) else {}
        received = row.get('quantite') or ''
        esi_ids = [str(v).strip() for v in (row.get('esi_ids') or []) if str(v).strip()]
        colis_par_esi = row.get('colis_par_esi') if isinstance(row.get('colis_par_esi'), dict) else {}
        numeros_colis = [
            _as_text(colis_par_esi.get(esi_id)).strip()
            for esi_id in esi_ids
            if _as_text(colis_par_esi.get(esi_id)).strip()
        ]
        vals = [
            ', '.join(esi_ids),
            row.get('reference'),
            row.get('designation') or row.get('description'),
            ', '.join(numeros_colis),
            row.get('dimensions'),
            received,
            bon.get('lieu_stockage') if row else ''
        ]
        for w, value in zip(widths, vals):
            rect(x, y0, w, row_h, fill=fill, stroke=LINE)
            fit_txt(x+4, y0+18, value, w-8, 6.3, False, 2, 8)
            x += w

    table_bottom = table_top - 38 - max_rows*row_h
    rect(30, table_bottom-22, 535, 22, fill=PALE, stroke=LINE)
    txt(350, table_bottom-14, 'TOTAL ARTICLES RECEPTIONNES', 7, True)
    total_received = sum(int(float(str(i.get('quantite') or 0).replace(',','.'))) for i in rows if str(i.get('quantite') or '').strip())
    txt(515, table_bottom-14, str(total_received), 8, True)

    # Comment and reception details.
    comment_y = table_bottom - 74
    rect(30, comment_y, 535, 46, stroke=LINE)
    rect(30, comment_y+30, 535, 16, fill=PALE, stroke=LINE)
    txt(38, comment_y+35, 'Commentaire :', 7.2, True)
    fit_txt(38, comment_y+19, bon.get('commentaire') or '-', 515, 7, False, 2, 9)

    sig_y = 72
    txt(305, sig_y+78, "Reception / controle a l'arrivee :", 9, True)
    txt(305, sig_y+57, 'Date :', 8); txt(350, sig_y+57, clean(bon.get('date_reception')), 8)
    txt(305, sig_y+36, 'Nom :', 8); txt(350, sig_y+36, clean(bon.get('receptionne_par')), 8)
    txt(305, sig_y+15, 'Signature : __________________________', 8)

    txt(30, 35, f"Reference : {clean(bon.get('reference'))}  |  Colis : {clean(', '.join(bon.get('colis') or []))}", 7, False, (0.35,0.40,0.45))
    txt(430, 35, 'Groupe ESI - Bon de reception', 7, False, (0.35,0.40,0.45))

    if logo_image_bytes and logo_w and logo_h:
        box_x, box_y, box_w, box_h = 30, 752, 118, 62
        scale = min(box_w / float(logo_w), box_h / float(logo_h))
        draw_w = logo_w * scale
        draw_h = logo_h * scale
        draw_x = box_x + (box_w - draw_w) / 2
        draw_y = box_y + (box_h - draw_h) / 2
        ops.extend([
            'q',
            f'{draw_w:.2f} 0 0 {draw_h:.2f} {draw_x:.2f} {draw_y:.2f} cm',
            '/Im1 Do',
            'Q',
        ])
    else:
        rect(30, 752, 118, 62, fill=NAVY)
        txt(50, 779, 'ESI', 24, True, WHITE)

    stream = '\n'.join(ops).encode('latin-1', errors='replace')
    if logo_image_bytes and logo_w and logo_h:
        image_obj = (
            f'<< /Type /XObject /Subtype /Image /Width {logo_w} /Height {logo_h} '
            f'/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length {len(logo_image_bytes)} >>\nstream\n'
        ).encode('latin-1') + logo_image_bytes + b'\nendstream'
        objects = [
            b'<< /Type /Catalog /Pages 2 0 R >>',
            b'<< /Type /Pages /Kids [5 0 R] /Count 1 >>',
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>',
            f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_W} {PAGE_H}] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> /XObject << /Im1 7 0 R >> >> /Contents 6 0 R >>'.encode('latin-1'),
            f'<< /Length {len(stream)} >>\nstream\n'.encode('latin-1') + stream + b'\nendstream',
            image_obj,
        ]
    else:
        objects = [
            b'<< /Type /Catalog /Pages 2 0 R >>',
            b'<< /Type /Pages /Kids [5 0 R] /Count 1 >>',
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>',
            f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_W} {PAGE_H}] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents 6 0 R >>'.encode('latin-1'),
            f'<< /Length {len(stream)} >>\nstream\n'.encode('latin-1') + stream + b'\nendstream',
        ]
    out = io.BytesIO(); out.write(b'%PDF-1.4\n')
    offsets=[]
    for i,obj in enumerate(objects,1):
        offsets.append(out.tell()); out.write(f'{i} 0 obj\n'.encode('latin-1')); out.write(obj); out.write(b'\nendobj\n')
    xref=out.tell(); out.write(f'xref\n0 {len(objects)+1}\n'.encode('latin-1')); out.write(b'0000000000 65535 f \n')
    for off in offsets: out.write(f'{off:010d} 00000 n \n'.encode('latin-1'))
    out.write(f'trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF'.encode('latin-1'))
    return out.getvalue()


def _build_blr_pdf_bytes(ticket, bon):
    return _build_reception_form_pdf_bytes(ticket, bon, source_type='enlevement')


def _build_labels_pdf_bytes(labels, kind="article"):
    """Génère un PDF d'étiquettes, une étiquette par page.

    - Étiquettes COLIS : format exact 100 x 148 mm avec le vrai logo ESI
      chargé depuis static/logo.png et affiché en haut à gauche sans déformation.
    - Étiquettes ARTICLE : format historique inchangé (A6), sans logo ajouté.
    """
    import io
    import textwrap as _tw

    # ------------------------------------------------------------------
    # Étiquettes ARTICLE : comportement historique strictement inchangé.
    # ------------------------------------------------------------------
    if kind != "colis":
        page_width, page_height = 298, 420
        margin = 24

        def pdf_escape(value):
            value = _as_text(value)
            return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            None,
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
        ]
        page_refs = []

        for label in labels or [{"titre": "ETIQUETTE"}]:
            content_obj_num = len(objects) + 1
            lines = []

            title = _as_text(label.get("titre") or "ARTICLE").strip()
            lines.append(("B", 20, title))

            principal = _as_text(label.get("principal")).strip()
            if principal:
                for part in _tw.wrap(principal, width=28) or [principal]:
                    lines.append(("B", 18, part))

            for key in ("esi_id", "dossier", "client", "reference", "designation", "quantite", "colis", "lieu", "bon"):
                value = _as_text(label.get(key)).strip()
                if not value:
                    continue
                label_name = {
                    "esi_id": "N° ESI",
                    "dossier": "Dossier",
                    "client": "Client",
                    "reference": "Article",
                    "designation": "Designation",
                    "quantite": "Quantite",
                    "colis": "Colis",
                    "lieu": "Stockage",
                    "bon": "Bon",
                }.get(key, key)
                text_line = f"{label_name} : {value}"
                for part in _tw.wrap(text_line, width=42) or [text_line]:
                    lines.append(("R", 11, part))

            stream_lines = []
            y = page_height - margin - 20
            for font_kind, size, line in lines:
                font = "F2" if font_kind == "B" else "F1"
                stream_lines += [
                    "BT",
                    f"/{font} {size} Tf",
                    f"{margin} {y} Td",
                    f"({pdf_escape(line)}) Tj",
                    "ET",
                ]
                y -= max(size + 8, 18)

            stream = "\n".join(stream_lines).encode("latin-1", errors="replace")
            objects.append(
                f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1")
                + stream + b"\nendstream"
            )
            page_obj_num = len(objects) + 1
            page = (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
                f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_obj_num} 0 R >>"
            )
            objects.append(page.encode("latin-1"))
            page_refs.append(f"{page_obj_num} 0 R")

        objects[1] = (
            f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>"
        ).encode("latin-1")

        pdf = io.BytesIO()
        pdf.write(b"%PDF-1.4\n")
        offsets = []
        for i, obj in enumerate(objects, start=1):
            offsets.append(pdf.tell())
            pdf.write(f"{i} 0 obj\n".encode("latin-1"))
            pdf.write(obj)
            pdf.write(b"\nendobj\n")

        xref_pos = pdf.tell()
        pdf.write(f"xref\n0 {len(objects)+1}\n".encode("latin-1"))
        pdf.write(b"0000000000 65535 f \n")
        for offset in offsets:
            pdf.write(f"{offset:010d} 00000 n \n".encode("latin-1"))
        pdf.write(
            (
                f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\n"
                f"startxref\n{xref_pos}\n%%EOF"
            ).encode("latin-1")
        )
        return pdf.getvalue()

    # ------------------------------------------------------------------
    # Étiquettes COLIS : 100 x 148 mm + vrai logo ESI en haut à gauche.
    # ------------------------------------------------------------------
    page_width = 100 * 72 / 25.4
    page_height = 148 * 72 / 25.4
    margin = 16

    def pdf_escape(value):
        value = _as_text(value)
        return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    # Charge le vrai logo du projet. Il n'est jamais recadré : on conserve son ratio.
    logo_path = APP_DIR / 'static' / 'logo.png'
    logo_image_bytes = None
    logo_w = logo_h = None
    if logo_path.exists():
        try:
            from PIL import Image
            with Image.open(logo_path) as im:
                # Le PDF utilise un JPEG RGB afin d'embarquer l'image sans dépendance externe.
                if im.mode != 'RGB':
                    bg = Image.new('RGB', im.size, 'white')
                    if 'A' in im.getbands():
                        bg.paste(im, mask=im.getchannel('A'))
                    else:
                        bg.paste(im.convert('RGB'))
                    im = bg
                logo_w, logo_h = im.size
                buf = io.BytesIO()
                im.save(buf, format='JPEG', quality=95)
                logo_image_bytes = buf.getvalue()
        except Exception as e:
            print(f'[ETIQUETTE COLIS] Logo ESI non charge: {e}')

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        None,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    ]

    image_obj_num = None
    if logo_image_bytes and logo_w and logo_h:
        image_obj_num = len(objects) + 1
        image_obj = (
            f'<< /Type /XObject /Subtype /Image /Width {logo_w} /Height {logo_h} '
            f'/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length {len(logo_image_bytes)} >>\nstream\n'
        ).encode('latin-1') + logo_image_bytes + b'\nendstream'
        objects.append(image_obj)

    page_refs = []

    for label in labels or [{"titre": "COLIS"}]:
        stream_lines = []

        # Logo en haut à gauche, avec zone réservée 92 x 48 points.
        logo_box_x = margin
        logo_box_y = page_height - 62
        logo_box_w = 92
        logo_box_h = 42
        if image_obj_num:
            scale = min(logo_box_w / float(logo_w), logo_box_h / float(logo_h))
            draw_w = logo_w * scale
            draw_h = logo_h * scale
            draw_x = logo_box_x
            draw_y = logo_box_y + (logo_box_h - draw_h) / 2
            stream_lines += [
                'q',
                f'{draw_w:.2f} 0 0 {draw_h:.2f} {draw_x:.2f} {draw_y:.2f} cm',
                '/Im1 Do',
                'Q',
            ]
        else:
            # Secours uniquement si static/logo.png est absent sur le serveur.
            stream_lines += [
                'BT', '/F2 20 Tf', f'{margin} {page_height - 40:.2f} Td', '(ESI) Tj', 'ET'
            ]

        # Titre COLIS en haut à droite du logo.
        title = _as_text(label.get('titre') or 'COLIS').strip()
        stream_lines += [
            'BT', '/F2 17 Tf', f'{page_width - 72:.2f} {page_height - 36:.2f} Td',
            f'({pdf_escape(title)}) Tj', 'ET'
        ]

        # Le N° COLIS reste l'information principale en grand.
        # Le N° BON DE RECEPTION est affiché plus bas à la place de l'ancien champ Bon.
        principal = _as_text(label.get('colis') or label.get('principal')).strip()
        y = page_height - 88
        if principal:
            stream_lines += [
                'BT', '/F2 9 Tf', f'{margin} {y:.2f} Td',
                '(N\260 COLIS) Tj', 'ET'
            ]
            y -= 18
            principal_size = 23 if len(principal) <= 18 else 19
            for part in _tw.wrap(principal, width=22) or [principal]:
                stream_lines += [
                    'BT', f'/F2 {principal_size} Tf', f'{margin} {y:.2f} Td',
                    f'({pdf_escape(part)}) Tj', 'ET'
                ]
                y -= principal_size + 9

        # Trait de séparation.
        y -= 2
        stream_lines += [
            '0.25 w', f'{margin} {y:.2f} m {page_width - margin:.2f} {y:.2f} l S'
        ]
        y -= 22

        fields = [
            ('dossier', 'Dossier'),
            ('client', 'Client'),
            ('lieu', 'Stockage'),
            ('bon', 'N° Bon reception'),
        ]
        for key, field_label in fields:
            value = _as_text(label.get(key)).strip()
            if not value:
                continue
            stream_lines += [
                'BT', '/F2 9 Tf', f'{margin} {y:.2f} Td',
                f'({pdf_escape(field_label.upper())}) Tj', 'ET'
            ]
            y -= 14
            wrapped = _tw.wrap(value, width=34) or [value]
            for part in wrapped[:3]:
                stream_lines += [
                    'BT', '/F1 11 Tf', f'{margin} {y:.2f} Td',
                    f'({pdf_escape(part)}) Tj', 'ET'
                ]
                y -= 14
            y -= 8

        stream = "\n".join(stream_lines).encode("latin-1", errors="replace")
        content_obj_num = len(objects) + 1
        objects.append(
            f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1")
            + stream + b"\nendstream"
        )

        page_obj_num = len(objects) + 1
        xobject = f" /XObject << /Im1 {image_obj_num} 0 R >>" if image_obj_num else ""
        page = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width:.4f} {page_height:.4f}] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >>{xobject} >> "
            f"/Contents {content_obj_num} 0 R >>"
        )
        objects.append(page.encode('latin-1'))
        page_refs.append(f'{page_obj_num} 0 R')

    objects[1] = (
        f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>"
    ).encode('latin-1')

    pdf = io.BytesIO()
    pdf.write(b'%PDF-1.4\n')
    offsets = []
    for i, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f'{i} 0 obj\n'.encode('latin-1'))
        pdf.write(obj)
        pdf.write(b'\nendobj\n')

    xref_pos = pdf.tell()
    pdf.write(f'xref\n0 {len(objects)+1}\n'.encode('latin-1'))
    pdf.write(b'0000000000 65535 f \n')
    for offset in offsets:
        pdf.write(f'{offset:010d} 00000 n \n'.encode('latin-1'))
    pdf.write(
        (
            f'trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\n'
            f'startxref\n{xref_pos}\n%%EOF'
        ).encode('latin-1')
    )
    return pdf.getvalue()

def _existing_colis_numbers(numero_dossier):
    """Retourne tous les numéros de colis déjà utilisés, quel que soit le type de réception."""
    numero_dossier = _as_text(numero_dossier).strip()
    if not numero_dossier: return set()
    used=set(); pattern=re.compile(r"^"+re.escape(numero_dossier)+r"-(\d+)$", re.I)
    def add(ref):
        m=pattern.fullmatch(_as_text(ref).strip())
        if m: used.add(int(m.group(1)))
    try:
        for ticket in list_tickets():
            enl = ticket.get("enlevement") or {}
            for bon in list(enl.get("bons_livraison") or []) + list(enl.get("bons_livraison_annules") or []):
                for ref in bon.get("colis") or []: add(ref)
            for rec in list(ticket.get("receptionsAvisArrivee") or []) + list(ticket.get("receptionsAvisArriveeAnnulees") or []):
                for ref in rec.get("colis") or []: add(ref)
        safe_dossier=urllib.parse.quote(numero_dossier,safe='')
        rows=supabase_rest_request("GET","articles",f"select=dernier_colis&dossier=eq.{safe_dossier}&limit=10000") or []
        for row in rows:
            for ref in _as_text(row.get("dernier_colis")).split(','): add(ref)
    except Exception as e: print(f"[COLIS] Lecture historique impossible pour {numero_dossier}: {e}")
    return used




def _resolve_colis_repartition(selected_items, raw_assignments, colis_refs):
    expected={}
    for item in selected_items:
        idx=int(item.get('index'))
        for unit_offset, esi_id in enumerate(item.get('esi_ids') or []): expected[(idx,unit_offset)]=esi_id
    assignments={}
    for entry in raw_assignments or []:
        try: key=(int(entry.get('index')),int(entry.get('unit_offset'))); ci=int(entry.get('colis_index'))
        except Exception: continue
        if key not in expected or ci<0 or ci>=len(colis_refs): continue
        assignments[key]=ci
    if set(assignments)!=set(expected): raise ValueError("Chaque article physique doit être associé à un colis.")
    used=set(assignments.values())
    if len(colis_refs)>len(expected): raise ValueError("Le nombre de colis ne peut pas dépasser le nombre d'articles réceptionnés.")
    if used != set(range(len(colis_refs))): raise ValueError("Chaque colis créé doit contenir au moins un article.")
    return {expected[k]:colis_refs[ci] for k,ci in assignments.items()}

def _apply_colis_to_selected_items(selected_items, colis_by_esi, reception_ref, lieu_stockage):
    for item in selected_items:
        mapping={esi:colis_by_esi.get(esi,'') for esi in item.get('esi_ids') or []}
        item['colis_par_esi']=mapping
        item['colis']=list(dict.fromkeys(x for x in mapping.values() if x))
        item['reception_ref']=reception_ref; item['lieu_stockage']=lieu_stockage

def _allocate_colis_numbers(numero_dossier, count):
    """Génère des références dossier-001, dossier-002... jamais déjà utilisées."""
    numero_dossier = _as_text(numero_dossier).strip()
    count = int(count)
    if not numero_dossier or count < 1:
        return []

    used = _existing_colis_numbers(numero_dossier)
    refs = []
    n = 1
    while len(refs) < count:
        if n not in used:
            refs.append(f"{numero_dossier}-{n:03d}")
            used.add(n)
        n += 1
    return refs


@app.route('/api/tickets/<ticket_id>/bon-livraison', methods=['POST'])
def api_create_bon_livraison(ticket_id):
    """Valide une réception partielle/totale et génère BL + étiquettes articles + étiquettes colis."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404

    module_normalise = _as_text(ticket.get('module')).replace("’", "'").strip()
    if module_normalise not in ("Demande d'enlèvement", "Demande d'enlevement") and not _as_text(ticket_id).startswith('ENL-'):
        return jsonify({'ok': False, 'error': "Ce ticket n'est pas une demande d'enlèvement"}), 400

    data = request.get_json(silent=True) or {}
    receptionne_par = _as_text(data.get('receptionne_par')).strip()
    lieu_stockage = _as_text(data.get('lieu_stockage')).strip()
    numero_dossier = _as_text(data.get('numero_dossier')).strip()

    # Secours serveur : si l'interface n'envoie pas le N° dossier,
    # on reprend automatiquement la référence principale du bon/ticket.
    enl_for_dossier = dict(ticket.get('enlevement') or {})
    if not numero_dossier:
        numero_dossier = _as_text(
            enl_for_dossier.get('numero_dossier')
            or enl_for_dossier.get('dossier_numero')
            or enl_for_dossier.get('numero_bon')
            or ticket.get('numeroDossier')
            or ticket.get('numero_dossier')
            or ticket.get('ref')
        ).strip()

    try:
        nombre_colis = int(data.get('nombre_colis') or 0)
    except (TypeError, ValueError):
        nombre_colis = 0

    items_reception = data.get('items_reception')
    selected_indexes = data.get('selected_indexes') or []
    colis_repartition = data.get('colis_repartition') or []

    if not receptionne_par:
        return jsonify({'ok': False, 'error': 'Nom et prénom du réceptionnaire obligatoires'}), 400
    if not lieu_stockage:
        return jsonify({'ok': False, 'error': 'Lieu de stockage obligatoire'}), 400
    if not numero_dossier:
        return jsonify({'ok': False, 'error': 'N° dossier obligatoire pour numéroter les colis'}), 400
    if nombre_colis < 1:
        return jsonify({'ok': False, 'error': 'Le nombre total de colis doit être supérieur ou égal à 1'}), 400

    # Compatibilité avec l'ancienne interface : si elle n'envoie que selected_indexes,
    # on réceptionne le reliquat complet de chaque article sélectionné.
    if not isinstance(items_reception, list):
        items_reception = []
        for raw_idx in selected_indexes if isinstance(selected_indexes, list) else []:
            try:
                items_reception.append({'index': int(raw_idx), 'quantite_recue': None})
            except Exception:
                pass

    if not items_reception:
        return jsonify({'ok': False, 'error': 'Sélectionne au moins un article à réceptionner'}), 400

    # Sécurise les anciennes demandes créées avant l'ajout du référentiel articles.
    try:
        _ensure_articles_for_ticket(ticket, save=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f"Impossible d'attribuer les numéros ESI : {e}"}), 500

    enl = dict(ticket.get('enlevement') or {})
    all_items = list(enl.get('items') or [])
    selected = []
    article_labels = []
    reception_esi_ids = []
    now = datetime.now()

    def _qty_int(value, default=0):
        try:
            return max(0, int(float(str(value).replace(',', '.'))))
        except Exception:
            return default

    def _item_fully_received(item):
        planned = max(1, _qty_int(item.get('quantite') or 1, 1))
        received = _qty_int(item.get('quantite_recue_totale'), -1)
        if received < 0:
            received = planned if item.get('receptionne') else 0
        return received >= planned

    if all_items and all(_item_fully_received(dict(x or {})) for x in all_items):
        return jsonify({
            'ok': False,
            'error': 'Réception clôturée : tous les articles ont déjà été réceptionnés.'
        }), 409

    for entry in items_reception:
        if not isinstance(entry, dict):
            continue
        try:
            idx = int(entry.get('index'))
        except Exception:
            continue
        if idx < 0 or idx >= len(all_items):
            continue

        item = dict(all_items[idx] or {})
        planned = max(1, _qty_int(item.get('quantite') or 1, 1))

        # Compatibilité avec les anciennes réceptions qui ne stockaient qu'un booléen.
        previous = _qty_int(item.get('quantite_recue_totale'), -1)
        if previous < 0:
            previous = planned if item.get('receptionne') else 0

        remaining = max(planned - previous, 0)
        if remaining <= 0:
            continue

        requested = entry.get('quantite_recue')
        qty_received = remaining if requested in (None, '', 0, '0') else _qty_int(requested, 0)
        if qty_received < 1:
            continue
        if qty_received > remaining:
            return jsonify({
                'ok': False,
                'error': f"Quantité reçue trop élevée pour {item.get('reference') or idx} : reste {remaining}"
            }), 400

        new_total = previous + qty_received

        history = list(item.get('receptions') or [])
        history.append({
            'date': now.isoformat(),
            'quantite': qty_received,
            'receptionne_par': receptionne_par,
            'lieu_stockage': lieu_stockage,
        })

        item['quantite_recue_totale'] = new_total
        item['receptionne'] = new_total >= planned
        item['receptionne_le'] = now.isoformat()
        item['receptionne_par'] = receptionne_par
        item['lieu_stockage'] = lieu_stockage
        item['receptions'] = history
        all_items[idx] = item

        received_esi_ids = _article_ids_for_received_units(item, previous, qty_received)
        reception_esi_ids.extend(received_esi_ids)

        selected_item = {
            'index': idx,
            'quantite': str(qty_received),
            'quantite_prevue': str(planned),
            'quantite_deja_recue': str(previous),
            'quantite_recue_totale': str(new_total),
            'designation': _as_text(item.get('designation')).strip(),
            'reference': _as_text(item.get('reference')).strip(),
            'dimensions': _as_text(item.get('dimensions')).strip(),
            'esi_ids': received_esi_ids,
        }
        selected.append(selected_item)

        # Une étiquette par unité réellement réceptionnée, avec son ESI-x unique.
        for unit_no, esi_id in enumerate(received_esi_ids, start=1):
            article_labels.append({
                'titre': 'ARTICLE',
                'principal': esi_id,
                'esi_id': esi_id,
                'dossier': numero_dossier,
                'client': enl.get('client') or ticket.get('dossier') or '',
                'reference': _as_text(item.get('reference')).strip(),
                'designation': _as_text(item.get('designation')).strip(),
                'quantite': f"{unit_no}/{qty_received}",
                'lieu': lieu_stockage,
            })

    if not selected:
        return jsonify({'ok': False, 'error': 'Aucun article valide sélectionné'}), 400

    with _BLR_LOCK:
        blr_ref = _next_blr_reference()
        colis_refs = _allocate_colis_numbers(numero_dossier, nombre_colis)
        try:
            colis_by_esi = _resolve_colis_repartition(selected, colis_repartition, colis_refs)
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
        _apply_colis_to_selected_items(selected, colis_by_esi, blr_ref, lieu_stockage)
        for label in article_labels:
            label['colis'] = colis_by_esi.get(label.get('esi_id'), '')

        bon = {
            'reference': blr_ref,
            'ticket_id': ticket_id,
            'client': enl.get('client') or ticket.get('dossier') or '',
            'numero_bon_enlevement': enl.get('numero_bon') or ticket.get('ref') or '',
            'numero_dossier': numero_dossier,
            'receptionne_par': receptionne_par,
            'lieu_stockage': lieu_stockage,
            'date_reception': now.strftime("%d/%m/%Y %H:%M"),
            'created_at': now.isoformat(),
            'nombre_colis': nombre_colis,
            'colis': colis_refs,
            'article_esi_ids': list(reception_esi_ids),
            'items': selected,
        }

        # BL principal
        pdf_bytes = _build_blr_pdf_bytes(ticket, bon)
        filename = f"{blr_ref}.pdf"
        storage_path = f"{ticket_id}/bons_livraison/{now.strftime('%Y%m%d%H%M%S')}_{filename}"

        # Étiquettes articles
        article_labels_bytes = _build_labels_pdf_bytes(article_labels, kind="article")
        article_labels_filename = f"{blr_ref}_etiquettes_articles.pdf"
        article_labels_path = f"{ticket_id}/bons_livraison/{now.strftime('%Y%m%d%H%M%S')}_{article_labels_filename}"

        # Étiquettes colis
        colis_labels = [{
            'titre': 'COLIS',
            'principal': colis_ref,
            'dossier': numero_dossier,
            'client': enl.get('client') or ticket.get('dossier') or '',
            'colis': colis_ref,
            'lieu': lieu_stockage,
            'bon': blr_ref,
        } for colis_ref in colis_refs]
        colis_labels_bytes = _build_labels_pdf_bytes(colis_labels, kind="colis")
        colis_labels_filename = f"{blr_ref}_etiquettes_colis.pdf"
        colis_labels_path = f"{ticket_id}/bons_livraison/{now.strftime('%Y%m%d%H%M%S')}_{colis_labels_filename}"

        try:
            supabase_upload_bytes(storage_path, pdf_bytes, "application/pdf")
            supabase_upload_bytes(article_labels_path, article_labels_bytes, "application/pdf")
            supabase_upload_bytes(colis_labels_path, colis_labels_bytes, "application/pdf")
        except Exception as e:
            print(f"[BLR] Erreur upload PDF: {e}")
            return jsonify({'ok': False, 'error': f'Impossible d’enregistrer les PDF : {e}'}), 500

        bon['filename'] = filename
        bon['storage_path'] = storage_path
        bon['etiquettes_articles_filename'] = article_labels_filename
        bon['etiquettes_articles_path'] = article_labels_path
        bon['etiquettes_colis_filename'] = colis_labels_filename
        bon['etiquettes_colis_path'] = colis_labels_path

        enl['items'] = all_items
        enl['references'] = [
            _as_text(x.get('reference')).strip()
            for x in all_items
            if _as_text(x.get('reference')).strip()
        ]
        bons = list(enl.get('bons_livraison') or [])
        bons.append(bon)
        enl['bons_livraison'] = bons
        ticket['enlevement'] = enl

        manager_sheets = list(ticket.get('managerSheets') or [])
        for name, size, path in [
            (filename, len(pdf_bytes), storage_path),
            (article_labels_filename, len(article_labels_bytes), article_labels_path),
            (colis_labels_filename, len(colis_labels_bytes), colis_labels_path),
        ]:
            manager_sheets.append({'name': name, 'size': size, 'path': path})
        ticket['managerSheets'] = manager_sheets
        ticket['updatedAt'] = now.isoformat()
        save_ticket(ticket)
        try:
            _update_article_logistics(
                reception_esi_ids,
                lieu_stockage=lieu_stockage,
                statut_logistique="Réceptionné",
                colis_by_esi=colis_by_esi,
                reception_ref=blr_ref,
                receptionne_par=receptionne_par,
            )
        except Exception as e:
            print(f"[ARTICLES] Mise à jour logistique BLR impossible: {e}")

    return jsonify({
        'ok': True,
        'reference': blr_ref,
        'filename': filename,
        'bon': bon,
        'colis': colis_refs,
        'etiquettes_articles_filename': article_labels_filename,
        'etiquettes_colis_filename': colis_labels_filename,
    })


@app.route('/api/tickets/<ticket_id>/status', methods=['PATCH'])
def api_update_status(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    ancien_statut = ticket.get('status')

    data = request.get_json(silent=True) or {}
    nouveau_statut = data.get('status', ancien_statut)

    ticket['status'] = nouveau_statut
    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)

    # L'envoi automatique SMTP est volontairement désactivé.
    # La notification se prépare maintenant via Outlook Web avec le bouton "Envoyer Notif".
    return jsonify({'ok': True})


@app.route('/api/tickets/<ticket_id>/annuler-enlevement', methods=['PATCH'])
def api_annuler_enlevement(ticket_id):
    """Annule un ticket du planning réception uniquement s'il n'a plus de réception active."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404

    module_normalise = _as_text(ticket.get('module')).replace("’", "'").strip()
    is_enlevement = (
        module_normalise in ("Demande d'enlèvement", "Demande d'enlevement")
        or _as_text(ticket_id).startswith('ENL-')
    )
    is_avis = (
        module_normalise == "Avis d'arrivée"
        or _as_text(ticket_id).startswith('ARR-')
    )
    if not (is_enlevement or is_avis):
        return jsonify({'ok': False, 'error': "Ce ticket n'appartient pas au planning réception"}), 400

    if is_avis:
        active_receptions = [
            r for r in (ticket.get('receptionsAvisArrivee') or [])
            if not bool((r or {}).get('annulee'))
        ]
    else:
        enl = ticket.get('enlevement') or {}
        active_receptions = [
            r for r in (enl.get('bons_livraison') or [])
            if not bool((r or {}).get('annulee'))
        ]

    if active_receptions:
        refs = [
            _as_text((r or {}).get('reference')).strip()
            for r in active_receptions
            if _as_text((r or {}).get('reference')).strip()
        ]
        detail = ', '.join(refs) if refs else f"{len(active_receptions)} réception(s) active(s)"
        return jsonify({
            'ok': False,
            'error': f"Impossible d'annuler le ticket : annule d'abord la/les réception(s) active(s) ({detail}).",
            'active_receptions': len(active_receptions),
            'references': refs,
        }), 409

    ticket['status'] = 'Annulé'
    ticket['annule_le'] = datetime.now().isoformat()
    ticket['updatedAt'] = ticket['annule_le']
    save_ticket(ticket)

    checked = load_ticket(ticket_id)
    if not checked or _as_text(checked.get('status')).strip().lower() not in ('annulé', 'annule'):
        return jsonify({'ok': False, 'error': "Le ticket n'a pas pu être confirmé comme annulé après sauvegarde."}), 500

    return jsonify({'ok': True, 'status': 'Annulé'})


@app.route('/api/tickets/<ticket_id>/notification-url')
def api_ticket_notification_url(ticket_id):
    """Prépare une URL Outlook Web préremplie pour envoyer la notification manuellement."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    charge_projet = (ticket.get("chargeProjet") or "").strip()
    email_dest = _find_project_manager_email(charge_projet)

    if not email_dest:
        return jsonify({
            'error': "Aucun email trouvé pour le chargé de projet dans les référentiels."
        }), 404

    module = ticket.get("module", "")
    dossier = (ticket.get("dossier") or "").strip()
    ref = (ticket.get("ref") or "").strip()
    preteur = (ticket.get("preteur") or "").strip()
    projet = (ticket.get("expo") or ticket.get("objet") or "").strip()
    lieu_rdv = (ticket.get("lieuRdv") or "").strip()
    date_rdv = (ticket.get("dateRdv") or "").strip()
    heure_rdv = (ticket.get("heureRdv") or "").strip()
    commentaire = ticket.get("commentaire", "")
    fiche = ticket.get("fiche") or {}

    subject = _format_ticket_notification_subject(ticket)

    # Lien direct vers le ticket dans le portail demandeur, sans mot de passe.
    base_url = request.host_url.rstrip('/')
    ticket_url = f"{base_url}/demandeur?ticket={urllib.parse.quote(ticket_id, safe='')}"

    if module == "Fiche de caisse":
        link_text = f"Consulter la fiche de caisse {dossier}-{ref}"
        intro = "La fiche de caisse suivante a été commandée :"
        details = f"""
        <p>
          <strong>Dossier :</strong> {dossier or '-'}<br>
          <strong>N° caisse / Référence :</strong> {ref or '-'}<br>
          <strong>Prêteur :</strong> {preteur or '-'}<br>
          <strong>Dimensions extérieures :</strong> {fiche.get('dimensionsExt') or '-'}<br>
          <strong>Prix de cession :</strong> {fiche.get('prixCession') or '-'}<br>
          <strong>Date mise à dispo :</strong> {datetime.fromisoformat(ticket.get('dateEmballage')).strftime('%d/%m/%Y') if ticket.get('dateEmballage') and ticket.get('dateEmballage') != '-' else '-'}
        </p>
        """
    elif module == "Demande de devis":
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter la demande de devis {label}"
        intro = "La demande de devis suivante a été finalisée :"
        details = f"""
        <p>
          <strong>Client / Dossier :</strong> {dossier or '-'}<br>
          <strong>Projet :</strong> {projet or '-'}<br>
          <strong>Chargé de projet :</strong> {charge_projet or '-'}
        </p>
        <p><strong>Commentaire :</strong><br>{(commentaire or '-').replace(chr(10), '<br>')}</p>
        """
    elif module == "Demande Aller voir":
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter le dossier {label}"
        intro = "La demande Aller Voir suivante a été finalisée :"
        date_rdv_fr = datetime.fromisoformat(date_rdv).strftime('%d/%m/%Y') if date_rdv and date_rdv != '-' else '-'
        details = f"""
        <p>
          <strong>Client / Dossier :</strong> {dossier or '-'}<br>
          <strong>Projet :</strong> {projet or '-'}<br>
          <strong>Lieu de rendez-vous :</strong> {lieu_rdv or '-'}<br>
          <strong>Date :</strong> {date_rdv_fr}<br>
          <strong>Heure :</strong> {heure_rdv or '-'}<br>
          <strong>Prêteur :</strong> {preteur or '-'}
        </p>
        """
    else:
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter le ticket {label}"
        intro = "La demande suivante a été finalisée :"
        details = f"""
        <p>
          <strong>Client / Dossier :</strong> {dossier or '-'}<br>
          <strong>Projet :</strong> {projet or '-'}<br>
          <strong>Chargé de projet :</strong> {charge_projet or '-'}
        </p>
        """

    body_html = f"""<html>
<body>
<p>Bonjour,</p>
<p>{intro}</p>
{details}
<p>Les documents associés sont disponibles dans ESI Tickets.</p>
<p>
  <a href=\"{ticket_url}\" style=\"background:#0284c7;color:#ffffff;padding:10px 16px;text-decoration:none;border-radius:6px;display:inline-block;font-weight:bold;\">
    {link_text}
  </a>
</p>
</body>
</html>"""

    # Outlook Web accepte le paramètre body dans le deeplink compose.
    # Le contenu HTML permet d'afficher un lien avec un libellé propre au lieu d'une URL brute.
    params = urllib.parse.urlencode({
        "to": email_dest,
        "subject": subject,
        "body": body_html
    })

    outlook_url = "https://outlook.office.com/mail/deeplink/compose?" + params

    return jsonify({
        'ok': True,
        'to': email_dest,
        'subject': subject,
        'body': body_html,
        'outlook_url': outlook_url,
        'ticket_url': ticket_url,
        'link_text': link_text
    })


def _build_notification_content(ticket, ticket_id, notification_mode='final'):
    """Construit les éléments de notification en HTML pour un brouillon .eml Outlook."""
    charge_projet = (ticket.get("chargeProjet") or "").strip()
    email_dest = _find_project_manager_email(charge_projet)

    if not email_dest:
        return None, "Aucun email trouvé pour le chargé de projet dans les référentiels."

    module = ticket.get("module", "")
    dossier = (ticket.get("dossier") or "").strip()
    ref = (ticket.get("ref") or "").strip()
    preteur = (ticket.get("preteur") or "").strip()
    projet = (ticket.get("expo") or ticket.get("objet") or "").strip()
    lieu_rdv = (ticket.get("lieuRdv") or "").strip()
    date_rdv = (ticket.get("dateRdv") or "").strip()
    heure_rdv = (ticket.get("heureRdv") or "").strip()
    commentaire = ticket.get("commentaire", "")
    fiche = ticket.get("fiche") or {}

    subject = _format_ticket_notification_subject(ticket)
    base_url = request.host_url.rstrip('/')
    ticket_url = f"{base_url}/demandeur?ticket={urllib.parse.quote(ticket_id, safe='')}"

    def esc(value):
        return html.escape(str(value or "-"))

    def nl2br(value):
        return html.escape(str(value or "-")).replace("\n", "<br>")

    if module == "Fiche de caisse":
        link_text = f"Consulter la fiche de caisse {dossier}-{ref}"
        intro = "La fiche de caisse suivante a été commandée :"
        details = f"""
        <p>
          <strong>Dossier :</strong> {esc(dossier)}<br>
          <strong>N° caisse / Référence :</strong> {esc(ref)}<br>
          <strong>Prêteur :</strong> {esc(preteur)}<br>
          <strong>Dimensions extérieures :</strong> {esc(fiche.get('dimensionsExt'))}<br>
          <strong>Prix de cession :</strong> {esc(fiche.get('prixCession'))}<br>
          <strong>Date mise à dispo :</strong> {esc(datetime.fromisoformat(ticket.get('dateEmballage')).strftime('%d/%m/%Y') if ticket.get('dateEmballage') and ticket.get('dateEmballage') != '-' else '-')}
        </p>
        """
    elif module == "Demande de devis":
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter la demande de devis {label}"
        intro = "La demande de devis suivante a été finalisée :"
        details = f"""
        <p>
          <strong>Client / Dossier :</strong> {esc(dossier)}<br>
          <strong>Projet :</strong> {esc(projet)}<br>
          <strong>Chargé de projet :</strong> {esc(charge_projet)}
        </p>
        <p><strong>Commentaire :</strong><br>{nl2br(commentaire)}</p>
        """
    elif module == "Demande Aller voir":
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter le dossier {label}"
        date_rdv_fr = datetime.fromisoformat(date_rdv).strftime('%d/%m/%Y') if date_rdv and date_rdv != '-' else '-'

        if notification_mode == 'validation':
            subject = f"[ESI Tickets] Créneau Aller voir validé - {label}".strip()
            intro = "Le créneau suivant a été validé :"
            details = f"""
            <p>
              <strong>Client / Dossier :</strong> {esc(dossier)}<br>
              <strong>Projet :</strong> {esc(projet)}<br>
              <strong>Lieu de rendez-vous :</strong> {esc(lieu_rdv)}<br>
              <strong>Date :</strong> {esc(date_rdv_fr)}<br>
              <strong>Heure :</strong> {esc(heure_rdv)}<br>
              <strong>Prêteur :</strong> {esc(preteur)}
            </p>
            <p>Le rendez-vous est désormais confirmé.</p>
            """
        else:
            intro = "La demande Aller Voir suivante a été finalisée :"
            details = f"""
            <p>
              <strong>Client / Dossier :</strong> {esc(dossier)}<br>
              <strong>Projet :</strong> {esc(projet)}<br>
              <strong>Lieu de rendez-vous :</strong> {esc(lieu_rdv)}<br>
              <strong>Date :</strong> {esc(date_rdv_fr)}<br>
              <strong>Heure :</strong> {esc(heure_rdv)}<br>
              <strong>Prêteur :</strong> {esc(preteur)}
            </p>
            """
    else:
        label = " ".join([x for x in [dossier, projet] if x]).strip() or ticket_id
        link_text = f"Consulter le ticket {label}"
        intro = "La demande suivante a été finalisée :"
        details = f"""
        <p>
          <strong>Client / Dossier :</strong> {esc(dossier)}<br>
          <strong>Projet :</strong> {esc(projet)}<br>
          <strong>Chargé de projet :</strong> {esc(charge_projet)}
        </p>
        """

    documents_line = ""
    if not (module == "Demande Aller voir" and notification_mode == "validation"):
        documents_line = "<p>Les documents associés sont disponibles dans ESI Tickets.</p>"

    body_html = f"""<!doctype html>
<html>
<body style="font-family:Arial,Helvetica,sans-serif;font-size:14px;color:#0f172a;line-height:1.45;">
<p>Bonjour,</p>
<p>{html.escape(intro)}</p>
{details}
{documents_line}
<p>
  <a href="{html.escape(ticket_url, quote=True)}" style="background:#0284c7;color:#ffffff;padding:10px 16px;text-decoration:none;border-radius:6px;display:inline-block;font-weight:bold;">
    {html.escape(link_text)}
  </a>
</p>
</body>
</html>"""

    body_text = f"""Bonjour,

{intro}

Dossier : {dossier or '-'}
Référence : {ref or '-'}
Prêteur : {preteur or '-'}

{link_text}
{ticket_url}

"""

    return {
        "to": email_dest,
        "subject": subject,
        "body_html": body_html,
        "body_text": body_text,
        "ticket_url": ticket_url,
        "link_text": link_text,
    }, None


@app.route('/api/tickets/<ticket_id>/notification-eml')
def api_ticket_notification_eml(ticket_id):
    """Génère un brouillon .eml HTML à ouvrir dans Outlook Desktop."""
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    notification_mode = (request.args.get('mode') or 'final').strip()
    content, error = _build_notification_content(ticket, ticket_id, notification_mode)
    if error:
        return jsonify({'error': error}), 404

    msg = EmailMessage()
    msg['To'] = content['to']
    msg['Subject'] = content['subject']
    msg['Date'] = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid(domain='esi-tickets.local')
    # Indique à Outlook que le fichier doit s'ouvrir comme un message non envoyé.
    msg['X-Unsent'] = '1'
    msg.set_content(
        content['body_text'],
        charset='utf-8',
        cte='8bit'
    )
    msg.add_alternative(
        content['body_html'],
        subtype='html',
        charset='utf-8',
        cte='8bit'
    )

    filename_base = "notification_" + safe_filename(ticket_id)
    eml_bytes = msg.as_bytes()

    import io
    return send_file(
        io.BytesIO(eml_bytes),
        as_attachment=True,
        download_name=f"{filename_base}.eml",
        mimetype='message/rfc822'
    )


@app.route('/api/tickets/<ticket_id>/manager-sheet', methods=['POST'])
def api_manager_sheet(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    files = request.files.getlist('files')
    if not files:
        single = request.files.get('file')
        if single:
            files = [single]

    valid_files = [fs for fs in files if fs and fs.filename]
    if not valid_files:
        return jsonify({'error': 'Fichier manquant'}), 400

    ticket_folder(ticket_id)  # conserve la création du dossier local historique
    manager_sheets = list(ticket.get('managerSheets') or [])
    legacy = ticket.get('managerSheet')
    if legacy and isinstance(legacy, dict) and legacy.get('name'):
        if not any(x.get('name') == legacy.get('name') for x in manager_sheets):
            manager_sheets.append(legacy)

    for fs in valid_files:
        content = fs.read()
        clean_name = safe_filename(fs.filename)
        storage_path = f"{ticket_id}/gestionnaire/{datetime.now().strftime('%Y%m%d%H%M%S')}_{clean_name}"

        try:
            supabase_upload_bytes(
                storage_path,
                content,
                fs.content_type
            )
        except Exception as e:
            print(f"[SUPABASE UPLOAD GESTIONNAIRE] Erreur : {e}")
            return jsonify({'ok': False, 'error': f'Erreur upload Supabase : {e}'}), 500

        manager_sheets = [x for x in manager_sheets if x.get('name') != fs.filename]
        manager_sheets.append({
            'name': fs.filename,
            'size': len(content),
            'path': storage_path
        })
    ticket['managerSheets'] = manager_sheets
    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

def _find_file_info(ticket, filename, kind):
    items = ticket.get('managerSheets') if kind == 'gestionnaire' else ticket.get('files')
    for f in items or []:
        if f.get('name') == filename:
            return f
    return None


def _redirect_to_signed_file(ticket_id, filename, kind):
    ticket = load_ticket(ticket_id)
    if not ticket:
        abort(404)

    file_info = _find_file_info(ticket, filename, kind)
    if not file_info:
        abort(404)

    storage_path = file_info.get('path')
    if not storage_path:
        abort(404)

    try:
        signed_url = supabase_signed_download_url(storage_path, expires_in=300)
    except Exception as e:
        # Secours : si l'URL signée échoue, on garde l'ancien comportement via Render.
        print(f"[SUPABASE SIGNED DOWNLOAD] Erreur, fallback Render : {e}")
        import io
        try:
            data = supabase_download_bytes(storage_path)
        except Exception as e2:
            print(f"[SUPABASE DOWNLOAD] Erreur : {e2}")
            abort(404)
        return send_file(io.BytesIO(data), as_attachment=True, download_name=filename)

    return redirect(signed_url)


@app.route('/api/tickets/<ticket_id>/download/<filename>')
def api_download_file(ticket_id, filename):
    return _redirect_to_signed_file(ticket_id, filename, 'demandeur')


@app.route('/api/tickets/<ticket_id>/download-sheet/<filename>')
def api_download_sheet(ticket_id, filename):
    return _redirect_to_signed_file(ticket_id, filename, 'gestionnaire')


@app.route('/api/tickets/<ticket_id>/fiche', methods=['GET'])
def api_get_fiche(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    return jsonify(ticket.get('fiche', {}))

@app.route('/api/tickets/<ticket_id>/fiche', methods=['POST'])
def api_save_fiche(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    data = request.get_json(silent=True) or {}
    longueur = data.get('longueur', '')
    largeur = data.get('largeur', '')
    hauteur = data.get('hauteur', '')
    dimensions_ext = " x ".join([v for v in [longueur, largeur, hauteur] if str(v).strip()])
    ancienne_fiche = ticket.get('fiche') or {}
    ticket['fiche'] = {
        'longueur': longueur,
        'largeur': largeur,
        'hauteur': hauteur,
        'dimensionsExt': dimensions_ext,
        'prixAchat': data.get('prixAchat', ''),
        'prixCession': data.get('prixCession', ''),
        'typeCaisseFiche': data.get('typeCaisseFiche', ''),
        'bilanCarbone': data.get('bilanCarbone', ''),
        'poids': data.get('poids', ''),
        'choixCaissier': data.get('choixCaissier', ''),
        # La localisation est renseignée depuis le planning réception.
        # On la conserve si la fiche est modifiée depuis l'écran gestionnaire.
        'localisation': ancienne_fiche.get('localisation', '')
    }
    save_ticket(ticket)
    return jsonify({'ok': True})



def _linked_articles_for_ticket(ticket):
    """Retourne les articles lies a une fiche de caisse avec dossier/reference a jour."""
    stored = ticket.get('articles_lies') or []
    esi_ids = []
    for item in stored:
        esi_id = _as_text(item.get('esi_id') if isinstance(item, dict) else item).strip()
        if esi_id and esi_id not in esi_ids:
            esi_ids.append(esi_id)

    if not esi_ids:
        return []

    rows_by_id = {}
    for offset in range(0, len(esi_ids), 100):
        part = esi_ids[offset:offset + 100]
        encoded = urllib.parse.quote(','.join(part), safe=',-_')
        rows = supabase_rest_request(
            'GET', 'articles',
            'select=esi_id,dossier,reference,type_objet&esi_id=in.(' + encoded + ')&limit=100'
        ) or []
        for row in rows:
            esi_id = _as_text(row.get('esi_id')).strip()
            type_objet = _as_text(row.get('type_objet') or 'PRODUIT').strip().upper()
            if not esi_id or type_objet == 'CONTENANT':
                continue
            rows_by_id[esi_id] = {
                'esi_id': esi_id,
                'dossier': _as_text(row.get('dossier')).strip(),
                'reference': _as_text(row.get('reference')).strip(),
            }

    return [rows_by_id[x] for x in esi_ids if x in rows_by_id]


@app.route('/api/tickets/<ticket_id>/articles-lies', methods=['GET', 'PUT'])
def api_ticket_articles_lies(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Ticket introuvable'}), 404
    if _as_text(ticket.get('module')).strip() != 'Fiche de caisse':
        return jsonify({'ok': False, 'error': "Ce ticket n'est pas une fiche de caisse"}), 400

    if request.method == 'GET':
        try:
            articles = _linked_articles_for_ticket(ticket)
            return jsonify({'ok': True, 'articles': articles, 'count': len(articles)})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e), 'articles': []}), 500

    data = request.get_json(silent=True) or {}
    raw_ids = data.get('esi_ids') or []
    if not isinstance(raw_ids, list):
        return jsonify({'ok': False, 'error': 'Liste des articles invalide'}), 400

    esi_ids = []
    for value in raw_ids:
        esi_id = _as_text(value).strip()
        if esi_id and esi_id not in esi_ids:
            esi_ids.append(esi_id)

    # Référence canonique de la caisse : N° dossier + N° caisse.
    # Exemple : dossier 101129, caisse 1 -> 101129-01.
    dossier_caisse = _as_text(ticket.get('dossier')).strip()
    numero_brut = _as_text(ticket.get('ref')).strip()
    numero_norm = _normalise_numero_caisse(numero_brut) if numero_brut else ''
    if not dossier_caisse or not numero_norm:
        return jsonify({
            'ok': False,
            'error': "Impossible de déterminer la référence de caisse (N° dossier ou N° caisse manquant)."
        }), 400

    numero_caisse = numero_norm.zfill(2) if numero_norm.isdigit() else numero_brut
    caisse_ref = f"{dossier_caisse}-{numero_caisse}"
    equivalent_refs = {caisse_ref, f"{dossier_caisse}-{numero_norm}"}

    # Mémorise l'ancienne sélection pour savoir quels articles ont été retirés.
    previous_ids = []
    for item in ticket.get('articles_lies') or []:
        previous_id = _as_text(item.get('esi_id') if isinstance(item, dict) else item).strip()
        if previous_id and previous_id not in previous_ids:
            previous_ids.append(previous_id)

    all_ids = list(dict.fromkeys(previous_ids + esi_ids))

    try:
        rows_by_id = {}
        for offset in range(0, len(all_ids), 100):
            part = all_ids[offset:offset + 100]
            if not part:
                continue
            encoded = urllib.parse.quote(','.join(part), safe=',-_')
            rows = supabase_rest_request(
                'GET', 'articles',
                'select=*&esi_id=in.(' + encoded + ')&limit=100'
            ) or []
            for row in rows:
                row_esi = _as_text(row.get('esi_id')).strip()
                if row_esi:
                    rows_by_id[row_esi] = dict(row)

        selected = []
        missing = []
        conflicts = []

        for esi_id in esi_ids:
            row = rows_by_id.get(esi_id)
            if not row:
                missing.append(esi_id)
                continue

            type_objet = _as_text(row.get('type_objet') or 'PRODUIT').strip().upper()
            if type_objet == 'CONTENANT':
                missing.append(esi_id)
                continue

            current_ref = _as_text(row.get('ref_caisse')).strip()
            if current_ref and current_ref not in equivalent_refs:
                conflicts.append({
                    'esi_id': esi_id,
                    'ref_caisse': current_ref,
                })
                continue

            selected.append({
                'esi_id': esi_id,
                'dossier': _as_text(row.get('dossier')).strip(),
                'reference': _as_text(row.get('reference')).strip(),
            })

        if missing:
            return jsonify({
                'ok': False,
                'error': 'Certains articles sont introuvables ou sont des CONTENANTS : ' + ', '.join(missing[:10])
            }), 400

        if conflicts:
            details = ', '.join(
                f"{x['esi_id']} ({x['ref_caisse']})" for x in conflicts[:10]
            )
            return jsonify({
                'ok': False,
                'error': "Certains articles sont déjà liés à une autre caisse : " + details
            }), 409

        now = datetime.now().isoformat()
        changed_articles = []

        with _ARTICLE_LOCK:
            try:
                # 1) Met à jour la table ARTICLES.
                for esi_id in all_ids:
                    row = rows_by_id.get(esi_id)
                    if not row:
                        continue

                    current_ref = _as_text(row.get('ref_caisse')).strip()
                    current_colis = _as_text(row.get('dernier_colis')).strip()

                    if esi_id in esi_ids:
                        # Article coché : le N° de caisse EST aussi le N° de colis.
                        target_ref = caisse_ref
                        target_colis = caisse_ref
                    elif esi_id in previous_ids:
                        # Article décoché : retire uniquement les valeurs qui correspondent
                        # à CETTE caisse, sans effacer une autre affectation éventuelle.
                        target_ref = '' if current_ref in equivalent_refs else current_ref
                        target_colis = '' if current_colis in equivalent_refs else current_colis
                    else:
                        continue

                    if current_ref == target_ref and current_colis == target_colis:
                        continue

                    patch = {
                        'ref_caisse': target_ref,
                        'dernier_colis': target_colis,
                        'updated_at': now,
                    }
                    merged = dict(row)
                    merged.update(patch)
                    patch['search_text'] = _article_search_text(merged)

                    supabase_rest_request(
                        'PATCH', 'articles',
                        'esi_id=eq.' + urllib.parse.quote(esi_id, safe='-'),
                        patch,
                        prefer='return=minimal'
                    )
                    changed_articles.append({
                        'esi_id': esi_id,
                        'ref_caisse': current_ref,
                        'dernier_colis': current_colis,
                        'updated_at': row.get('updated_at'),
                        'search_text': row.get('search_text'),
                    })

                # 2) Conserve aussi la sélection dans le ticket comme auparavant.
                ticket['articles_lies'] = selected
                ticket['updatedAt'] = now
                save_ticket(ticket)

            except Exception:
                # Si la sauvegarde échoue, remet au mieux les articles dans leur état précédent.
                for old in reversed(changed_articles):
                    try:
                        rollback_patch = {
                            'ref_caisse': old.get('ref_caisse') or '',
                            'dernier_colis': old.get('dernier_colis') or '',
                            'updated_at': old.get('updated_at') or now,
                            'search_text': old.get('search_text') or '',
                        }
                        supabase_rest_request(
                            'PATCH', 'articles',
                            'esi_id=eq.' + urllib.parse.quote(old['esi_id'], safe='-'),
                            rollback_patch,
                            prefer='return=minimal'
                        )
                    except Exception as rollback_error:
                        print(
                            f"[ARTICLES LIES] Rollback impossible pour {old.get('esi_id')}: "
                            f"{rollback_error}"
                        )
                raise

        return jsonify({
            'ok': True,
            'articles': selected,
            'count': len(selected),
            'ref_caisse': caisse_ref,
            'numero_colis': caisse_ref,
        })

    except Exception as e:
        print(f"[ARTICLES LIES] Erreur synchro caisse {ticket_id}: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/export/excel')
def api_export_excel():
    try:
        from openpyxl import Workbook
        import io
        import re
    except Exception:
        return jsonify({'error': "openpyxl non installé"}), 500

    tickets = list_tickets()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append([
        "ID","Module","Statut","Date création","Dossier / Client",
        "Réf / N° caisse","Chargé de projet","Projet / Expo",
        "Type de caisse","Dimensions","Prix devis",
        "Prix d'achat","Prix cession","Commentaire","Choix du caissier",
        "Date RDV","Heure RDV","Lieu RDV"
    ])

    def parse_euro(value):
        if value is None:
            return None
        txt = str(value).strip()
        if not txt or txt == '-':
            return None
        txt = txt.replace('\xa0', ' ').replace('€', '').replace(' ', '')
        txt = txt.replace(',', '.')
        txt = re.sub(r'[^0-9.\-]', '', txt)
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None

    for t in tickets:
        fiche = t.get('fiche', {}) or {}
        ws.append([
            t.get('id',''),
            t.get('module',''),
            t.get('status',''),
            t.get('createdAt',''),
            t.get('dossier',''),
            t.get('ref',''),
            t.get('chargeProjet',''),
            t.get('expo') or t.get('objet',''),
            t.get('typeCaisse',''),
            t.get('dimensions',''),
            parse_euro(t.get('prixDevis','')),
            parse_euro(fiche.get('prixAchat','')),
            parse_euro(fiche.get('prixCession','')),
            t.get('commentaire',''),
            fiche.get('choixCaissier',''),
            t.get('dateRdv',''),
            t.get('heureRdv',''),
            t.get('lieuRdv','')
        ])

    for row in range(2, ws.max_row + 1):
        for col in [11, 12, 13]:
            ws.cell(row=row, column=col).number_format = '#,##0.00 €'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="tickets_esi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route('/api/tickets/<ticket_id>/export-pdf')
def api_export_ticket_pdf(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    import io
    import textwrap

    def clean(value):
        if value is None or value == '':
            return '-'
        return str(value).replace('\r', ' ').replace('\n', ' ')

    def add_wrapped(lines, label, value):
        text = label + " : " + clean(value)
        for part in textwrap.wrap(text, width=82) or [text]:
            lines.append(part)

    lines = []
    lines.append("ESI TICKETS - DETAIL TICKET")
    lines.append("=" * 70)
    lines.append("")
    add_wrapped(lines, "ID", ticket.get('id'))
    add_wrapped(lines, "Module", ticket.get('module'))
    add_wrapped(lines, "Statut", ticket.get('status'))
    add_wrapped(lines, "Dossier / Client", ticket.get('dossier'))
    add_wrapped(lines, "Reference", ticket.get('ref'))
    add_wrapped(lines, "Charge de projet", ticket.get('chargeProjet'))
    add_wrapped(lines, "Projet / Expo", ticket.get('expo') or ticket.get('objet'))
    add_wrapped(lines, "Preteur", ticket.get('preteur'))
    add_wrapped(lines, "Type de caisse", ticket.get('typeCaisse'))
    add_wrapped(lines, "Dimensions", ticket.get('dimensions'))
    add_wrapped(lines, "Prix devis", ticket.get('prixDevis'))
    add_wrapped(lines, "Lieu RDV", ticket.get('lieuRdv'))
    add_wrapped(lines, "Date RDV", ticket.get('dateRdv'))
    add_wrapped(lines, "Heure RDV", ticket.get('heureRdv'))

    lines.append("")
    lines.append("COMMENTAIRE / INFORMATIONS")
    lines.append("-" * 70)
    commentaire = clean(ticket.get('commentaire'))
    for part in textwrap.wrap(commentaire, width=82) or ['-']:
        lines.append(part)

    fiche = ticket.get('fiche') or {}
    if fiche:
        lines.append("")
        lines.append("INFORMATIONS FICHE")
        lines.append("-" * 70)
        add_wrapped(lines, "Dimensions exterieures", fiche.get('dimensionsExt'))
        add_wrapped(lines, "Prix achat", fiche.get('prixAchat'))
        add_wrapped(lines, "Type caisse fiche", fiche.get('typeCaisseFiche'))
        add_wrapped(lines, "Bilan carbone", fiche.get('bilanCarbone'))
        add_wrapped(lines, "Poids", fiche.get('poids'))
        add_wrapped(lines, "Choix caissier", fiche.get('choixCaissier'))

    lines.append("")
    lines.append("DOCUMENTS DU DEMANDEUR")
    lines.append("-" * 70)
    files = ticket.get('files') or []
    if files:
        for f in files:
            lines.append("- " + clean(f.get('name')))
    else:
        lines.append("- Aucun document")

    manager_sheets = ticket.get('managerSheets') or []
    if manager_sheets:
        lines.append("")
        lines.append("DOCUMENTS GESTIONNAIRE")
        lines.append("-" * 70)
        for f in manager_sheets:
            lines.append("- " + clean(f.get('name')))

    lines.append("")
    lines.append("NOTES / ACTIONS A PREVOIR")
    lines.append("-" * 70)
    lines.append("")
    lines.append("_" * 70)
    lines.append("")
    lines.append("_" * 70)
    lines.append("")
    lines.append("_" * 70)

    def pdf_escape(value):
        value = str(value)
        value = value.replace("\\", "\\\\")
        value = value.replace("(", "\\(")
        value = value.replace(")", "\\)")
        return value

    page_width, page_height = 595, 842
    margin_left = 42
    y_start = 800
    line_height = 14
    max_lines = 53
    chunks = [lines[i:i+max_lines] for i in range(0, len(lines), max_lines)] or [["Ticket vide"]]

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        None,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    ]

    page_refs = []
    for chunk in chunks:
        content_obj_num = len(objects) + 1
        content_lines = ["BT", "/F1 10 Tf", f"{margin_left} {y_start} Td"]
        first = True
        for line in chunk:
            if not first:
                content_lines.append(f"0 -{line_height} Td")
            first = False
            content_lines.append(f"({pdf_escape(line)}) Tj")
        content_lines.append("ET")

        stream = "\n".join(content_lines).encode("latin-1", errors="replace")
        objects.append(f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") + stream + b"\nendstream")

        page_obj_num = len(objects) + 1
        page = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_obj_num} 0 R >>"
        )
        objects.append(page.encode("latin-1"))
        page_refs.append(f"{page_obj_num} 0 R")

    objects[1] = f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>".encode("latin-1")

    pdf = io.BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = []

    for i, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f"{i} 0 obj\n".encode("latin-1"))
        pdf.write(obj)
        pdf.write(b"\nendobj\n")

    xref_pos = pdf.tell()
    pdf.write(f"xref\n0 {len(objects)+1}\n".encode("latin-1"))
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets:
        pdf.write(f"{offset:010d} 00000 n \n".encode("latin-1"))

    trailer = f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF"
    pdf.write(trailer.encode("latin-1"))
    pdf.seek(0)

    return send_file(
        pdf,
        as_attachment=True,
        download_name=f"{ticket.get('id','ticket')}.pdf",
        mimetype='application/pdf'
    )




@app.route('/api/restart')
def api_restart():
    import os
    os._exit(0)

@app.route('/splash')
def splash():
    return """
    <html>
    <head>
        <title>ESI Tickets</title>
        <style>
            body{margin:0;display:flex;justify-content:center;align-items:center;height:100vh;background:linear-gradient(180deg,#eef6fb,#f6f8fb);font-family:Arial;}
            .box{text-align:center;}
            img{width:120px;margin-bottom:20px;}
            h1{margin:0;color:#0284c7;}
            p{color:#64748b;}
        </style>
        <script>
            setTimeout(()=>{window.location.href="/demandeur";},1500);
        </script>
    </head>
    <body>
        <div class="box">
            <img id="splashLogo" src="/static/logo.jpg" onerror="
                const logos=['/static/logo.png','/static/logo%20esi.jpg'];
                const idx=Number(this.dataset.idx||0);
                if(idx<logos.length){this.dataset.idx=idx+1;this.src=logos[idx];}
                else{this.style.display='none';}
            ">
            <h1>ESI Tickets</h1>
            <p>Chargement en cours...</p>
        </div>
    </body>
    </html>
    """


@app.route('/api/tickets/<ticket_id>/validate-aller-voir', methods=['POST'])
def api_validate_aller_voir(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    ticket['status'] = 'En cours'
    ticket['validatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

@app.route('/api/tickets/<ticket_id>/calendar.ics')
def api_ticket_calendar_ics(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    date_rdv = ticket.get('dateRdv')
    heure_rdv = ticket.get('heureRdv')
    if not date_rdv or not heure_rdv or date_rdv == '-' or heure_rdv == '-':
        return jsonify({'error': 'Date/heure manquante'}), 400

    from datetime import timedelta
    start = datetime.fromisoformat(f"{date_rdv}T{heure_rdv}:00")
    end = start + timedelta(hours=2)

    def fmt(dt):
        return dt.strftime('%Y%m%dT%H%M%S')

    lieu = (ticket.get('lieuRdv') or '').strip()
    dossier = (ticket.get('dossier') or '').strip()
    summary = f"Aller voir - {lieu} - {dossier}"

    ics = f"""BEGIN:VCALENDAR
VERSION:2.0
BEGIN:VEVENT
DTSTART:{fmt(start)}
DTEND:{fmt(end)}
SUMMARY:{summary}
END:VEVENT
END:VCALENDAR"""

    from flask import Response
    return Response(ics, mimetype='text/calendar')



def open_browser():
    webbrowser.open('http://127.0.0.1:5050/splash')

ensure_shared_root()
init_db()

if __name__ == '__main__':
      app.run(host='127.0.0.1', port=5050, debug=False)
